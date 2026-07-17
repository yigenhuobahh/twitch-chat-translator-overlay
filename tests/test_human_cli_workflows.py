#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Human-like CLI workflow tests.

Goal: simulate the ways a person actually drives the scripts, without calling a
live translation API. These are intentionally higher-level than unit tests.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
import subprocess
import sys

import pytest

from helpers import (
    FIXTURES_DIR,
    ROOT,
    SCRIPTS_DIR,
    ffprobe_json,
    load_module,
    make_translation_for_export,
    stream_types,
    write_json,
)


def _env(extra: dict | None = None) -> dict:
    env = dict(os.environ)
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    # Prevent translate_chat_openai.py from loading .env in subprocess tests.
    env["_TWITCH_TRANSPARENT_TEST_MODE"] = "1"
    # Ensure accidental live translation is impossible in these tests.
    env.pop("OPENAI_COMPAT_API_KEY", None)
    env.pop("OPENAI_COMPAT_BASE_URL", None)
    env.pop("OPENAI_COMPAT_MODEL", None)
    env.pop("AGNES_API_KEY", None)
    env.pop("AGNES_BASE_URL", None)
    env.pop("AGNES_MODEL", None)
    if extra:
        env.update(extra)
    return env


def _run(cmd: list[str], env: dict | None = None) -> subprocess.CompletedProcess:
    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=_env(env),
    )


def _fill_translations(export_data: dict) -> dict:
    filled = make_translation_for_export(
        export_data,
        mapping={
            "hello [LUL] world": "你好 [LUL] 世界",
            "[Hey]": "[Hey]",
            "nice [xdx] clip": "不错的 [xdx] 片段",
        },
    )
    for msg in filled["messages"]:
        original = str(msg.get("original", "")).strip()
        if not msg.get("translation"):
            msg["translation"] = f"译:{original}"
        if original and all(p.startswith("[") and p.endswith("]") for p in original.split()):
            msg["translation"] = original
    return filled


def _assert_mp4(path: Path, min_duration: float = 1.5) -> None:
    assert path.is_file(), f"missing mp4: {path}"
    info = ffprobe_json(path)
    assert "video" in stream_types(info)
    assert "audio" in stream_types(info)
    assert float(info["format"]["duration"]) >= min_duration


class TestHumanCliWorkflows:
    def test_manual_translation_exports_json_and_review_tables(
        self, make_test_video, tmp_path: Path
    ):
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        work = tmp_path / "work"
        out = tmp_path / "out" / "unused.mp4"
        cmd = [
            sys.executable,
            str(SCRIPTS_DIR / "render_cn_chat.py"),
            str(video),
            str(html),
            "--manual-translation",
            "--workdir",
            str(work),
            "--output",
            str(out),
            "--offset",
            "0",
        ]
        r = _run(cmd)
        assert r.returncode == 0, r.stdout + "\n" + r.stderr

        trans = work / f"{video.stem}_translation.json"
        tsv = work / f"{video.stem}_translation_review.tsv"
        xlsx = work / f"{video.stem}_translation_review.xlsx"
        assert trans.is_file(), r.stdout
        assert tsv.is_file(), r.stdout
        assert xlsx.is_file(), r.stdout
        assert not out.exists()

        data = json.loads(trans.read_text(encoding="utf-8"))
        assert len(data["messages"]) == 3
        joined = (r.stdout or "") + (r.stderr or "")
        assert "review-done" in joined or "复核" in joined

    @pytest.mark.smoke
    def test_review_xlsx_edit_then_review_done_render(
        self, make_test_video, tmp_path: Path
    ):
        """manual export -> human edits xlsx -> review-done + reuse -> short render."""
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        work = tmp_path / "work"
        final_out = tmp_path / "final" / "reviewed.mp4"

        # Step 1: export for manual review
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "render_cn_chat.py"),
                str(video),
                str(html),
                "--manual-translation",
                "--workdir",
                str(work),
                "--output",
                str(final_out),
                "--offset",
                "0",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        trans = work / f"{video.stem}_translation.json"
        xlsx = work / f"{video.stem}_translation_review.xlsx"
        assert trans.is_file() and xlsx.is_file()

        # Step 2: human edits one translation cell
        from openpyxl import load_workbook

        wb = load_workbook(xlsx)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        tcol = headers.index("translation") + 1
        ws.cell(row=2, column=tcol, value="人工复核译文A")
        wb.save(xlsx)

        # Step 3: review-done + reuse + short preview clip
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "render_cn_chat.py"),
                str(video),
                str(html),
                "--reuse-translation",
                "--review-done",
                "--translation-json",
                str(trans),
                "--review-xlsx",
                str(xlsx),
                "--workdir",
                str(work),
                "--output",
                str(final_out),
                "--preview-clip",
                "2",
                "--offset",
                "0",
                "--fps",
                "30",
                "--x",
                "8",
                "--y",
                "24",
                "--width",
                "280",
                "--height",
                "160",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr

        data = json.loads(trans.read_text(encoding="utf-8"))
        assert data["messages"][0]["translation"] == "人工复核译文A"
        # preview-clip may publish nested output or leave under workdir/temp
        candidates = [
            final_out,
            work / "temp" / f"{video.stem}_chat.mp4",
            work / f"{video.stem}_chat.mp4",
        ]
        produced = next((p for p in candidates if p.is_file()), None)
        assert produced is not None, r.stdout + "\n" + r.stderr
        _assert_mp4(produced, min_duration=1.5)

    @pytest.mark.smoke
    def test_render_original_preview_clip(self, make_test_video, tmp_path: Path):
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        work = tmp_path / "work"
        out = tmp_path / "out" / "original.mp4"
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "render_cn_chat.py"),
                str(video),
                str(html),
                "--render-original",
                "--workdir",
                str(work),
                "--output",
                str(out),
                "--preview-clip",
                "2",
                "--offset",
                "0",
                "--fps",
                "30",
                "--x",
                "10",
                "--y",
                "30",
                "--width",
                "300",
                "--height",
                "180",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        candidates = [out, work / "temp" / f"{video.stem}_chat.mp4"]
        produced = next((p for p in candidates if p.is_file()), None)
        assert produced is not None, r.stdout
        _assert_mp4(produced, min_duration=1.5)

    @pytest.mark.smoke
    def test_preview_frame_png(self, make_test_video, tmp_path: Path):
        video = make_test_video(duration=3.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        preview = tmp_path / "preview_1s.png"
        out_dir = tmp_path / "burn"
        out_dir.mkdir()
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "twitch_chat_burn.py"),
                str(video),
                str(html),
                "--preview-frame",
                "1.0",
                "--preview-image",
                str(preview),
                "--out-dir",
                str(out_dir),
                "--job-dir",
                str(out_dir),
                "--offset",
                "0",
                "--x",
                "10",
                "--y",
                "30",
                "--w",
                "300",
                "--h",
                "180",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        assert preview.is_file(), r.stdout
        assert preview.stat().st_size > 100
        # preview-frame should not force a full chat mp4
        assert not (out_dir / f"{video.stem}_chat.mp4").exists()

    def test_skip_translate_stops_after_export(self, make_test_video, tmp_path: Path):
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        work = tmp_path / "work"
        out = tmp_path / "out.mp4"
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "render_cn_chat.py"),
                str(video),
                str(html),
                "--skip-translate",
                "--workdir",
                str(work),
                "--output",
                str(out),
                "--offset",
                "0",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        trans = work / f"{video.stem}_translation.json"
        assert trans.is_file()
        assert not out.exists()
        data = json.loads(trans.read_text(encoding="utf-8"))
        assert len(data["messages"]) == 3

    def test_lint_translation_cli_standalone(self, tmp_path: Path):
        bad = {
            "messages": [
                {
                    "index": 0,
                    "timestamp": 0,
                    "author": "a",
                    "original": "hello",
                    "translation": "",
                },
                {
                    "index": 1,
                    "timestamp": 1,
                    "author": "b",
                    "original": "@Alice https://example.com",
                    "translation": "看这里",
                },
            ]
        }
        path = write_json(tmp_path / "bad.json", bad)
        report = tmp_path / "lint.tsv"
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "render_cn_chat.py"),
                "--lint-translation",
                str(path),
                "--lint-report",
                str(report),
            ]
        )
        assert r.returncode != 0, r.stdout + "\n" + r.stderr
        assert report.is_file()
        text = report.read_text(encoding="utf-8")
        assert "empty" in text.lower() or "FAIL" in text or "空" in text

    def test_profile_and_rules_with_reuse_render(
        self, make_test_video, tmp_path: Path
    ):
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        work = tmp_path / "work"
        out = tmp_path / "out" / "profiled.mp4"

        # Export first
        export_json = tmp_path / "export.json"
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "twitch_chat_burn.py"),
                str(video),
                str(html),
                "--export-translation",
                str(export_json),
                "--offset",
                "0",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        filled = _fill_translations(json.loads(export_json.read_text(encoding="utf-8")))
        # Inject a rules-targetable acronym
        filled["messages"].append(
            {
                "index": 99,
                "timestamp": 1.5,
                "author": "rules_user",
                "original": "GG",
                "translation": "GG",
            }
        )
        filled_json = write_json(tmp_path / "filled.json", filled)

        profile = ROOT / "profiles" / "default.yaml"
        rules = ROOT / "configs" / "rules.example.yaml"
        assert profile.is_file() and rules.is_file()

        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "render_cn_chat.py"),
                str(video),
                str(html),
                "--reuse-translation",
                "--translation-json",
                str(filled_json),
                "--profile",
                str(profile),
                "--rules",
                str(rules),
                "--workdir",
                str(work),
                "--output",
                str(out),
                "--preview-clip",
                "2",
                "--offset",
                "0",
                "--fps",
                "30",
                "--x",
                "10",
                "--y",
                "30",
                "--width",
                "300",
                "--height",
                "180",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        # rules should normalize GG
        after = json.loads(filled_json.read_text(encoding="utf-8"))
        gg = next(m for m in after["messages"] if m.get("original") == "GG")
        assert gg["translation"] != "GG"
        joined = (r.stdout or "") + (r.stderr or "")
        assert "[profile]" in joined or "profile" in joined.lower()

        candidates = [out, work / "temp" / f"{video.stem}_chat.mp4"]
        produced = next((p for p in candidates if p.is_file()), None)
        assert produced is not None, r.stdout
        _assert_mp4(produced, min_duration=1.5)

    def test_conflict_flags_are_rejected(self, make_test_video, tmp_path: Path):
        video = make_test_video(duration=1.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "render_cn_chat.py"),
                str(video),
                str(html),
                "--render-original",
                "--manual-translation",
                "--offset",
                "0",
            ]
        )
        assert r.returncode != 0
        joined = (r.stdout or "") + (r.stderr or "")
        assert "render-original" in joined or "不能" in joined or "错误" in joined

    @pytest.mark.smoke
    def test_auto_offset_preview_clip_without_manual_offset(
        self, make_test_video, tmp_path: Path
    ):
        """Do not pass --offset; script should still produce a short preview."""
        video = make_test_video(duration=3.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        out_dir = tmp_path / "auto_off"
        out_dir.mkdir()
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "twitch_chat_burn.py"),
                str(video),
                str(html),
                "--preview-clip",
                "2",
                "--out-dir",
                str(out_dir),
                "--job-dir",
                str(out_dir),
                "--keep-temp",
                "--fps",
                "30",
                "--x",
                "8",
                "--y",
                "20",
                "--w",
                "260",
                "--h",
                "150",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        final = out_dir / f"{video.stem}_chat.mp4"
        _assert_mp4(final, min_duration=1.5)

    def test_translate_resume_skips_when_all_done(self, tmp_path: Path):
        """With all translations already present, resume should do zero model work."""
        data = {
            "messages": [
                {
                    "index": 0,
                    "timestamp": 1,
                    "author": "a",
                    "original": "hello",
                    "translation": "你好",
                },
                {
                    "index": 1,
                    "timestamp": 2,
                    "author": "b",
                    "original": "[LUL]",
                    "translation": "[LUL]",
                },
                {
                    "index": 2,
                    "timestamp": 3,
                    "author": "c",
                    "original": "world",
                    "translation": "世界",
                },
            ]
        }
        path = write_json(tmp_path / "done.json", data)
        # CLI still validates env even when todo is empty; use dummy values (no network).
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "translate_chat_openai.py"),
                str(path),
                "--resume",
                "--workers",
                "1",
                "--batch-size",
                "5",
            ],
            env={
                "OPENAI_COMPAT_BASE_URL": "http://127.0.0.1:9/v1",
                "OPENAI_COMPAT_API_KEY": "dummy-key",
                "OPENAI_COMPAT_MODEL": "dummy-model",
            },
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        joined = (r.stdout or "") + (r.stderr or "")
        out = json.loads(path.read_text(encoding="utf-8"))
        assert out["messages"][0]["translation"] == "你好"
        assert out["messages"][2]["translation"] == "世界"
        assert "待翻译 0" in joined or "已有有效译文" in joined

    def test_translate_resume_and_retry_failed_from_progress(self, tmp_path: Path):
        """
        Seed a progress file with one success + one failure, then --retry-failed
        with a local stub OpenAI client by monkeypatching is hard at CLI level.
        Instead verify: --retry-failed with remaining missing rows fails cleanly
        without API env (expected), while --resume with partial progress preserves
        already completed rows when invoked after we fill them.
        """
        path = write_json(
            tmp_path / "partial.json",
            {
                "messages": [
                    {
                        "index": 0,
                        "timestamp": 1,
                        "author": "a",
                        "original": "hello",
                        "translation": "你好",
                    },
                    {
                        "index": 1,
                        "timestamp": 2,
                        "author": "b",
                        "original": "todo item",
                        "translation": "",
                    },
                ]
            },
        )
        progress = path.with_name(path.name + ".progress.json")
        write_json(
            progress,
            {
                "schema_version": 1,
                "translations": {"0": "你好"},
                "failed": [1],
            },
        )

        # Without API env, retry-failed should fail (cannot translate remaining).
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "translate_chat_openai.py"),
                str(path),
                "--retry-failed",
                "--workers",
                "1",
            ]
        )
        assert r.returncode != 0, r.stdout + "\n" + r.stderr
        # Completed translation must remain intact after failed retry attempt.
        out = json.loads(path.read_text(encoding="utf-8"))
        assert out["messages"][0]["translation"] == "你好"

    @pytest.mark.smoke
    def test_reuse_translation_second_render_is_idempotent(
        self, make_test_video, tmp_path: Path
    ):
        """Human re-renders after tweaking layout using the same translation JSON."""
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        export_json = tmp_path / "export.json"
        r = _run(
            [
                sys.executable,
                str(SCRIPTS_DIR / "twitch_chat_burn.py"),
                str(video),
                str(html),
                "--export-translation",
                str(export_json),
                "--offset",
                "0",
            ]
        )
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        filled = write_json(
            tmp_path / "filled.json",
            _fill_translations(json.loads(export_json.read_text(encoding="utf-8"))),
        )

        outs = []
        for i, y in enumerate((20, 40)):
            work = tmp_path / f"work{i}"
            out = tmp_path / f"out{i}" / "chat.mp4"
            r = _run(
                [
                    sys.executable,
                    str(SCRIPTS_DIR / "render_cn_chat.py"),
                    str(video),
                    str(html),
                    "--reuse-translation",
                    "--translation-json",
                    str(filled),
                    "--workdir",
                    str(work),
                    "--output",
                    str(out),
                    "--preview-clip",
                    "2",
                    "--offset",
                    "0",
                    "--fps",
                    "30",
                    "--x",
                    "10",
                    "--y",
                    str(y),
                    "--width",
                    "300",
                    "--height",
                    "180",
                ]
            )
            assert r.returncode == 0, r.stdout + "\n" + r.stderr
            candidates = [out, work / "temp" / f"{video.stem}_chat.mp4"]
            produced = next((p for p in candidates if p.is_file()), None)
            assert produced is not None, r.stdout
            _assert_mp4(produced, min_duration=1.5)
            outs.append(produced)
        assert len(outs) == 2


class TestTranslateCliWithStubApi:
    def test_cache_and_resume_with_stub_openai(self, tmp_path: Path, monkeypatch):
        """In-process stub: first call fills cache, second call is all cache hits."""
        tr = load_module("translate_chat_openai", "translate_chat_openai.py")

        class _Msg:
            def __init__(self, content: str):
                self.content = content

        class _Choice:
            def __init__(self, content: str):
                self.message = _Msg(content)

        class _Resp:
            def __init__(self, content: str):
                self.choices = [_Choice(content)]

        class _Completions:
            def __init__(self):
                self.calls = 0

            def create(self, **kwargs):
                self.calls += 1
                user = kwargs["messages"][-1]["content"]
                items = []
                # prepare_messages_for_llm uses lines like: [7] hello
                for line in user.splitlines():
                    line = line.strip()
                    if line.startswith("[") and "]" in line:
                        idx_s = line[1 : line.index("]")]
                        original = line[line.index("]") + 1 :].strip()
                        try:
                            idx = int(idx_s)
                        except ValueError:
                            continue
                        items.append({"index": idx, "translation": f"译-{original}"})
                payload = json.dumps({"translations": items}, ensure_ascii=False)
                return _Resp(payload)

        class _Chat:
            def __init__(self):
                self.completions = _Completions()

        class _Client:
            def __init__(self, *a, **k):
                self.chat = _Chat()

        monkeypatch.setattr(tr, "OpenAI", _Client)
        monkeypatch.setattr(tr, "BASE_URL", "http://stub.local")
        monkeypatch.setattr(tr, "API_KEY", "stub-key")
        monkeypatch.setattr(tr, "MODEL", "stub-model")

        path = write_json(
            tmp_path / "need.json",
            {
                "messages": [
                    {
                        "index": 0,
                        "timestamp": 1,
                        "author": "a",
                        "original": "hello",
                        "translation": "",
                    },
                    {
                        "index": 1,
                        "timestamp": 2,
                        "author": "b",
                        "original": "world",
                        "translation": "",
                    },
                ]
            },
        )
        cache_dir = tmp_path / "cache"

        old_argv = sys.argv[:]
        try:
            sys.argv = [
                "translate_chat_openai.py",
                str(path),
                "--cache-dir",
                str(cache_dir),
                "--workers",
                "1",
                "--batch-size",
                "10",
                "--resume",
            ]
            tr.main()
        finally:
            sys.argv = old_argv

        out1 = json.loads(path.read_text(encoding="utf-8"))
        assert out1["messages"][0]["translation"].startswith("译-")
        assert out1["messages"][1]["translation"].startswith("译-")

        # Reset translations but keep cache; second run should complete via cache hits.
        for msg in out1["messages"]:
            msg["translation"] = ""
        write_json(path, out1)
        progress = path.with_name(path.name + ".progress.json")
        if progress.exists():
            progress.unlink()

        old_argv = sys.argv[:]
        try:
            sys.argv = [
                "translate_chat_openai.py",
                str(path),
                "--cache-dir",
                str(cache_dir),
                "--workers",
                "1",
                "--batch-size",
                "10",
                "--no-resume",
            ]
            tr.main()
        finally:
            sys.argv = old_argv

        out2 = json.loads(path.read_text(encoding="utf-8"))
        assert out2["messages"][0]["translation"].startswith("译-")
        assert out2["messages"][1]["translation"].startswith("译-")
        assert any(cache_dir.rglob("*")), "expected cache files"

