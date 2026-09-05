#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Regression: export wipe guard, stream timestamps, stale run_meta, progress fp."""

from __future__ import annotations

import json
import os
from pathlib import Path
import sys
import time

import pytest

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))


def test_write_export_refuses_nonempty_without_force(tmp_path: Path):
    import twitch_chat_burn as burn

    path = tmp_path / "t.json"
    path.write_text(
        json.dumps(
            {
                "messages": [
                    {"index": 0, "author": "a", "original": "hi", "translation": "你好"},
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    chat = {
        "messages": [
            {
                "author": "a",
                "timestamp": 1.0,
                "stream_timestamp": 100.0,
                "fragments": [{"type": "text", "text": ": hi"}],
            }
        ]
    }
    with pytest.raises(FileExistsError, match="非空 translation"):
        burn.write_export_translation_json(path, chat, force=False)
    # force overwrites
    payload = burn.write_export_translation_json(path, chat, force=True)
    data = json.loads(path.read_text(encoding="utf-8"))
    assert data["messages"][0]["translation"] == ""
    assert data["time_base"] == "stream"
    assert data["messages"][0]["timestamp"] == 100.0
    assert payload["schema_version"] == 2


def test_export_uses_stream_timestamp_not_video_relative():
    import twitch_chat_burn as burn

    chat = {
        "messages": [
            {
                "author": "a",
                "timestamp": 5.0,  # already video-relative after offset
                "stream_timestamp": 3605.0,
                "fragments": [{"type": "text", "text": ": hi"}],
            }
        ]
    }
    payload = burn.build_export_translation_payload(
        chat, offset_info={"offset": 3600.0, "mode": "manual"}
    )
    assert payload["export_offset"] == 3600.0
    assert payload["messages"][0]["timestamp"] == 3605.0
    assert payload["messages"][0]["stream_timestamp"] == 3605.0


def test_import_survives_offset_change_with_stream_export():
    """Export stream ts, re-apply different offset on HTML, import still applies."""
    from chat_window import apply_time_offset
    import twitch_chat_burn as burn

    # HTML as parsed (stream absolute)
    chat = {
        "messages": [
            {
                "author": "alice",
                "timestamp": 1000.0,
                "fragments": [{"type": "text", "text": ": hello"}],
            }
        ]
    }
    # Simulate export after offset A=1000 → video ts 0, stream 1000
    apply_time_offset(chat["messages"], 1000.0)
    payload = burn.build_export_translation_payload(
        chat, offset_info={"offset": 1000.0, "mode": "manual"}
    )
    payload["messages"][0]["translation"] = "你好"

    # Fresh parse + different offset B=900
    chat2 = {
        "messages": [
            {
                "author": "alice",
                "timestamp": 1000.0,
                "fragments": [{"type": "text", "text": ": hello"}],
            }
        ]
    }
    apply_time_offset(chat2["messages"], 900.0)
    replaced, _s, warnings = burn.apply_imported_translations(chat2, payload)
    assert replaced == 1
    assert chat2["messages"][0]["fragments"][0]["text"] == "你好"
    assert not any("时间戳不一致" in w for w in warnings)


def test_legacy_import_uses_video_relative_when_no_stream_meta():
    from chat_window import apply_time_offset
    import twitch_chat_burn as burn

    chat = {
        "messages": [
            {
                "author": "alice",
                "timestamp": 10.0,
                "fragments": [{"type": "text", "text": ": hello"}],
            }
        ]
    }
    apply_time_offset(chat["messages"], 0.0)
    # legacy: timestamp is video-relative matching current HTML timestamp
    legacy = {
        "messages": [
            {
                "index": 0,
                "author": "alice",
                "timestamp": 10.0,
                "original": "hello",
                "translation": "你好",
            }
        ]
    }
    replaced, _s, warnings = burn.apply_imported_translations(chat, legacy)
    assert replaced == 1


def test_apply_time_offset_preserves_stream_timestamp():
    from chat_window import apply_time_offset

    msgs = [{"timestamp": 500.0, "author": "u"}]
    apply_time_offset(msgs, 100.0)
    assert msgs[0]["stream_timestamp"] == 500.0
    assert msgs[0]["timestamp"] == 400.0
    # second apply with different offset must still use original stream base
    apply_time_offset(msgs, 200.0)
    assert msgs[0]["stream_timestamp"] == 500.0
    assert msgs[0]["timestamp"] == 300.0


def test_is_live_run_meta_dead_pid_and_stale():
    import run_meta as rm

    now = time.time()
    assert rm.is_live_run_meta({"status": "success"}, now=now) is False
    assert rm.is_live_run_meta({"status": "running", "pid": 99999999}, now=now) is False
    # Fresh running without pid → live (fail closed)
    assert (
        rm.is_live_run_meta(
            {"status": "running", "updated_at": time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(now))},
            now=now,
        )
        is True
    )
    # Stale updated_at
    old = time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(now - 7 * 3600))
    assert (
        rm.is_live_run_meta(
            {"status": "running", "updated_at": old},
            stale_after_sec=6 * 3600,
            now=now,
        )
        is False
    )
    # Current process is alive
    assert (
        rm.is_live_run_meta(
            {
                "status": "running",
                "pid": os.getpid(),
                "updated_at": time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(now)),
            },
            now=now,
        )
        is True
    )


def test_clean_skips_only_truly_live_jobs(tmp_path: Path):
    import process_util as pu
    import run_meta as rm

    out = tmp_path / "out"
    live = out / "job_live"
    stale = out / "job_stale"
    for d in (live, stale):
        d.mkdir(parents=True)
        (d / ".twitch_overlay_job").write_text("1", encoding="utf-8")
    rm.write_run_meta(live, {"status": "running", "pid": os.getpid()})
    # dead pid + old stamp
    old = time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(time.time() - 9 * 3600))
    (stale / "run_meta.json").write_text(
        json.dumps({"status": "running", "pid": 99999999, "updated_at": old}),
        encoding="utf-8",
    )
    assert pu._is_live_tool_job(live) is True
    assert pu._is_live_tool_job(stale) is False


def test_progress_resume_ignores_missing_fingerprint(tmp_path: Path):
    import translate_chat_openai as tr

    json_path = tmp_path / "t.json"
    data = {
        "messages": [
            {"index": 0, "author": "a", "original": "hello", "translation": ""},
            {"index": 1, "author": "b", "original": "world", "translation": ""},
        ]
    }
    json_path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    progress = {
        "schema_version": 1,
        "translations": {"0": "你好", "1": "世界"},
        # fingerprints intentionally missing (legacy)
        "failed": [],
        "target_language": "zh",
        "context": "livestream chat",
    }
    prog_path = tr.progress_path_for(json_path)
    prog_path.write_text(json.dumps(progress), encoding="utf-8")

    # Simulate the seed loop from main()
    progress_loaded = tr.load_progress(prog_path)
    progress_map = {int(k): v for k, v in (progress_loaded.get("translations") or {}).items()}
    progress_fps = progress_loaded.get("fingerprints") or {}
    translation_map = {}
    for msg in data["messages"]:
        idx = msg["index"]
        if idx in progress_map and str(progress_map[idx]).strip():
            fp_now = tr.fingerprint_message(msg)
            fp_old = str(progress_fps.get(str(idx)) or progress_fps.get(idx) or "").strip()
            if not fp_old or fp_old != fp_now:
                continue
            translation_map[idx] = progress_map[idx]
    assert translation_map == {}  # no fingerprints → do not trust

    # With matching fingerprints, should accept
    progress_fps = {str(i): tr.fingerprint_message(m) for i, m in enumerate(data["messages"])}
    translation_map = {}
    for msg in data["messages"]:
        idx = msg["index"]
        if idx in progress_map and str(progress_map[idx]).strip():
            fp_now = tr.fingerprint_message(msg)
            fp_old = str(progress_fps.get(str(idx)) or "").strip()
            if not fp_old or fp_old != fp_now:
                continue
            translation_map[idx] = progress_map[idx]
    assert translation_map[0] == "你好"
    assert translation_map[1] == "世界"


def test_pipeline_export_auto_skips_when_filled(tmp_path: Path, monkeypatch, capsys):
    import render_cn_chat as pipe

    tj = tmp_path / "t.json"
    tj.write_text(
        json.dumps(
            {"messages": [{"index": 0, "original": "a", "translation": "甲"}]},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    called = {"n": 0}

    def fake_run(cmd, **kwargs):
        called["n"] += 1

    monkeypatch.setattr(pipe, "run", fake_run)
    pipe._export_translation_json(
        burn=tmp_path / "burn.py",
        video=tmp_path / "v.mp4",
        chat_html=tmp_path / "c.html",
        trans_json=tj,
        force=False,
    )
    assert called["n"] == 0
    out = capsys.readouterr().out
    assert "跳过导出" in out or "非空" in out

    pipe._export_translation_json(
        burn=tmp_path / "burn.py",
        video=tmp_path / "v.mp4",
        chat_html=tmp_path / "c.html",
        trans_json=tj,
        force=True,
        offset=12.5,
    )
    assert called["n"] == 1


def test_pipeline_export_forwards_offset(tmp_path: Path, monkeypatch):
    import render_cn_chat as pipe

    seen = {"cmd": None}

    def fake_run(cmd, **kwargs):
        seen["cmd"] = list(cmd)

    monkeypatch.setattr(pipe, "run", fake_run)
    tj = tmp_path / "empty.json"
    pipe._export_translation_json(
        burn=tmp_path / "burn.py",
        video=tmp_path / "v.mp4",
        chat_html=tmp_path / "c.html",
        trans_json=tj,
        force=False,
        offset=42.0,
    )
    assert seen["cmd"] is not None
    assert "--offset" in seen["cmd"]
    assert seen["cmd"][seen["cmd"].index("--offset") + 1] == "42.0"
    assert "--export-translation" in seen["cmd"]
    assert seen["cmd"][seen["cmd"].index("--out-dir") + 1] == str(tj.parent)


def test_publish_output_respects_backup_prev(tmp_path: Path):
    import render_cn_chat as pipe

    src = tmp_path / "src.mp4"
    dst = tmp_path / "out.mp4"
    src.write_bytes(b"new")
    dst.write_bytes(b"old")
    pipe.publish_output(src, dst, backup_prev=False)
    assert dst.read_bytes() == b"new"
    assert not (tmp_path / "out.mp4.bak").exists()

    src2 = tmp_path / "src2.mp4"
    src2.write_bytes(b"newer")
    dst.write_bytes(b"prev")
    pipe.publish_output(src2, dst, backup_prev=True)
    assert dst.read_bytes() == b"newer"
    assert (tmp_path / "out.mp4.bak").is_file()


def test_pipeline_rejects_output_that_replaces_source_video(tmp_path: Path, monkeypatch):
    import render_cn_chat as pipe

    monkeypatch.setattr(pipe, "DRY_RUN", False)
    video = tmp_path / "source.mp4"
    chat = tmp_path / "chat.html"
    video.write_bytes(b"video")
    chat.write_text("<html></html>", encoding="utf-8")
    monkeypatch.setattr(pipe, "validate_source_media", lambda *_args, **_kwargs: pytest.fail("must reject before media check"))
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "render_cn_chat.py",
            str(video),
            str(chat),
            "--output",
            str(video),
            "--render-original",
            "--yes",
        ],
    )

    with pytest.raises(pipe.PipelineError, match="output"):
        pipe._main()


def test_resume_ignores_json_when_progress_lang_incompatible():
    """Mirror product seed rules: lang/context wipe must not trust filled JSON."""
    progress_compatible = False
    resume = True
    trust_existing_json = bool(resume and progress_compatible)
    messages = [
        {"index": 1, "original": "hello", "translation": "你好"},  # stale zh
        {"index": 2, "original": "world", "translation": ""},
    ]
    translation_map = {}
    todo = []
    for msg in messages:
        idx = msg["index"]
        existing = str(msg.get("translation", "") or "").strip()
        if trust_existing_json and existing:
            translation_map[idx] = existing
            continue
        if idx not in translation_map:
            todo.append(idx)
    assert translation_map == {}
    assert todo == [1, 2]


def test_burn_cli_export_refuses_nonempty_without_force(tmp_path: Path):
    """Real CLI: second export without --force-export must exit non-zero and keep JSON."""
    import subprocess

    html = ROOT / "tests" / "fixtures" / "twitchdownloader_chat.html"
    assert html.is_file()
    export_json = tmp_path / "t.json"
    missing_video = tmp_path / "missing.mp4"
    burn = SCRIPTS / "twitch_chat_burn.py"
    env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}

    def run_export(*, force: bool = False):
        cmd = [
            sys.executable,
            str(burn),
            str(missing_video),
            str(html),
            "--export-translation",
            str(export_json),
            "--offset",
            "0",
        ]
        if force:
            cmd.append("--force-export")
        return subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            cwd=str(ROOT),
        )

    r1 = run_export()
    assert r1.returncode == 0, (r1.stdout or "") + (r1.stderr or "")
    data = json.loads(export_json.read_text(encoding="utf-8"))
    assert data.get("time_base") == "stream"
    # Simulate hand-filled translation
    data["messages"][0]["translation"] = "已手翻"
    export_json.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    before = export_json.read_text(encoding="utf-8")

    r2 = run_export(force=False)
    assert r2.returncode != 0
    joined = (r2.stdout or "") + (r2.stderr or "")
    assert "非空" in joined or "拒绝" in joined or "force-export" in joined.lower()
    assert export_json.read_text(encoding="utf-8") == before

    r3 = run_export(force=True)
    assert r3.returncode == 0, (r3.stdout or "") + (r3.stderr or "")
    after = json.loads(export_json.read_text(encoding="utf-8"))
    assert after["messages"][0]["translation"] == ""


def test_clean_all_removes_stale_running_job(tmp_path: Path):
    """--clean --clean-all should delete crashed/stale running jobs, keep live pid jobs."""
    import subprocess

    import process_util as pu
    import run_meta as rm

    out = tmp_path / "out"
    out.mkdir()
    live = pu.make_job_dir(out, prefix="job_")
    stale = pu.make_job_dir(out, prefix="job_")
    rm.write_run_meta(live, {"status": "running", "pid": os.getpid()})
    old = time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime(time.time() - 9 * 3600))
    (stale / "run_meta.json").write_text(
        json.dumps({"status": "running", "pid": 99999999, "updated_at": old}),
        encoding="utf-8",
    )
    (stale / "junk.bin").write_bytes(b"x")

    dummy_video = tmp_path / "v.mp4"
    dummy_video.write_bytes(b"\x00")
    dummy_html = tmp_path / "c.html"
    dummy_html.write_text("<html></html>", encoding="utf-8")
    r = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "twitch_chat_burn.py"),
            str(dummy_video),
            str(dummy_html),
            "--clean",
            "--clean-all",
            "--out-dir",
            str(out),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env={**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"},
        cwd=str(ROOT),
    )
    assert r.returncode == 0, (r.stdout or "") + (r.stderr or "")
    joined = (r.stdout or "") + (r.stderr or "")
    assert live.is_dir(), "live pid job must be skipped"
    assert not stale.exists(), f"stale running job should be cleaned: {joined}"
    assert "skip live" in joined.lower() or "live" in joined.lower()


def test_fallback_manual_mentions_partial_translations(tmp_path: Path, monkeypatch, capsys):
    import render_cn_chat as pipe

    tj = tmp_path / "t.json"
    tj.write_text(
        json.dumps(
            {
                "messages": [
                    {"index": 0, "translation": "已译"},
                    {"index": 1, "translation": ""},
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    entry = tmp_path / "Program Files" / "A&B" / "twitch-chat-overlay.exe"
    monkeypatch.setattr(sys, "argv", [str(entry)])
    monkeypatch.setattr(pipe, "export_review_tsv", lambda *a, **k: None)
    monkeypatch.setattr(pipe, "export_review_xlsx", lambda *a, **k: None)
    pipe._fallback_manual_after_export(
        video=tmp_path / "v.mp4",
        chat_html=tmp_path / "c.html",
        trans_json=tj,
        review_tsv=tmp_path / "r.tsv",
        review_xlsx=tmp_path / "r.xlsx",
        workdir=None,
        final_output=tmp_path / "o.mp4",
        reason="API down",
    )
    out = capsys.readouterr().out
    assert "1" in out
    assert "非空" in out or "已有" in out
    assert pipe.current_cli_invocation() in out
    assert sys.executable not in pipe.current_cli_invocation()
    assert "render_cn_chat.py" not in out


def test_review_table_export_creates_missing_parent_dirs(tmp_path: Path):
    """复核表导出到不存在的父目录时必须自动创建（TSV mkdir 修复的守门回归）。

    用户显式传 --review-tsv/--review-xlsx 指向尚不存在的子目录（如
    <workdir>/review/r.tsv）时，缺 mkdir 会直接 FileNotFoundError，
    而此时翻译早已完成，损失的是整张复核表。
    """
    import render_cn_chat as pipe

    json_path = tmp_path / "t.json"
    json_path.write_text(
        json.dumps(
            {
                "messages": [
                    {
                        "index": 0,
                        "author": "a",
                        "timestamp": 1.0,
                        "original": "hi",
                        "translation": "你好",
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    nested_tsv = tmp_path / "review" / "deep" / "r.tsv"
    pipe.export_review_tsv(json_path, nested_tsv)
    assert nested_tsv.is_file()

    nested_xlsx = tmp_path / "review-x" / "r.xlsx"
    pipe.export_review_xlsx(json_path, nested_xlsx)
    assert nested_xlsx.is_file()


# ---------------------------------------------------------------------------
# C2·BOM: 手工编辑器写入的 BOM 不得让护栏失效
# ---------------------------------------------------------------------------


def _write_bom_json(path: Path, payload: dict) -> None:
    path.write_bytes(b"\xef\xbb\xbf" + json.dumps(payload, ensure_ascii=False).encode("utf-8"))


def test_bom_translation_json_counts_nonempty(tmp_path: Path):
    """带 BOM 的翻译 JSON 也要被 nonempty 计数识别（护栏前提）。"""
    import translation_io

    path = tmp_path / "t.json"
    _write_bom_json(
        path,
        {"messages": [{"index": 0, "translation": "你好"}]},
    )
    assert translation_io.translation_json_nonempty_count(path) == 1


def test_write_export_refuses_existing_bom_nonempty_without_force(tmp_path: Path):
    """已有 BOM 且非空译文的 JSON，force=False 导出必须拒绝覆盖。"""
    import translation_io

    path = tmp_path / "t.json"
    _write_bom_json(
        path,
        {"messages": [{"index": 0, "translation": "手翻成果"}]},
    )
    chat = {
        "messages": [
            {
                "author": "a",
                "timestamp": 1.0,
                "stream_timestamp": 1.0,
                "fragments": [{"type": "text", "text": ": hi"}],
            }
        ]
    }
    with pytest.raises(FileExistsError, match="非空 translation"):
        translation_io.write_export_translation_json(path, chat, force=False)


def test_load_json_reads_bom_file(tmp_path: Path):
    """translate_chat_openai.load_json 必须能读带 BOM 的翻译 JSON。"""
    import translate_chat_openai as tr

    path = tmp_path / "t.json"
    _write_bom_json(path, {"messages": [{"index": 0, "original": "hi"}]})
    data = tr.load_json(path)
    assert data["messages"][0]["original"] == "hi"


def test_bom_json_roundtrip_via_burn_import(tmp_path: Path):
    """BOM 翻译 JSON 经 burn 导入门面（open+json.load）正常回填。"""
    import twitch_chat_burn as burn

    chat = {
        "messages": [
            {
                "author": "alice",
                "timestamp": 1000.0,
                "fragments": [{"type": "text", "text": ": hello"}],
            }
        ]
    }
    path = tmp_path / "t.json"
    _write_bom_json(
        path,
        {
            "messages": [
                {
                    "index": 0,
                    "author": "alice",
                    "timestamp": 1000.0,
                    "original": "hello",
                    "translation": "你好",
                }
            ]
        },
    )
    with open(path, encoding="utf-8-sig") as f:
        trans_data = json.load(f)
    replaced, _stripped, warnings = burn.apply_imported_translations(chat, trans_data)
    assert replaced == 1
    assert chat["messages"][0]["fragments"][0]["text"] == "你好"
    assert not any("不一致" in w for w in warnings)


# ---------------------------------------------------------------------------
# O·导出 tmp 唯一化: 并发写同一 export_path 不得互踩
# ---------------------------------------------------------------------------


def test_concurrent_export_writes_same_path_atomically(tmp_path: Path):
    import threading

    import translation_io

    export_path = tmp_path / "t.json"
    chat = {
        "messages": [
            {
                "author": "a",
                "timestamp": 1.0,
                "stream_timestamp": 1.0,
                "fragments": [{"type": "text", "text": ": hi"}],
            }
        ]
    }
    errors: list[Exception] = []

    def worker():
        try:
            translation_io.write_export_translation_json(export_path, chat, force=True)
        except Exception as e:  # pragma: no cover - 失败时聚到断言里
            errors.append(e)

    threads = [threading.Thread(target=worker) for _ in range(2)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    assert not errors
    data = json.loads(export_path.read_text(encoding="utf-8"))
    assert data["schema_version"] == 2
    assert len(data["messages"]) == 1
    # 无残留 .tmp（mkstemp 唯一名 + finally unlink）
    assert not list(tmp_path.glob("*.json.*.tmp"))


# ---------------------------------------------------------------------------
# R3·XLSX 公式注入: 以 = + - @ 开头的内容不得被 Excel 当公式执行
# ---------------------------------------------------------------------------


def test_review_xlsx_formula_injection_roundtrip(tmp_path: Path):
    """original="=1+1" 导出后不得成为公式；导入剥掉防注入前缀保真还原。"""
    from openpyxl import load_workbook

    import render_cn_chat as pipeline

    src = {
        "messages": [
            {
                "index": 0,
                "timestamp": 1.0,
                "author": "a",
                "original": "=1+1",
                "translation": "已译",
            }
        ]
    }
    json_path = tmp_path / "t.json"
    json_path.write_text(json.dumps(src, ensure_ascii=False), encoding="utf-8")
    xlsx = tmp_path / "r.xlsx"

    pipeline.export_review_xlsx(json_path, xlsx)

    wb = load_workbook(xlsx)
    ws = wb.active
    # 回读断言：original 单元格不是公式（data_type 'f' 即公式）
    original_cell = ws.cell(row=2, column=4)
    assert original_cell.data_type != "f"
    assert str(original_cell.value).startswith("'")

    # 回写 translation 后导入：剥离防注入前缀，还原原文内容
    src["messages"][0]["translation"] = "=2+2"
    json_path.write_text(json.dumps(src, ensure_ascii=False), encoding="utf-8")
    pipeline.export_review_xlsx(json_path, xlsx)
    pipeline.import_review_xlsx(json_path, xlsx)
    data = json.loads(json_path.read_text(encoding="utf-8"))
    assert data["messages"][0]["translation"] == "=2+2"


def test_review_xlsx_original_stays_text_type(tmp_path: Path):
    """original/translation 列必须按文本（'@'）存储，防止类型漂移。"""
    from openpyxl import load_workbook

    import render_cn_chat as pipeline

    src = {
        "messages": [
            {
                "index": 0,
                "timestamp": 1.0,
                "author": "a",
                "original": "2024-01-02",
                "translation": "1.5倍上分",
            }
        ]
    }
    json_path = tmp_path / "t.json"
    json_path.write_text(json.dumps(src, ensure_ascii=False), encoding="utf-8")
    xlsx = tmp_path / "r.xlsx"
    pipeline.export_review_xlsx(json_path, xlsx)

    wb = load_workbook(xlsx)
    ws = wb.active
    assert ws.cell(row=2, column=4).number_format == "@"
    assert ws.cell(row=2, column=5).number_format == "@"


# ---------------------------------------------------------------------------
# R5·severity 合并单调升级
# ---------------------------------------------------------------------------


def test_review_issue_map_severity_monotonic_upgrade(tmp_path: Path):
    """同 index 多条 issue：severity 只升不降（[OK, WARN]→WARN 等）。"""
    import review_tables as rt

    json_path = tmp_path / "t.json"
    json_path.write_text("{}", encoding="utf-8")

    def issue(sev: str) -> dict:
        return {"index": 0, "severity": sev, "code": f"c_{sev}", "message": sev}

    def merge(*sevs: str) -> dict:
        return rt._review_issue_map(
            json_path,
            data={"messages": []},
            lint_fn=lambda *_a, **_k: [issue(s) for s in sevs],
        )[0]

    assert merge("WARN", "OK")["severity"] == "WARN"
    assert merge("OK", "WARN")["severity"] == "WARN"
    assert merge("WARN", "FAIL")["severity"] == "FAIL"
    assert merge("FAIL", "WARN", "OK")["severity"] == "FAIL"
    # 未知 severity 归一为 WARN，不抛异常
    assert merge("OK", "WEIRD")["severity"] == "WARN"


# ---------------------------------------------------------------------------
# 导入身份门·细节回归（tests-6）
#   1) 重复 index 行：后写覆盖先写 + 警告文案
#   2) 手编 JSON 缺 author/original 字段：不得因此被拒
#   3) 时间戳容差 0.51s 边界（0.5 应用 / 0.6 跳过）+ 时间戳无法解析分支
# ---------------------------------------------------------------------------


def _single_msg_chat(author: str = "alice", ts: float = 100.0) -> dict:
    return {
        "messages": [
            {
                "author": author,
                "timestamp": ts,
                "stream_timestamp": ts,
                "fragments": [{"type": "text", "text": ": hello"}],
            }
        ]
    }


def test_import_duplicate_index_last_write_wins_and_warns():
    """同一 JSON 出现重复 index：后写覆盖先写，且必须发"后写覆盖先写"警告。

    变异验证锚点：若删掉 dup 检测块（仅抑制警告、保留 trans_map 赋值），
    本测试的警告断言必须失败；若删块连带丢赋值，覆盖断言也会失败。
    """
    import translation_io

    chat = _single_msg_chat()
    trans = {
        "schema_version": 2,
        "time_base": "stream",
        "messages": [
            {"index": 0, "author": "alice", "timestamp": 100.0, "original": "hello", "translation": "先写"},
            {"index": 0, "author": "alice", "timestamp": 100.0, "original": "hello", "translation": "后写"},
        ],
    }
    replaced, _stripped, warnings = translation_io.apply_imported_translations(chat, trans)
    assert replaced == 1
    assert chat["messages"][0]["fragments"][0]["text"] == "后写"
    dup = [w for w in warnings if "重复 index" in w]
    assert dup, f"missing dup-index warning, warnings={warnings!r}"
    assert "后写覆盖先写" in dup[0]
    assert "0" in dup[0]
    # 只有 index=0 一行（重复不计），消息与 map 条数一致，不应有条数不一致警告
    assert not any("条数" in w for w in warnings)


def test_import_row_without_author_and_original_still_applies():
    """手编 JSON 缺 author、缺 original 字段：时间戳匹配时正常导入。

    变异验证锚点：若 author 守卫（exp_author is not None）被改成无条件比较，
    None != 'alice' 会判为身份不一致而跳过导入，本测试失败。
    """
    import translation_io

    chat = _single_msg_chat(author="alice", ts=100.0)
    trans = {
        "schema_version": 2,
        "time_base": "stream",
        "messages": [
            {"index": 0, "timestamp": 100.0, "translation": "你好"},
        ],
    }
    replaced, _stripped, warnings = translation_io.apply_imported_translations(chat, trans)
    assert replaced == 1
    assert chat["messages"][0]["fragments"][0]["text"] == "你好"
    assert not any("不一致" in w for w in warnings)


@pytest.mark.parametrize(
    ("exp_ts", "should_apply"),
    [
        (100.5, True),   # |Δ| = 0.5，未超 0.51 容差 → 应用
        (100.6, False),  # |Δ| = 0.6，超 0.51 容差 → 跳过
    ],
)
def test_import_timestamp_tolerance_boundary(exp_ts: float, should_apply: bool):
    """时间戳容差边界 |Δ|∈{0.5, 0.6}：0.5 应用、0.6 跳过。

    变异验证锚点：0.51 → 1e9 时 0.6 也被放行（should_apply=False 分支失败）；
    0.51 → 0 时 0.5 也被拒（should_apply=True 分支失败）。
    """
    import translation_io

    chat = _single_msg_chat(author="alice", ts=100.0)
    trans = {
        "schema_version": 2,
        "time_base": "stream",
        "messages": [
            {"index": 0, "author": "alice", "timestamp": exp_ts, "translation": "你好"},
        ],
    }
    replaced, _stripped, warnings = translation_io.apply_imported_translations(chat, trans)
    if should_apply:
        assert replaced == 1
        assert chat["messages"][0]["fragments"][0]["text"] == "你好"
        assert not any("时间戳不一致" in w for w in warnings)
    else:
        assert replaced == 0
        assert chat["messages"][0]["fragments"][0]["text"] == ": hello"
        assert any("时间戳不一致(stream)" in w for w in warnings)


def test_import_unparseable_timestamp_is_flagged_and_skipped():
    """timestamp 为非数值字符串：走"时间戳无法解析"分支，行被跳过不误贴。"""
    import translation_io

    chat = _single_msg_chat(author="alice", ts=100.0)
    trans = {
        "schema_version": 2,
        "time_base": "stream",
        "messages": [
            {"index": 0, "author": "alice", "timestamp": "not-a-number", "translation": "你好"},
        ],
    }
    replaced, _stripped, warnings = translation_io.apply_imported_translations(chat, trans)
    assert replaced == 0
    assert chat["messages"][0]["fragments"][0]["text"] == ": hello"
    unparseable = [w for w in warnings if "时间戳无法解析" in w]
    assert unparseable, f"missing unparseable-ts warning, warnings={warnings!r}"
    # 实际文案带 repr 的 JSON 值：'时间戳无法解析: 翻译 JSON='not-a-number''
    assert "'not-a-number'" in unparseable[0]
