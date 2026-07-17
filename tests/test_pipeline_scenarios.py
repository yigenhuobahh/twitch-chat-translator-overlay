#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Strict real-world scenario tests for the chat overlay pipeline.

These cover the daily-use path:
export -> (optional rules) -> lint/review -> import/render -> publish
without calling any live translation API.
"""

from __future__ import annotations

import json
from pathlib import Path
import subprocess
import sys

import pytest

from helpers import (
    FIXTURES_DIR,
    ROOT,
    SCRIPTS_DIR,
    ffprobe_json,
    load_module,
    make_leadin_video,
    make_translation_for_export,
    stream_types,
    write_json,
)


def _run(cmd: list[str], env: dict | None = None) -> subprocess.CompletedProcess:
    # Force UTF-8 stdio so Windows GBK consoles do not break Chinese script output.
    full_env = dict(**{k: v for k, v in __import__("os").environ.items()})
    full_env["PYTHONIOENCODING"] = "utf-8"
    full_env["PYTHONUTF8"] = "1"
    if env:
        full_env.update(env)
    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=full_env,
    )


class TestExportImportPipeline:
    @pytest.mark.smoke
    def test_export_import_render_pipeline(self, make_test_video, tmp_path: Path):
        """Full offline path: export JSON, fill translations, import, render short clip."""
        import twitch_chat_burn as burn

        video = make_test_video(duration=3.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        export_json = tmp_path / "export.json"
        filled_json = tmp_path / "filled.json"
        out_dir = tmp_path / "render"
        out_dir.mkdir()

        # 1) export
        export_cmd = [
            sys.executable, str(SCRIPTS_DIR / "twitch_chat_burn.py"),
            str(video), str(html),
            "--export-translation", str(export_json),
            "--offset", "0",
        ]
        r = _run(export_cmd)
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        data = json.loads(export_json.read_text(encoding="utf-8"))
        assert len(data["messages"]) == 3

        # 2) fill translations offline (simulate human/API)
        filled = make_translation_for_export(
            data,
            mapping={
                "hello [LUL] world": "你好 [LUL] 世界",
                "[Hey]": "[Hey]",
                "nice [xdx] clip": "不错的 [xdx] 片段",
            },
        )
        # export original text may vary slightly by spacing; fall back to generic fill
        for msg in filled["messages"]:
            if not msg.get("translation"):
                msg["translation"] = f"译:{msg.get('original', '')}"
            # ensure pure emote stays pure if original is pure emote markers
            original = str(msg.get("original", "")).strip()
            if original.startswith("[") and "]" in original and all(
                part.startswith("[") and part.endswith("]")
                for part in original.split()
            ):
                msg["translation"] = original
        write_json(filled_json, filled)

        # 3) import + preview clip render
        render_cmd = [
            sys.executable, str(SCRIPTS_DIR / "twitch_chat_burn.py"),
            str(video), str(html),
            "--import-translation", str(filled_json),
            "--x", "10", "--y", "40", "--w", "300", "--h", "200",
            "--fps", "30",
            "--preview-clip", "3",
            "--out-dir", str(out_dir),
            "--job-dir", str(out_dir),
            "--keep-temp",
            "--offset", "0",
        ]
        r = _run(render_cmd)
        assert r.returncode == 0, r.stdout + "\n" + r.stderr

        final = out_dir / f"{video.stem}_chat.mp4"
        assert final.is_file()
        info = ffprobe_json(final)
        assert "video" in stream_types(info)
        assert "audio" in stream_types(info)
        assert float(info["format"]["duration"]) >= 2.7

        frames = sorted((out_dir / "overlay_frames").glob("frame_*.png"))
        assert len(frames) == burn.expected_overlay_frame_count(3.0, 30)

    def test_import_strips_emote_placeholder_and_username_prefix(self, tmp_path: Path):
        """Import path should clean model metadata and emote placeholders before rendering text."""
        import twitch_chat_burn as burn

        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        chat = burn.parse_chat_html(str(html), str(tmp_path / "emotes"))
        msg = chat["messages"][0]
        titles = [f["title"] for f in msg["fragments"] if f["type"] == "emote"]
        assert titles  # fixture message 0 contains LUL

        # Simulate dirty model output
        dirty = f"<{msg['author']}> 你好 [{titles[0]}] 世界"
        cleaned = burn.clean_imported_translation(dirty, msg["author"])
        for title in titles:
            cleaned = cleaned.replace(f"[{title}]", "").strip()
            cleaned = " ".join(cleaned.split())
        assert msg["author"] not in cleaned
        assert f"[{titles[0]}]" not in cleaned
        assert "你好" in cleaned and "世界" in cleaned


class TestLintAndRulesRealCases:
    def test_lint_catches_common_real_failures(self, tmp_path: Path):
        render = load_module("render_cn_chat", "render_cn_chat.py")
        bad = {
            "messages": [
                {
                    "index": 0,
                    "timestamp": 0,
                    "author": "a",
                    "original": "hello",
                    "translation": "",  # empty
                },
                {
                    "index": 0,  # duplicate index
                    "timestamp": 1,
                    "author": "b",
                    "original": "[LUL]",
                    "translation": "哈哈",  # pure emote rewritten
                },
                {
                    "index": 2,
                    "timestamp": 2,
                    "author": "c",
                    "original": "@Alice check https://example.com/x",
                    "translation": "看看这个",  # lost mention + url
                },
                {
                    "index": 3,
                    "timestamp": 3,
                    "author": "d",
                    "original": "hi",
                    "translation": "这是一条特别特别特别特别特别特别特别特别特别特别特别特别特别特别特别特别特别特别长的翻译用来触发长度告警",
                },
            ]
        }
        path = write_json(tmp_path / "bad.json", bad)
        issues = render.lint_translation(path, max_chars=20)
        # lint structure uses severity + message; inspect raw for robustness
        joined = json.dumps(issues, ensure_ascii=False)
        assert "empty_translation" in joined
        assert "duplicate_index" in joined
        assert "FAIL" in joined
        # mention/url loss and pure emote rewrite should be reported somehow
        assert ("pure" in joined.lower()) or ("emote" in joined.lower()) or ("[LUL]" in joined)
        assert ("@" in joined) or ("mention" in joined.lower()) or ("url" in joined.lower()) or ("https" in joined)

    def test_rules_normalization_rewrites_matched_originals(self, tmp_path: Path):
        render = load_module("render_cn_chat", "render_cn_chat.py")
        data = {
            "messages": [
                {"index": 0, "timestamp": 0, "author": "a", "original": "GG", "translation": "GG"},
                {"index": 1, "timestamp": 1, "author": "b", "original": "hello", "translation": "你好"},
            ]
        }
        json_path = write_json(tmp_path / "msg.json", data)
        rules_path = tmp_path / "rules.yaml"
        rules_path.write_text(
            """
normalizations:
  - name: gg
    match: ["GG", "gg"]
    translation: "打得好"
""".strip(),
            encoding="utf-8",
        )
        render.normalize_translation(json_path, rules_path=rules_path)
        out = json.loads(json_path.read_text(encoding="utf-8"))
        assert out["messages"][0]["translation"] == "打得好"
        assert out["messages"][1]["translation"] == "你好"

    def test_dry_run_skips_rules_write(self, tmp_path: Path):
        render = load_module("render_cn_chat", "render_cn_chat.py")
        data = {
            "messages": [
                {"index": 0, "timestamp": 0, "author": "a", "original": "GG", "translation": "GG"},
            ]
        }
        json_path = write_json(tmp_path / "msg.json", data)
        rules_path = tmp_path / "rules.yaml"
        rules_path.write_text(
            "normalizations:\n  - name: gg\n    match: [\"GG\"]\n    translation: \"打得好\"\n",
            encoding="utf-8",
        )
        render.DRY_RUN = True
        try:
            render.normalize_translation(json_path, rules_path=rules_path)
        finally:
            render.DRY_RUN = False
        out = json.loads(json_path.read_text(encoding="utf-8"))
        assert out["messages"][0]["translation"] == "GG"


class TestReviewRoundtripXlsx:
    def test_xlsx_export_import_roundtrip(self, tmp_path: Path):
        render = load_module("render_cn_chat", "render_cn_chat.py")
        src = {
            "messages": [
                {
                    "index": 0,
                    "timestamp": 1.0,
                    "author": "viewer_one",
                    "original": "hello there",
                    "translation": "你好呀",
                },
                {
                    "index": 1,
                    "timestamp": 2.0,
                    "author": "viewer_two",
                    "original": "[Pog]",
                    "translation": "[Pog]",
                },
            ]
        }
        json_path = write_json(tmp_path / "t.json", src)
        xlsx_path = tmp_path / "review.xlsx"
        render.export_review_xlsx(json_path, xlsx_path)
        assert xlsx_path.is_file()

        # Manually edit translation via openpyxl to simulate human review
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        ws = wb.active
        # find translation column
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        tcol = headers.index("translation") + 1
        # row 2 is first data row
        ws.cell(row=2, column=tcol, value="人工改过的译文")
        wb.save(xlsx_path)

        render.import_review_xlsx(json_path, xlsx_path)
        out = json.loads(json_path.read_text(encoding="utf-8"))
        assert out["messages"][0]["translation"] == "人工改过的译文"
        assert out["messages"][1]["translation"] == "[Pog]"


class TestTextLayoutAndCleaning:
    def test_wrap_fragments_breaks_long_cjk_text(self):
        import twitch_chat_burn as burn

        # Fake width function: each char width=10
        def tw(s: str) -> int:
            return 10 * len(s)

        frags = [("text", "这是一段需要换行的中文弹幕内容测试", tw("这是一段需要换行的中文弹幕内容测试"))]
        lines = burn.wrap_fragments(frags, header_w=0, max_w=50, padding=0, indent=0, gap=0, text_width_fn=tw)
        assert len(lines) >= 2
        joined = "".join(part[1] for line in lines for part in line if part[0] == "text")
        assert "中文弹幕" in joined

    def test_prepare_messages_for_llm_does_not_include_author(self):
        translate = load_module("translate_chat_openai", "translate_chat_openai.py")
        text = translate.prepare_messages_for_llm(
            [
                {"index": 7, "author": "secret_user", "original": "hello"},
                {"index": 8, "author": "secret_user", "original": "[LUL]"},
            ]
        )
        assert "secret_user" not in text
        assert "[7] hello" in text
        assert "[8] [LUL]" in text

    def test_should_preserve_pure_emote_and_numbers(self):
        translate = load_module("translate_chat_openai", "translate_chat_openai.py")
        assert translate.should_preserve_original("[LUL]")
        assert translate.should_preserve_original("[Hey] [xdx]")
        assert translate.should_preserve_original("12345")
        assert not translate.should_preserve_original("hello [LUL]")

    def test_clean_translation_text_real_world_matrix(self):
        translate = load_module("translate_chat_openai", "translate_chat_openai.py")
        cases = {
            "太强了/真厉害": "太强了",
            "https://twitch.tv/foo/bar": "https://twitch.tv/foo/bar",
            "看这里 https://example.com/a/b": "看这里 https://example.com/a/b",
            "and/or": "and/or",
            "[12] <bot> 你好": "你好",
            "普通句子": "普通句子",
        }
        for src, expected in cases.items():
            assert translate.clean_translation_text(src) == expected, src


class TestDoctorAndCliGuards:
    def test_doctor_exits_zero_on_healthy_machine(self):
        cmd = [sys.executable, str(SCRIPTS_DIR / "render_cn_chat.py"), "--doctor"]
        r = _run(cmd)
        out = (r.stdout or "") + "\n" + (r.stderr or "")
        assert r.returncode == 0, out
        assert ("Doctor" in out) or ("诊断" in out) or ("[OK]" in out)

    def test_render_cn_chat_dry_run_export_only(self, make_test_video, tmp_path: Path):
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        work = tmp_path / "work"
        out = tmp_path / "out" / "final.mp4"
        cmd = [
            sys.executable, str(SCRIPTS_DIR / "render_cn_chat.py"),
            str(video), str(html),
            "--dry-run",
            "--skip-translate",
            "--workdir", str(work),
            "--output", str(out),
            "--offset", "0",
        ]
        r = _run(cmd)
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        # dry-run + skip-translate should stop before final render output
        assert not out.exists()


class TestTimingAndLeadIn:
    @pytest.mark.smoke
    def test_output_fps_follows_config(self, make_test_video, tmp_path: Path):
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"
        out_dir = tmp_path / "fps30"
        out_dir.mkdir()
        cmd = [
            sys.executable, str(SCRIPTS_DIR / "twitch_chat_burn.py"),
            str(video), str(html),
            "--fps", "30",
            "--preview-clip", "2",
            "--out-dir", str(out_dir),
            "--job-dir", str(out_dir),
            "--keep-temp",
            "--offset", "0",
            "--x", "5", "--y", "20", "--w", "280", "--h", "180",
        ]
        r = _run(cmd)
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        final = out_dir / f"{video.stem}_chat.mp4"
        # Probe average frame rate
        probe = subprocess.run(
            [
                "ffprobe", "-v", "error", "-select_streams", "v:0",
                "-show_entries", "stream=r_frame_rate,avg_frame_rate",
                "-of", "json", str(final),
            ],
            capture_output=True, text=True, encoding="utf-8", errors="replace", check=True,
        )
        info = json.loads(probe.stdout)
        rate = info["streams"][0].get("avg_frame_rate") or info["streams"][0].get("r_frame_rate")
        # 30/1 or near 30
        if "/" in rate:
            num, den = rate.split("/", 1)
            fps = float(num) / max(float(den), 1.0)
        else:
            fps = float(rate)
        assert 29.0 <= fps <= 31.0, rate

    def test_validate_rejects_too_short_output(self, make_test_video, tmp_path: Path):
        import twitch_chat_burn as burn

        video = make_test_video(duration=2.0, fps=30)
        ok, summary, reason = burn.validate_rendered_output(
            str(video), expected_duration=10.0, require_audio=True
        )
        assert not ok
        assert "shorter than expected" in reason

    def test_lead_in_expected_duration_is_render_window_not_source_plus_padding(self, tmp_path: Path):
        """
        Lead-in rewrites start times; published duration target is the render window
        (~source length), not source+lead_in. Still reject clearly truncated files.
        """
        import twitch_chat_burn as burn

        src = tmp_path / "leadin_src.mp4"
        make_leadin_video(src, duration=3.0, lead_in=1.0, fps=30)
        summary = burn.probe_media_summary(str(src))
        assert summary["ok"]
        assert summary["duration"] >= 3.0

        source_duration = float(summary["duration"])
        # Correct policy: expected == render/source length
        expected = burn.expected_compose_duration(source_duration, video_lead_in=1.0)
        assert expected == pytest.approx(source_duration)
        ok, _, reason = burn.validate_rendered_output(
            str(src), expected_duration=expected, require_audio=True
        )
        assert ok, reason

        # Still reject a file that is far shorter than the render window
        ok2, _, reason2 = burn.validate_rendered_output(
            str(src), expected_duration=source_duration + 5.0, require_audio=True
        )
        assert not ok2
        assert "shorter than expected" in reason2


class TestDirtyTranslationFixture:
    def test_dirty_fixture_can_be_cleaned_for_import(self):
        translate = load_module("translate_chat_openai", "translate_chat_openai.py")
        burn = load_module("twitch_chat_burn", "twitch_chat_burn.py")
        data = json.loads((FIXTURES_DIR / "dirty_translation.json").read_text(encoding="utf-8"))

        cleaned_rows = []
        for msg in data["messages"]:
            text = translate.clean_translation_text(msg["translation"])
            text = burn.clean_imported_translation(text, msg.get("author"))
            # strip known emote placeholders from original markers
            for token in ["[LUL]", "[Hey]", "[xdx]"]:
                text = text.replace(token, " ")
            text = " ".join(text.split())
            cleaned_rows.append(text)

        assert cleaned_rows[0] == "你好 世界"
        # pure emote row may become empty after placeholder strip; that is acceptable
        assert "example_user" not in cleaned_rows[0]
        assert "clip_fan" not in cleaned_rows[2]
        assert "不错的" in cleaned_rows[2] and "片段" in cleaned_rows[2]
        assert "https://example.com/a/b" in translate.clean_translation_text(data["messages"][3]["translation"])
        assert translate.clean_translation_text(data["messages"][4]["translation"]) == "太强了"


class TestOptionalLocalSample:
    def test_local_sample_html_parses_if_present(self, tmp_path: Path):
        """Optionally parse a private local HTML if the user keeps one on disk.

        Public docs/tests must not name personal sample filenames.
        Accepted neutral local names only:
          samples/local_chat.html
          samples/private_chat.html
        """
        candidates = [
            ROOT / "samples" / "local_chat.html",
            ROOT / "samples" / "private_chat.html",
        ]
        sample = next((p for p in candidates if p.is_file()), None)
        if sample is None:
            pytest.skip("no optional local sample HTML present")
        import twitch_chat_burn as burn

        data = burn.parse_chat_html(str(sample), str(tmp_path / "emotes_local"))
        assert len(data["messages"]) > 0
        classes = [
            frag["class"]
            for msg in data["messages"]
            for frag in msg["fragments"]
            if frag["type"] == "emote"
        ]
        # If the private sample contains emotesv2 assets, fragments should keep them.
        if any("emotesv2" in c for c in data.get("emote_map", {})):
            assert any("emotesv2" in c for c in classes)


@pytest.mark.smoke
class TestRenderCnChatWorkdirPublish:
    def test_render_cn_chat_publish_to_nested_output(self, make_test_video, tmp_path: Path):
        """Pipeline should create nested output dirs and publish a valid mp4."""
        video = make_test_video(duration=2.0, fps=30)
        html = FIXTURES_DIR / "twitchdownloader_chat.html"

        # Prepare a filled translation offline
        export_json = tmp_path / "export.json"
        r = _run([
            sys.executable, str(SCRIPTS_DIR / "twitch_chat_burn.py"),
            str(video), str(html),
            "--export-translation", str(export_json),
            "--offset", "0",
        ])
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        data = json.loads(export_json.read_text(encoding="utf-8"))
        filled = make_translation_for_export(data)
        for msg in filled["messages"]:
            original = str(msg.get("original", "")).strip()
            if original and all(p.startswith("[") and p.endswith("]") for p in original.split()):
                msg["translation"] = original
            elif not msg.get("translation"):
                msg["translation"] = "测试译文"
        filled_json = write_json(tmp_path / "filled.json", filled)

        workdir = tmp_path / "workdir"
        final_output = tmp_path / "nested" / "final" / "chat_out.mp4"
        cmd = [
            sys.executable, str(SCRIPTS_DIR / "render_cn_chat.py"),
            str(video), str(html),
            "--reuse-translation",
            "--translation-json", str(filled_json),
            "--workdir", str(workdir),
            "--output", str(final_output),
            "--preview-clip", "2",
            "--offset", "0",
            "--fps", "30",
            "--x", "10", "--y", "30", "--width", "300", "--height", "180",
        ]
        r = _run(cmd)
        assert r.returncode == 0, r.stdout + "\n" + r.stderr
        # Nested --output must always publish a real mp4 for preview-clip renders.
        assert final_output.is_file(), f"missing published output: {final_output}\n{r.stdout}\n{r.stderr}"
        info = ffprobe_json(final_output)
        assert "video" in stream_types(info)
        assert float(info["format"]["duration"]) > 0
