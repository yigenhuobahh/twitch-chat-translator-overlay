#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""人工复核表（TSV/XLSX）导出导入、翻译质检（lint）与 YAML/发布/规则清洗簇。

从 render_cn_chat.py 原样搬出（搬运而非重写）：本模块保持纯函数 —— 不读
模块级可变全局；dry-run 与日志经显式参数注入，由调用方（render_cn_chat 的
薄包装）传入 `dry_run=DRY_RUN, log=log`，以保留测试按模块属性 monkeypatch
DRY_RUN 的语义。render_cn_chat 对这些名字保持 re-export/包装，外部签名不变。

刀五迁入 load_yaml_file / load_yaml_rules / load_profile / publish_output /
normalize_translation（原样搬运）： PipelineError 一并在此单源定义
（render_cn_chat re-export），使 YAML/发布簇能抛出同一错误类而不产生
review_tables -> render_cn_chat 的反向导入。
"""

from __future__ import annotations

import json
import os
from pathlib import Path
import re
import shutil
import tempfile

from common_utils import atomic_replace_with_retry, atomic_write_json


class PipelineError(SystemExit):
    """流水线错误（原 render_cn_chat.PipelineError，单源迁至此处）。

    继承 SystemExit，CLI 退出码链保持不变；render_cn_chat 顶部 re-export
    该名字，`from render_cn_chat import PipelineError` 等历史消费者拿到的
    是同一个类对象。
    """


# 复核表 severity 单调升级顺序：同一条消息命中多条规则时取最严重者，
# 而不是被后到的 OK/WARN 覆盖。
_SEVERITY_RANK = {"OK": 0, "WARN": 1, "FAIL": 2}


def _review_issue_map(json_path: Path, max_chars: int = 90, data: dict | None = None, lint_fn=None):
    """Map message index -> (severity, codes, notes) from lint, without printing a full report.

    data: 调用方已解析好的翻译 JSON；缺省 None 时自行读盘解析（失败仅告警返回空表）。
    lint_fn: 缺省用本模块 lint_translation；render_cn_chat 的包装传入自己的
    （可能被测试 monkeypatch 的）lint_translation，保持历史 spy 语义。
    """
    if data is None:
        try:
            data = json.loads(json_path.read_text(encoding="utf-8-sig"))
        except FileNotFoundError:
            print(f"[WARN] 复核表 lint 跳过：找不到 {json_path}", flush=True)
            return {}
        except (OSError, UnicodeError, json.JSONDecodeError) as e:
            # Do not fail export on bad JSON here; surface why lint columns are empty.
            print(f"[WARN] 复核表 lint 跳过：无法解析 {json_path}: {e}", flush=True)
            return {}
    # Reuse lint rules quietly; suppress console noise via stdout redirect.
    import contextlib
    import io
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        try:
            issues = (lint_fn or lint_translation)(json_path, report_path=None, max_chars=max_chars, data=data)
        except SystemExit:
            issues = []
        except Exception as e:
            print(f"[WARN] 复核表 lint 失败: {e}", flush=True)
            issues = []
    by_index: dict = {}
    for issue in issues:
        idx = issue.get("index")
        bucket = by_index.setdefault(idx, {"severity": "OK", "codes": [], "notes": []})
        sev = str(issue.get("severity", "WARN")).upper()
        if sev not in _SEVERITY_RANK:
            sev = "WARN"
        if _SEVERITY_RANK[sev] > _SEVERITY_RANK[bucket["severity"]]:
            bucket["severity"] = sev
        code = str(issue.get("code", ""))
        if code:
            bucket["codes"].append(code)
        note = str(issue.get("message", ""))
        if note:
            bucket["notes"].append(note)
    return by_index


def _xlsx_formula_sanitize(value):
    """对以 = + - @ \\t 开头的字符串前置单引号，阻断 Excel 电子表格注入。

    只处理 str 类型；单引号是 Excel 官方"按文本处理"的标记，导入侧
    _strip_formula_quote 会剥掉它，保真往返。
    """
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@", "\t"):
        return "'" + value
    return value


def _strip_formula_quote(value):
    """剥掉导出侧为防注入添加的前置单引号（只在我们自己加的约定内剥离）。"""
    if isinstance(value, str) and value.startswith("'") and len(value) > 1 and value[1:2] in ("=", "+", "-", "@", "\t"):
        return value[1:]
    return value


def _review_rows(
    json_path: Path,
    include_lint: bool = True,
    max_chars: int = 90,
    data: dict | None = None,
    issue_map: dict | None = None,
):
    if data is None:
        data = json.loads(json_path.read_text(encoding="utf-8-sig"))
    if include_lint:
        if issue_map is None:
            issue_map = _review_issue_map(json_path, max_chars=max_chars, data=data)
    else:
        issue_map = {}
    rows = []
    for msg in data.get("messages", []):
        idx = msg.get("index", "")
        info = issue_map.get(idx) or issue_map.get(str(idx)) or {}
        severity = info.get("severity", "OK") if include_lint else ""
        codes = ",".join(info.get("codes") or []) if include_lint else ""
        notes = " | ".join(info.get("notes") or []) if include_lint else ""
        rows.append([
            idx,
            msg.get("timestamp", ""),
            msg.get("author", ""),
            str(msg.get("original", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
            str(msg.get("translation", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
            severity,
            codes,
            notes,
        ])
    return rows


def _prepare_review_export(json_path: Path, max_chars: int = 90, lint_fn=None) -> tuple[dict, dict]:
    """翻译 JSON 只解析一次、lint 只跑一次；结果供 TSV/XLSX 两个导出函数共用。"""
    data = json.loads(json_path.read_text(encoding="utf-8-sig"))
    issue_map = _review_issue_map(json_path, max_chars=max_chars, data=data, lint_fn=lint_fn)
    return data, issue_map


def export_review_tsv(
    json_path: Path,
    review_path: Path,
    *,
    data: dict | None = None,
    issue_map: dict | None = None,
    dry_run: bool = False,
    log=None,
):
    """导出人工复核 TSV。translation 列可直接编辑后再导入。

    data/issue_map: 由 _prepare_review_export 预计算传入（JSON/lint 只算一次）；
    缺省时函数内部自行计算。
    """
    if dry_run:
        (log or print)(f"[dry-run] 跳过复核表 TSV 写出: {review_path}")
        return
    lines = ["index\ttimestamp\tauthor\toriginal\ttranslation\tlint_severity\tlint_codes\tlint_notes"]
    for row in _review_rows(json_path, include_lint=True, data=data, issue_map=issue_map):
        # 防 Excel 公式注入：以 = + - @ \t 开头的字符串前置单引号（TSV 里
        # 同样生效——Excel 打开 TSV 时也会执行公式）。
        lines.append(
            "\t".join(
                str(_xlsx_formula_sanitize(v)) if i in (2, 3, 4) else str(v)
                for i, v in enumerate(row)
            )
        )
    review_path.parent.mkdir(parents=True, exist_ok=True)
    # 原子写：同目录唯一 tmp（utf-8-sig 逐字节同 write_text）+ retry replace，
    # 中断/共享冲突不会半写复核表（C-5，与 atomic_write_json 同款模式）。
    fd, tmp_name = tempfile.mkstemp(prefix=f".{review_path.name}.", suffix=".tmp", dir=str(review_path.parent))
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8-sig") as file:
            file.write("\n".join(lines) + "\n")
        atomic_replace_with_retry(tmp_path, review_path)
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
    print(f"\n[人工复核] 已导出中英对照 TSV: {review_path}")


def export_review_xlsx(
    json_path: Path,
    review_path: Path,
    *,
    data: dict | None = None,
    issue_map: dict | None = None,
    dry_run: bool = False,
    log=None,
):
    """导出带列宽、换行和冻结表头的人工复核 XLSX。

    data/issue_map: 与 export_review_tsv 相同的预计算入参；dry-run 下不写出。
    """
    if dry_run:
        (log or print)(f"[dry-run] 跳过复核表 XLSX 写出: {review_path}")
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise SystemExit("错误: 导出 XLSX 需要 openpyxl，请先运行 python -m pip install openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    header = ["index", "timestamp", "author", "original", "translation", "lint_severity", "lint_codes", "lint_notes"]
    ws.append(header)
    for row in _review_rows(json_path, include_lint=True, data=data, issue_map=issue_map):
        # 防 Excel 公式注入：在 _review_rows 的 \t\r\n 归一之后、append 之前
        # 处理（仅 str 列 C=author / D=original / E=translation）。
        ws.append([_xlsx_formula_sanitize(v) if i in (2, 3, 4) else v for i, v in enumerate(row)])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fail_fill = PatternFill("solid", fgColor="F8CBAD")
    warn_fill = PatternFill("solid", fgColor="FFE699")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 8, "B": 10, "C": 20, "D": 50, "E": 50, "F": 12, "G": 24, "H": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        # 类型漂移防护：original/translation 一律按文本存储，防止 "=1+1"
        # 之类内容被 Excel 重解释成数字/公式后类型丢失。
        row[3].number_format = "@"
        row[4].number_format = "@"
        row[3].alignment = Alignment(vertical="top", wrap_text=True)
        row[4].alignment = Alignment(vertical="top", wrap_text=True)
        sev = str(row[5].value or "").upper()
        if sev == "FAIL":
            row[5].fill = fail_fill
        elif sev == "WARN":
            row[5].fill = warn_fill

    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 36

    review_path.parent.mkdir(parents=True, exist_ok=True)
    # 原子写：先存到带 .xlsx 后缀的同目录 tmp 名（openpyxl 按扩展名定格式），
    # 再 retry replace——共享冲突/中断不会留半写 XLSX（C-5）。
    tmp_path = review_path.parent / (review_path.name + f".{os.getpid()}.tmp.xlsx")
    try:
        wb.save(tmp_path)
        atomic_replace_with_retry(tmp_path, review_path)
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
    print(f"[人工复核] 已导出更适合 Excel/WPS 的 XLSX: {review_path}")


def import_review_xlsx(json_path: Path, review_path: Path, *, dry_run: bool = False):
    """把人工复核 XLSX 的 translation 列回写到 JSON。"""
    if dry_run:
        print(f"[dry-run] 跳过 XLSX 回写: {review_path} -> {json_path}")
        return
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("错误: 读取 XLSX 需要 openpyxl，请先运行 python -m pip install openpyxl") from e
    if not review_path.is_file():
        raise SystemExit(f"错误: 找不到人工复核文件: {review_path}")
    data = json.loads(json_path.read_text(encoding="utf-8-sig"))
    by_index = {int(m.get("index")): m for m in data.get("messages", []) if str(m.get("index", "")).isdigit()}
    wb = load_workbook(review_path)
    ws = wb.active
    # 维度预检：损坏/伪造文件可能声明天文数字的维度，max_row 驱动的逐行
    # 遍历会变成小时级假死；超阈值直接判损坏拒绝处理。
    if ws.max_row > 1_000_000 or ws.max_column > 64:
        raise SystemExit("复核表维度异常，疑似损坏文件")
    header = [ws.cell(row=1, column=i).value for i in range(1, 9)]
    required = ["index", "timestamp", "author", "original", "translation"]
    if header[:5] != required:
        # Backward compatible with old 5-column review sheets.
        header5 = [ws.cell(row=1, column=i).value for i in range(1, 6)]
        if header5 != required:
            raise SystemExit(
                "错误: XLSX 表头不匹配，请保持 index/timestamp/author/original/translation 五列"
                "（可选附加 lint_severity/lint_codes/lint_notes）"
            )
    changed = 0
    for row_no in range(2, ws.max_row + 1):
        idx_value = ws.cell(row=row_no, column=1).value
        if idx_value is None:
            continue
        try:
            idx = int(idx_value)
        except ValueError:
            print(f"警告: 第 {row_no} 行 index 非数字，已跳过")
            continue
        if idx not in by_index:
            print(f"警告: 第 {row_no} 行 index={idx} 不存在，已跳过")
            continue
        raw_cell = ws.cell(row=row_no, column=5).value
        translation = str(raw_cell or "").strip()
        # 剥掉导出侧防注入添加的前置单引号，保真往返（只剥我们自己的约定）。
        translation = _strip_formula_quote(translation)
        # Empty cells must not wipe existing non-empty translations on writeback.
        existing = str(by_index[idx].get("translation", "") or "").strip()
        if not translation and existing:
            continue
        if by_index[idx].get("translation") != translation:
            by_index[idx]["translation"] = translation
            changed += 1
    atomic_write_json(json_path, data)
    print(f"\n[人工复核] 已从 XLSX 回写 {changed} 条修改到: {json_path}")


def import_review_tsv(json_path: Path, review_path: Path, *, dry_run: bool = False):
    """把人工复核 TSV 的 translation 列回写到 JSON。"""
    if dry_run:
        print(f"[dry-run] 跳过 TSV 回写: {review_path} -> {json_path}")
        return
    if not review_path.is_file():
        raise SystemExit(f"错误: 找不到人工复核文件: {review_path}")
    data = json.loads(json_path.read_text(encoding="utf-8-sig"))
    by_index = {int(m.get("index")): m for m in data.get("messages", []) if str(m.get("index", "")).isdigit()}
    lines = review_path.read_text(encoding="utf-8-sig").splitlines()
    if not lines:
        raise SystemExit(f"错误: 人工复核文件为空: {review_path}")
    header = lines[0].split("\t")
    if len(header) < 5 or header[:5] != ["index", "timestamp", "author", "original", "translation"]:
        raise SystemExit(
            "错误: TSV 表头不匹配，请保持 index/timestamp/author/original/translation 五列"
            "（可选附加 lint_severity/lint_codes/lint_notes）"
        )
    changed = 0
    for line_no, line in enumerate(lines[1:], start=2):
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            print(f"警告: 第 {line_no} 行列数不足，已跳过")
            continue
        try:
            idx = int(parts[0])
        except ValueError:
            print(f"警告: 第 {line_no} 行 index 非数字，已跳过")
            continue
        if idx not in by_index:
            print(f"警告: 第 {line_no} 行 index={idx} 不存在，已跳过")
            continue
        # TSV 导出同样加了防注入 ' 前缀；回写时与 XLSX 侧对称剥掉。
        translation = _strip_formula_quote(parts[4].strip())
        # Empty cells must not wipe existing non-empty translations on writeback.
        existing = str(by_index[idx].get("translation", "") or "").strip()
        if not translation and existing:
            continue
        if by_index[idx].get("translation") != translation:
            by_index[idx]["translation"] = translation
            changed += 1
    atomic_write_json(json_path, data)
    print(f"\n[人工复核] 已从 TSV 回写 {changed} 条修改到: {json_path}")


LINT_URL_RE = re.compile(r"https?://\S+")
LINT_MENTION_RE = re.compile(r"@[A-Za-z0-9_]+")
LINT_BRACKET_TOKEN_RE = re.compile(r"\[[^\]]+\]")
# 消歧空白：旧写法 (?:\s*\[[^\]]+\]\s*)+ 的 \s* 与 [^\]]+（可吞空格）存在歧义，
# 对"尾部未闭合 token"的输入会灾难性回溯（n=26 约 5 秒、n=30 约 81 秒）。
# 改为 token 之间必须出现显式分隔空白，回溯线性，语义不变（纯 [emote] 序列，
# 含多空格分隔与首尾空白）。注意：此处与 translate_chat_openai 的 PURE_PRESERVE_RE
# 规则不同，不要互相照抄。
LINT_PURE_EMOTE_RE = re.compile(r"^\s*\[[^\]]+\](?:\s+\[[^\]]+\])*\s*$")


def _lint_issue(issues, idx, code, message, severity="WARN", original="", translation=""):
    issues.append({
        "index": idx,
        "severity": severity,
        "code": code,
        "message": message,
        "original": original,
        "translation": translation,
    })


def lint_translation(
    json_path: Path,
    report_path: Path | None = None,
    max_chars: int = 90,
    max_ratio: float = 2.8,
    data: dict | None = None,
    dry_run: bool = False,
):
    """检查翻译 JSON 中的常见可疑问题，返回 issue 列表。

    data: 调用方已解析好的翻译 JSON（复核表导出等场景避免同一文件反复读盘解析）；
    缺省 None 时从 json_path 读取。对外位置签名保持兼容。
    """
    if not json_path.is_file():
        raise SystemExit(f"错误: 翻译 JSON 不存在: {json_path}")
    if data is None:
        try:
            data = json.loads(json_path.read_text(encoding="utf-8-sig"))
        except json.JSONDecodeError as e:
            raise SystemExit(f"错误: JSON 解析失败: {json_path}: {e}")

    issues = []
    messages = data.get("messages")
    if not isinstance(messages, list):
        _lint_issue(issues, "", "schema_missing_messages", "顶层字段 messages 缺失或不是数组", "FAIL")
        messages = []

    seen_indexes = set()
    for pos, msg in enumerate(messages):
        if not isinstance(msg, dict):
            _lint_issue(issues, pos, "schema_message_not_object", "消息条目不是对象", "FAIL")
            continue

        idx = msg.get("index", pos)
        original = str(msg.get("original", ""))
        translation = str(msg.get("translation", "")) if msg.get("translation") is not None else ""
        original_s = original.strip()
        translation_s = translation.strip()

        for required in ["index", "original", "translation"]:
            if required not in msg:
                _lint_issue(issues, idx, "schema_missing_field", f"缺少字段: {required}", "FAIL", original, translation)

        if idx in seen_indexes:
            _lint_issue(issues, idx, "duplicate_index", "index 重复", "FAIL", original, translation)
        seen_indexes.add(idx)

        if not translation_s:
            _lint_issue(issues, idx, "empty_translation", "translation 为空", "FAIL", original, translation)
            continue

        is_pure_emote = bool(LINT_PURE_EMOTE_RE.fullmatch(original_s))
        if is_pure_emote and translation_s != original_s:
            _lint_issue(issues, idx, "pure_emote_changed", "纯 emote 消息应保持原样", "WARN", original, translation)

        original_mentions = set(LINT_MENTION_RE.findall(original))
        missing_mentions = sorted(m for m in original_mentions if m not in translation)
        if missing_mentions:
            _lint_issue(issues, idx, "mention_lost", "翻译丢失 @用户名: " + ", ".join(missing_mentions), "WARN", original, translation)

        original_urls = set(LINT_URL_RE.findall(original))
        missing_urls = sorted(u for u in original_urls if u not in translation)
        if missing_urls:
            _lint_issue(issues, idx, "url_lost", "翻译丢失 URL: " + ", ".join(missing_urls), "WARN", original, translation)

        original_brackets = set(LINT_BRACKET_TOKEN_RE.findall(original))
        missing_brackets = sorted(b for b in original_brackets if b not in translation)
        if missing_brackets:
            _lint_issue(issues, idx, "bracket_token_lost", "翻译丢失方括号 token/emote: " + ", ".join(missing_brackets), "WARN", original, translation)

        if not is_pure_emote and len(translation_s) > max_chars:
            _lint_issue(issues, idx, "too_long", f"翻译超过 {max_chars} 字，可能不适合弹幕显示", "WARN", original, translation)
        elif not is_pure_emote and original_s and len(translation_s) / max(1, len(original_s)) > max_ratio and len(translation_s) > 24:
            _lint_issue(issues, idx, "expansion_ratio_high", f"翻译长度超过原文 {max_ratio:.1f} 倍", "WARN", original, translation)

    fail_count = sum(1 for issue in issues if issue["severity"] == "FAIL")
    warn_count = sum(1 for issue in issues if issue["severity"] == "WARN")
    print(f"\n[翻译质检] 文件: {json_path}")
    print(f"  消息数: {len(messages)}")
    print(f"  FAIL: {fail_count}, WARN: {warn_count}")

    if issues:
        for issue in issues[:80]:
            print(f"  [{issue['severity']}] #{issue['index']} {issue['code']}: {issue['message']}")
        if len(issues) > 80:
            print(f"  ... 还有 {len(issues) - 80} 条未显示")
    else:
        print("  未发现确定性规则问题。")

    if report_path:
        if dry_run:
            # dry-run 不落盘：质检报告写出统一在写侧拦截（lint 计算本身只读，照常执行）。
            print(f"  [dry-run] 跳过质检报告写出: {report_path}")
        else:
            lines = ["index\tseverity\tcode\tmessage\toriginal\ttranslation"]
            for issue in issues:
                row = [
                    str(issue["index"]),
                    issue["severity"],
                    issue["code"],
                    issue["message"],
                    str(issue.get("original", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
                    str(issue.get("translation", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
                ]
                lines.append("\t".join(row))
            # 原子写：同目录唯一 tmp + retry replace，报告不因共享冲突/中断半写（C-5）。
            fd, tmp_name = tempfile.mkstemp(
                prefix=f".{report_path.name}.", suffix=".tmp", dir=str(report_path.parent)
            )
            tmp_path = Path(tmp_name)
            try:
                with os.fdopen(fd, "w", encoding="utf-8-sig") as file:
                    file.write("\n".join(lines) + "\n")
                atomic_replace_with_retry(tmp_path, report_path)
            finally:
                try:
                    tmp_path.unlink(missing_ok=True)
                except OSError:
                    pass
            print(f"  质检报告: {report_path}")

    return issues


# ---------------------------------------------------------------------------
# 刀五迁入（自 render_cn_chat.py 原样搬运）：YAML 规则/Profile 解析、
# 渲染产物发布、翻译 JSON 规则清洗。除 normalize_translation 的 DRY_RUN
# 改为显式 dry_run 参数注入外，函数体一概 verbatim。
# ---------------------------------------------------------------------------


def load_yaml_file(yaml_path: Path, label: str):
    try:
        import yaml
    except ImportError as e:
        # PipelineError subclasses SystemExit, so CLI exit codes are unchanged
        # while callers can catch every failure uniformly.
        raise PipelineError(
            f"错误: 使用 {label} 需要安装 PyYAML，请运行 pip install PyYAML"
        ) from e
    if not yaml_path.is_file():
        raise PipelineError(f"错误: {label} 文件不存在: {yaml_path}")
    try:
        data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, yaml.YAMLError) as e:
        raise PipelineError(f"Invalid {label} YAML {yaml_path}: {e}") from e
    if data is None:
        return {}
    if not isinstance(data, dict):
        raise PipelineError(f"Invalid {label} YAML {yaml_path}: root must be a mapping")
    return data


def load_yaml_rules(rules_path: Path):
    """Load normalizations + optional preserve_patterns from a rules YAML.

    Returns a dict: {"normalizations": [...], "preserve_patterns": [compiled re, ...]}.
    preserve_patterns skip rule application when original matches (translate path
    still runs separately; this only protects the rules-normalize pass).
    """
    data = load_yaml_file(rules_path, "规则")
    rules = []
    if "normalizations" not in data:
        raw_rules = []
    else:
        raw_rules = data["normalizations"]
    if not isinstance(raw_rules, list):
        raise PipelineError(
            f"Invalid rules YAML {rules_path}: normalizations must be a list"
        )
    for rule_index, item in enumerate(raw_rules):
        if not isinstance(item, dict):
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: normalizations[{rule_index}] must be a mapping"
            )
        targets = item.get("match", [])
        if isinstance(targets, str):
            targets = [targets]
        elif not isinstance(targets, list):
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: normalizations[{rule_index}].match must be a string or list"
            )
        if not all(isinstance(target, (str, int, float)) for target in targets):
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: normalizations[{rule_index}].match contains a non-scalar value"
            )
        translation = item.get("translation")
        if translation is None:
            continue
        rules.append({
            "name": item.get("name", "unnamed"),
            "match": {str(x) for x in targets},
            "translation": str(translation),
        })
    preserve = []
    preserve_raw = data.get("preserve_patterns")
    if preserve_raw is None:
        preserve_raw = []
    if not isinstance(preserve_raw, list):
        raise PipelineError(
            f"Invalid rules YAML {rules_path}: preserve_patterns must be a list"
        )
    for pattern_index, pat in enumerate(preserve_raw):
        text = str(pat)
        # 灾难性回溯防护:超长 pattern(嵌套量词类)编译/搜索代价不可控,
        # rules YAML 属用户自担配置,但仍在编译入口设硬上限并明确报错。
        if len(text) > 500:
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: preserve_patterns[{pattern_index}] "
                f"超过 {500} 字符上限({len(text)}),拒绝编译(疑似灾难性回溯风险)"
            )
        try:
            preserve.append(re.compile(text))
        except re.error as e:
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: preserve_patterns[{pattern_index}] is not a valid regex: {e}"
            ) from e
    return {"normalizations": rules, "preserve_patterns": preserve}


def publish_output(src_path: Path, dst_path: Path, *, backup_prev: bool = True):
    """Copy rendered output to the final path using a temp file + atomic replace.

    When backup_prev is True (default), rename an existing dst to dst.bak first and
    restore it if the replace fails — matching burn's default backup contract.
    """
    src_path = Path(src_path)
    dst_path = Path(dst_path)
    if not src_path.is_file():
        raise PipelineError(f"错误: 渲染输出不存在: {src_path}")
    if src_path.resolve() == dst_path.resolve():
        return dst_path
    dst_path.parent.mkdir(parents=True, exist_ok=True)
    backup = None
    backup_created = False
    if backup_prev and dst_path.is_file():
        backup = Path(str(dst_path) + ".bak")
        try:
            if backup.is_file():
                backup.unlink()
            dst_path.rename(backup)
            backup_created = True
            print(f"  [backup] {backup}")
        except OSError as e:
            print(f"  warning: cannot backup {dst_path}: {e}")
            backup = None
            backup_created = False
    fd, tmp_name = tempfile.mkstemp(prefix=dst_path.stem + ".", suffix=".partial.mp4", dir=str(dst_path.parent))
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        try:
            shutil.copy2(src_path, tmp_path)
            atomic_replace_with_retry(tmp_path, dst_path)
        except OSError:
            if backup_created and backup is not None and backup.is_file() and not dst_path.is_file():
                try:
                    backup.rename(dst_path)
                    print(f"  已从备份恢复: {dst_path}")
                except OSError as restore_err:
                    print(f"  警告: 无法从备份恢复 {backup}: {restore_err}")
            raise
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except OSError:
            pass
    try:
        src_path.unlink()
    except OSError:
        pass
    return dst_path


def normalize_translation(json_path: Path, rules_path: Path | None = None, *, dry_run: bool = False):
    """规则清洗（原 render_cn_chat.normalize_translation，原样搬运）。

    dry_run: 由调用方（render_cn_chat 薄包装）注入模块全局 DRY_RUN，
    保持测试按模块属性 monkeypatch DRY_RUN 的语义；本模块自身不读全局。
    """
    if not rules_path:
        print("\n[规则清洗] 未指定 --rules，跳过规则清洗。")
        return
    if dry_run:
        print(f"\n[dry-run] 跳过规则清洗写入: {json_path}")
        return
    loaded = load_yaml_rules(rules_path)
    rules = loaded.get("normalizations") or []
    preserve_patterns = loaded.get("preserve_patterns") or []
    if not rules:
        print(f"\n[规则清洗] 规则文件无 normalizations: {rules_path}")
        return
    data = json.loads(json_path.read_text(encoding="utf-8-sig"))
    changed = []
    for msg in data.get("messages", []):
        original = str(msg.get("original", ""))
        if any(p.search(original) for p in preserve_patterns):
            continue
        for rule in rules:
            if original in rule["match"] and msg.get("translation") != rule["translation"]:
                changed.append((msg.get("index"), rule["name"], original, msg.get("translation"), rule["translation"]))
                msg["translation"] = rule["translation"]
                break
    atomic_write_json(json_path, data)
    if changed:
        print(f"\n[规则清洗] 已应用 {len(changed)} 条修改，规则文件: {rules_path}")
        for idx, rule_name, original, old, new in changed:
            print(f"  [{idx}] {rule_name}: {original!r}: {old!r} -> {new!r}")
    else:
        print(f"\n[规则清洗] 无需修改，规则文件: {rules_path}")


def load_profile(profile_path: Path):
    data = load_yaml_file(profile_path, "Profile")
    glossary_value = data.get("glossary")
    if glossary_value is not None and not isinstance(glossary_value, dict):
        raise PipelineError(
            f"Invalid Profile YAML {profile_path}: glossary must be a mapping"
        )
    preserve_value = data.get("preserve")
    if preserve_value is not None and not isinstance(preserve_value, list):
        raise PipelineError(
            f"Invalid Profile YAML {profile_path}: preserve must be a list"
        )
    style_value = data.get("translation_style")
    if style_value is not None and not isinstance(style_value, dict):
        raise PipelineError(
            f"Invalid Profile YAML {profile_path}: translation_style must be a mapping"
        )
    context_parts = []
    if data.get("context"):
        context_parts.append(str(data["context"]))

    glossary = data.get("glossary") or {}
    if glossary:
        terms = []
        for src, dst in glossary.items():
            terms.append(f"  {src} -> {dst}")
        context_parts.append(
            "**术语词典 / Glossary (必须严格遵守 / MUST follow strictly)**\n"
            + "\n".join(terms)
        )

    preserve = data.get("preserve") or []
    if preserve:
        context_parts.append("需要保留 / Preserve: " + ", ".join(map(str, preserve)))

    style = data.get("translation_style") or {}
    if style:
        style_lines = [f"{k}: {v}" for k, v in style.items()]
        context_parts.append("翻译风格 / Translation style:\n" + "\n".join(style_lines))

    return "\n\n".join(context_parts), data
