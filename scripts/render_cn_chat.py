#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一键生成翻译后的 Twitch 聊天覆盖视频
================================

输入：源视频 + Twitch 原始/导出的聊天 HTML
输出：带翻译后 chat overlay 的视频（默认 <视频名>_chat.mp4）

流程：
1. twitch_chat_burn.py --export-translation 导出待翻译 JSON
2. translate_chat_openai.py 使用 OpenAI-compatible 接口并发翻译 JSON
3. 可选 YAML 规则清洗（例如频道梗、术语替换）
4. twitch_chat_burn.py --import-translation 渲染并合成视频

示例：
  python render_cn_chat.py --init
  python render_cn_chat.py --doctor
  python render_cn_chat.py --job jobs/example_job.yaml
  python render_cn_chat.py video.mp4 chat.html --mode preview --render-original
  python render_cn_chat.py video.mp4 chat.html --reuse-translation --rules configs/rules.example.yaml
  python render_cn_chat.py video.mp4 chat.html --profile profiles/default.yaml
  python render_cn_chat.py video.mp4 chat.html --preview-frame 60 --preview-image preview.png
"""

import argparse
import importlib.util
import json
import os
from pathlib import Path
import platform
import re
import shutil
import subprocess
import sys
import tempfile

# Allow sibling imports when loaded as a script or via importlib from tests.
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from chat_window import apply_preview_first_defaults
from common_utils import (
    current_cli_invocation,
    current_cli_script,
    detect_cjk_font,
    ensure_utf8_stdio,
    load_dotenv_if_present,
    positive_float_arg,
    quote_cli_arg,
    require_executable,
    resolve_font_paths,
    resolve_public_resource,
    safe_which,
    validate_positive_float,
)
from env_bootstrap import (
    maybe_prompt_offer_fixes,
    maybe_prompt_offer_td_cli,
    offer_fixes,
    offer_td_cli_guide,
    prepend_tools_ffmpeg_to_path,
    print_readiness_report,
    probe_translate_api,
)
from job_config import (
    apply_job_to_namespace,
    load_job_file,
    save_last_job,
    validate_job_media_paths,
)
from job_wizard import run_job_wizard, run_list_jobs
from layout_preset import apply_layout_preset_to_namespace, load_layout_preset
from media_health import validate_media_health
from process_util import (
    clean_companion_flags_error,
    clean_temp_artifacts,
    install_process_cleanup_handlers,
    is_dangerous_publish_path,
    run_tracked,
)
from render_preset import apply_render_preset_to_namespace, load_render_preset
from task_events import emit_task_event
from task_results import write_task_result
from ux_setup import print_setup_next_steps, run_init

ensure_utf8_stdio()

_TASK_STAGE_BY_PROGRAM = {
    "translate_chat_openai.py": "translate",
    "twitch_chat_burn.py": "render",
}
load_dotenv_if_present()


# ---------------------------------------------------------------------------
# Dual-CLI flag forwarding (pipeline → twitch_chat_burn)
#
# Shared flags live in the tables below and are applied via append_* helpers.
# When adding a new shared burn flag: add (attr, flag, kind) to the right
# table, expose it on the pipeline argparse, and the contract tests in
# tests/test_cli_flag_forward.py will catch missing forwards.
#
# kind:
#   always      – always emit ``flag value`` (attr must exist)
#   opt         – emit if attr is present and not None/""
#   opt_truthy  – emit if attr is truthy
#   flag        – emit bare flag if attr is truthy
# ---------------------------------------------------------------------------

# Burn CLI options that are *not* general pipeline flags (path/mode specific).
# Pipeline may still pass some of these at call sites (e.g. --import-translation
# on render, --out-dir when --workdir is set) but they are not part of the
# shared append_* tables.
BURN_ONLY_FLAGS: tuple[str, ...] = (
    "export-translation",
    "import-translation",
    "force-export",
    "strict-import",  # pipeline has a thin forward when importing (see append_strict_import_arg)
    "job-dir",
    "no-job-dir",
    "out-dir",
)

# (attr, flag, kind) — chat-layer / final fps
FPS_FORWARD_SPECS: tuple[tuple[str, str, str], ...] = (
    ("fps", "--fps", "always"),
    ("output_fps", "--output-fps", "opt"),
)

# (attr, flag, kind) — layout / stack / ratios (layout-preset applied onto args first)
LAYOUT_FORWARD_SPECS: tuple[tuple[str, str, str], ...] = (
    ("max_visible", "--max-visible", "opt"),
    ("msg_lifetime", "--msg-lifetime", "opt"),
    ("max_message_lines", "--max-message-lines", "opt"),
    ("min_visible_seconds", "--min-visible-seconds", "opt"),
    ("arrival_interval", "--arrival-interval", "opt"),
    ("stack_mode", "--stack-mode", "opt"),
    ("x_ratio", "--x-ratio", "opt"),
    ("y_ratio", "--y-ratio", "opt"),
    ("width_ratio", "--width-ratio", "opt"),
    ("height_ratio", "--height-ratio", "opt"),
    ("font_size_ratio", "--font-size-ratio", "opt"),
    ("emote_height", "--emote-height", "opt"),
    ("lazy_message_images", "--lazy-message-images", "flag"),
)

# (attr, flag, kind) — encode / static-frame reuse
PERF_FORWARD_SPECS: tuple[tuple[str, str, str], ...] = (
    ("encoder", "--encoder", "always"),
    ("video_preset", "--video-preset", "opt_truthy"),
    ("crf", "--crf", "always"),
    ("video_bitrate", "--video-bitrate", "opt_truthy"),
    ("maxrate", "--maxrate", "opt_truthy"),
    ("bufsize", "--bufsize", "opt_truthy"),
    ("audio_codec", "--audio-codec", "always"),
    ("audio_bitrate", "--audio-bitrate", "always"),
    ("overlay_codec", "--overlay-codec", "always"),
    ("webm_crf", "--webm-crf", "always"),
    ("webm_cpu_used", "--webm-cpu-used", "always"),
    ("no_reuse_static_frames", "--no-reuse-static-frames", "flag"),
    ("no_skip_blank_frames", "--no-skip-blank-frames", "flag"),
    ("blank_hold_seconds", "--blank-hold-seconds", "always"),
)

# Flat list of every shared forward flag (for contract tests / docs).
SHARED_FORWARD_FLAGS: tuple[str, ...] = tuple(
    flag
    for _attr, flag, _kind in (
        *FPS_FORWARD_SPECS,
        *LAYOUT_FORWARD_SPECS,
        *PERF_FORWARD_SPECS,
    )
) + (
    # companion of lazy_message_images (emitted only when lazy is on)
    "--message-image-cache-size",
)


def _append_flag_specs(cmd: list, args, specs: tuple[tuple[str, str, str], ...]) -> list:
    """Apply a table of (attr, flag, kind) to *cmd* from *args*."""
    for attr, flag, kind in specs:
        if kind == "always":
            cmd.extend([flag, str(getattr(args, attr))])
            continue
        if not hasattr(args, attr):
            continue
        val = getattr(args, attr)
        if kind == "opt":
            if val is not None and val != "":
                cmd.extend([flag, str(val)])
        elif kind == "opt_truthy":
            if val:
                cmd.extend([flag, str(val)])
        elif kind == "flag":
            if val:
                cmd.append(flag)
        else:
            raise ValueError(f"unknown flag-forward kind: {kind!r} for {attr}")
    return cmd


def append_fps_args(cmd, args):
    """Forward chat-layer fps and optional final output fps."""
    return _append_flag_specs(cmd, args, FPS_FORWARD_SPECS)


def append_layout_burn_args(cmd: list, args) -> list:
    """Forward layout / lazy-memory flags to twitch_chat_burn.py."""
    _append_flag_specs(cmd, args, LAYOUT_FORWARD_SPECS)
    # cache size only meaningful with lazy mode (matches prior behavior)
    if getattr(args, "lazy_message_images", False):
        cmd.extend(
            [
                "--message-image-cache-size",
                str(getattr(args, "message_image_cache_size", 256)),
            ]
        )
    # layout-preset already applied onto args; no need to forward YAML path
    return cmd


def append_perf_encode_args(cmd: list, args) -> list:
    """Forward performance/encode flags to twitch_chat_burn.py."""
    return _append_flag_specs(cmd, args, PERF_FORWARD_SPECS)


def append_strict_import_arg(cmd: list, args) -> list:
    """Forward --strict-import when pipeline is driving burn --import-translation."""
    if getattr(args, "strict_import", False):
        cmd.append("--strict-import")
    return cmd


def append_shared_burn_args(cmd: list, args) -> list:
    """Forward all shared fps/layout/encode flags (not burn-only path flags)."""
    _append_flag_specs(cmd, args, FPS_FORWARD_SPECS)
    append_layout_burn_args(cmd, args)
    _append_flag_specs(cmd, args, PERF_FORWARD_SPECS)
    return cmd


DRY_RUN = False
VERBOSE = False
QUIET = False
_TASK_RESULT_CONTEXT: dict[str, object] = {"mode": "unknown", "artifacts": []}


class PipelineError(SystemExit):
    pass


def mark_manual_translation_required() -> None:
    """Record that a requested translated task stopped for human input."""
    _TASK_RESULT_CONTEXT["terminal_state"] = "manual_required"


def validate_source_media(video: Path, *, mode: str, dry_run: bool = False) -> None:
    """Fail before translation when the local input cannot be decoded safely."""
    selected_mode = str(mode or "fast").lower()
    if selected_mode == "off":
        log("[media] 输入视频健康检查已关闭。")
        emit_task_event("stage_skipped", stage="source_media_check", reason="disabled", completed=0, total=1)
        return
    if dry_run:
        log(f"[dry-run] 跳过输入视频健康检查（{selected_mode}）。")
        emit_task_event("stage_skipped", stage="source_media_check", reason="dry_run", completed=0, total=1)
        return

    label = "完整解码" if selected_mode == "decode" else "快速"
    log(f"[media] {label}检查输入视频，发现问题会在翻译/渲染前停止…")
    emit_task_event("stage_started", stage="source_media_check", completed=0, total=1)
    # Local workflows may legitimately use silent video. Validate its video
    # stream and any present audio stream without turning silence into failure.
    health = validate_media_health(video, mode=selected_mode, require_audio=False)
    if not health.ok:
        emit_task_event("stage_failed", stage="source_media_check", completed=0, total=1)
        raise PipelineError(
            "错误: 输入视频健康检查失败，已在翻译或渲染前停止。\n"
            f"  详情: {health.reason()}\n"
            "  建议: 重新下载有问题的片段，或使用 --source-media-check fast 进行快速复查。"
        )
    for warning in health.warnings:
        log(f"[media] 提示: {warning}")
    emit_task_event("stage_completed", stage="source_media_check", completed=1, total=1)


def log(msg, level="info"):
    if QUIET and level == "info":
        return
    if VERBOSE or level != "debug":
        print(msg, flush=True)


def _stdin_is_interactive() -> bool:
    """True when we can pause for the user (real TTY)."""
    try:
        if sys.stdin is None or not sys.stdin.isatty():
            return False
    except Exception:
        return False
    try:
        name = getattr(sys.stdin, "name", "") or ""
        if name in ("nul", "NUL", "/dev/null"):
            return False
    except Exception:
        pass
    return True


def _render_preview_clip(
    *,
    video: Path,
    chat_html: Path,
    trans_json: Path,
    args,
    workdir: Path | None,
    seconds: float,
    burn: Path,
) -> Path | None:
    """Render a short preview clip with translated chat overlay. Returns output path or None on failure."""
    preview_dir = (workdir / "temp") if workdir else Path("outputs") / "_preview"
    if is_dangerous_publish_path(preview_dir):
        print(f"  [FAIL] 预览目录在系统路径下，已拒绝: {preview_dir}", flush=True)
        return None
    preview_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        sys.executable, str(burn), str(video), str(chat_html),
        "--x", str(args.x), "--y", str(args.y),
        "--w", str(args.width), "--h", str(args.height),
        "--font-size", str(args.font_size),
        "--font-path", args.font_path,
        "--font-bold-path", args.font_bold_path,
        "--bg-alpha", str(args.bg_alpha),
        "--import-translation", str(trans_json),
        "--preview-clip", str(seconds),
        "--out-dir", str(preview_dir),
    ]
    append_strict_import_arg(cmd, args)
    append_shared_burn_args(cmd, args)
    if args.offset is not None:
        cmd.extend(["--offset", str(args.offset)])
    if getattr(args, "preview_dense", False):
        cmd.append("--preview-dense")

    log(f"\n[预览] 渲染 {seconds}s 预览片段...")
    try:
        run(cmd, error_hint="预览渲染失败")
    except PipelineError as e:
        print(f"  [FAIL] 预览渲染失败: {e}", flush=True)
        return None

    # burn compose names preview clips as <stem>_chat.mp4 (same as full burns);
    # also accept any *_preview_*s.mp4 if naming changes later.
    candidates = list(preview_dir.glob(f"{video.stem}_chat.mp4"))
    if not candidates:
        candidates = list(preview_dir.glob(f"{video.stem}_preview_*s.mp4"))
    if not candidates:
        # Job-dir layout: out_dir may contain job_*/<stem>_chat.mp4
        candidates = list(preview_dir.glob(f"**/job_*/{video.stem}_chat.mp4"))
    if candidates:
        # Prefer newest if multiple
        preview_out = max(candidates, key=lambda p: p.stat().st_mtime)
    else:
        print(f"  [WARN] 未找到预览输出文件（期望 {video.stem}_chat.mp4）", flush=True)
        return None
    log(f"[预览] 已生成: {preview_out}")
    # Best-effort open the preview (Windows).
    if os.name == "nt":
        try:
            os.startfile(str(preview_out))
        except OSError:
            pass
    return preview_out


def pause_after_translation_for_review(
    *,
    trans_json: Path,
    review_xlsx: Path,
    review_tsv: Path,
    auto_continue: bool = False,
    # Preview support
    video: Path | None = None,
    chat_html: Path | None = None,
    args=None,
    workdir: Path | None = None,
    burn: Path | None = None,
) -> str:
    """After API/rules translation: export Excel and wait for Enter before render.

    Returns:
      "continue" — proceed to render (optionally after user edited XLSX; caller may re-import)
      "stop" — user chose to stop here (same spirit as --review)
    Non-interactive / --yes / dry-run: prints paths and continues without blocking.
    """
    # Always refresh review tables so user has something to open.
    try:
        export_review_tsv(trans_json, review_tsv)
        export_review_xlsx(trans_json, review_xlsx)
    except Exception as e:
        log(f"[WARN] 导出复核表失败（仍可继续渲染）: {e}")

    print("\n======== 翻译已完成 · 渲染前确认 ========", flush=True)
    print(f"  翻译 JSON : {trans_json}", flush=True)
    if review_xlsx.is_file():
        print(f"  Excel 复核: {review_xlsx}", flush=True)
        print("  （请打开 XLSX，检查/修改最后一列 translation）", flush=True)
    if review_tsv.is_file():
        print(f"  TSV 备份  : {review_tsv}", flush=True)
    print("----------------------------------------", flush=True)

    if auto_continue or DRY_RUN or not _stdin_is_interactive():
        if not auto_continue and not DRY_RUN:
            print("  （非交互终端：自动继续渲染。交互运行时会在此等待回车。）", flush=True)
        else:
            print("  （--yes / dry-run：不暂停，继续渲染）", flush=True)
        return "continue"

    # Best-effort open Excel for the user (Windows).
    if review_xlsx.is_file() and os.name == "nt":
        try:
            os.startfile(str(review_xlsx))  # type: ignore[attr-defined]
            print("  已尝试用默认程序打开 Excel。", flush=True)
        except OSError:
            pass

    # workdir is optional: without it, _render_preview_clip writes under outputs/_preview.
    can_preview = all(v is not None for v in (video, chat_html, args, burn))
    while True:
        print("  回车 = 继续渲染（若改过 XLSX 会先自动回写）", flush=True)
        if can_preview:
            print("  P    = 先渲染一小段预览片段（默认 10 秒；无 --workdir 时写到 outputs/_preview）", flush=True)
            print("  P 30 = 渲染 30 秒预览片段", flush=True)
        print("  S    = 先停在这里，稍后用 --review-done 再渲染", flush=True)
        try:
            raw = input("请选择 [回车继续" + (" / P 预览" if can_preview else "") + " / S 停止]: ").strip()
        except EOFError:
            raw = ""

        low = raw.lower()

        if low in ("s", "stop", "q", "quit"):
            print("\n[OK] 已暂停。改完 Excel 后可用：", flush=True)
            resume_hint = (
                f"{current_cli_invocation()} "
                f"{quote_cli_arg(video or 'video.mp4')} "
                f"{quote_cli_arg(chat_html or 'chat.html')} "
                f"--reuse-translation --review-done "
                f"--translation-json {quote_cli_arg(trans_json)} "
                f"--review-xlsx {quote_cli_arg(review_xlsx)}"
            )
            if workdir is not None:
                resume_hint += f" --workdir {quote_cli_arg(workdir)}"
            output = getattr(args, "output", None)
            if output:
                resume_hint += f" --output {quote_cli_arg(output)}"
            print(f"  {resume_hint}", flush=True)
            return "stop"

        parts = raw.split(None, 1)
        if can_preview and parts and parts[0].lower() == "p":
            # "P" -> 10s; "P 30" -> 30s
            seconds = 10.0
            if len(parts) == 2:
                try:
                    seconds = validate_positive_float(
                        "preview seconds", float(parts[1]), maximum=3600.0
                    )
                except (TypeError, ValueError):
                    print("  秒数无效（须在 0 到 3600 之间），使用默认 10 秒", flush=True)
                    seconds = 10.0
            preview_path = _render_preview_clip(
                video=video,
                chat_html=chat_html,
                trans_json=trans_json,
                args=args,
                workdir=workdir,
                seconds=seconds,
                burn=burn,
            )
            if preview_path is not None:
                print(f"  预览已生成: {preview_path}", flush=True)
                print("  请检查预览效果，然后回车继续渲染或 S 停止。", flush=True)
            continue

        # Empty / unknown -> continue to render
        return "continue"


def run(cmd, cwd=None, error_hint=""):
    launcher = Path(str(cmd[0])).stem.lower()
    program_arg = cmd[1] if len(cmd) > 1 and launcher.startswith("python") else cmd[0]
    program = Path(str(program_arg)).name
    stage = _TASK_STAGE_BY_PROGRAM.get(program)
    if DRY_RUN:
        log(f"[dry-run] {' '.join(str(c) for c in cmd)}")
        emit_task_event("command_skipped", program=program, reason="dry_run")
        if stage:
            emit_task_event("stage_skipped", stage=stage, reason="dry_run", completed=0, total=1)
        return
    log("\n$ " + " ".join(f'"{c}"' if " " in str(c) else str(c) for c in cmd))
    emit_task_event("command_started", program=program)
    if stage:
        emit_task_event("stage_started", stage=stage, completed=0, total=1)
    env = os.environ.copy()
    env.setdefault("PYTHONIOENCODING", "utf-8")
    try:
        # Inherit stdio (None) so child progress remains visible, but still track
        # the process tree for Ctrl+C / atexit cleanup.
        p = run_tracked(cmd, cwd=cwd, text=False, env=env, stdout=None, stderr=None)
    except FileNotFoundError as e:
        emit_task_event("command_failed", program=program, reason="not_found")
        if stage:
            emit_task_event("stage_failed", stage=stage, reason="not_found", completed=0, total=1)
        hint = error_hint or "找不到可执行文件，请确认已安装并加入 PATH"
        raise PipelineError(f"错误: {hint}\n  详情: {e}")
    emit_task_event("command_exited", program=program, returncode=p.returncode)
    if stage:
        emit_task_event(
            "stage_completed" if p.returncode == 0 else "stage_failed",
            stage=stage,
            completed=1 if p.returncode == 0 else 0,
            total=1,
        )
    if p.returncode != 0:
        hint = error_hint or "命令执行失败"
        raise PipelineError(f"错误: {hint} (exit code {p.returncode})")


def _translation_nonempty_count(path: Path) -> int:
    """Count non-empty translation fields in a translation JSON (0 if missing)."""
    if not path.is_file():
        return 0
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return 0
    items = data.get("messages") if isinstance(data, dict) else None
    if not isinstance(items, list):
        return 0
    return sum(1 for it in items if isinstance(it, dict) and str(it.get("translation", "") or "").strip())


def _post_download_next_steps(video: Path, chat_html: Path, *, download_only: bool, yes: bool) -> int:
    """After assets land: print paths and optionally interactive next-step menu."""
    print("\n======== 下载完成 ========")
    print(f"  视频: {video}")
    print(f"  聊天: {chat_html}")
    print("  下一步示例:")
    print(
        f"    {current_cli_invocation()} {quote_cli_arg(video)} {quote_cli_arg(chat_html)} "
        f"--mode preview --render-original --preview-clip 10"
    )
    print(
        f"    {current_cli_invocation()} {quote_cli_arg(video)} {quote_cli_arg(chat_html)} --manual-translation"
    )
    if download_only or yes or not _stdin_is_interactive():
        return 0
    print("\n请选择下一步:")
    print("  [1] 预览短片（原文 10s）")
    print("  [2] 导出人工翻译表")
    print("  [3] 翻译出片（API 可用则自动译）")
    print("  [0] 结束（仅保留已下载文件）")
    try:
        choice = input("请选择 [0-3] (默认 1): ").strip() or "1"
    except EOFError:
        return 0
    if choice in ("0", "q", "quit"):
        return 0
    if choice == "2":
        return _run_pipeline_with_media(
            video,
            chat_html,
            "--manual-translation",
            "--yes",
        )
    if choice == "3":
        return _run_pipeline_with_media(video, chat_html, "--mode", "full", "--yes")
    # default preview
    return _run_pipeline_with_media(
        video,
        chat_html,
        "--mode",
        "preview",
        "--render-original",
        "--preview-clip",
        "10",
        "--yes",
    )


def _run_pipeline_with_media(video: Path, chat_html: Path, *extra: str) -> int:
    """Re-enter this script with local media (same interpreter)."""
    cmd = [
        sys.executable,
        str(Path(__file__).resolve()),
        str(video),
        str(chat_html),
        *extra,
    ]
    print("\n$ " + " ".join(f'"{c}"' if " " in str(c) else str(c) for c in cmd), flush=True)
    try:
        p = run_tracked(cmd, stdout=None, stderr=None, text=False)
        return int(p.returncode)
    except Exception as e:
        print(f"[FAIL] 无法继续 pipeline: {e}", flush=True)
        return 1


def _parse_cli_segments(raw_segments) -> list[tuple[str, str]]:
    """Parse --segment BEGIN-END values into (begin, end) pairs."""
    from twitch_download import TwitchDownloadError, parse_segment_line

    pairs: list[tuple[str, str]] = []
    for raw in raw_segments or []:
        text = str(raw or "").strip()
        if not text:
            continue
        # Accept "begin-end", "begin end", or "begin,end"
        try:
            seg = parse_segment_line(text)
        except TwitchDownloadError:
            # Rare: hyphen-only form failed earlier heuristics — try space on first '-'
            if "-" in text and " " not in text and "," not in text:
                left, _, right = text.partition("-")
                seg = parse_segment_line(f"{left} {right}")
            else:
                raise
        if seg is None:
            raise TwitchDownloadError(f"无效 --segment: {raw!r}")
        pairs.append((seg.begin, seg.end))
    return pairs


def _run_download_flow(args) -> int:
    """CLI entry for --download: fetch VOD/clip + HTML via TwitchDownloaderCLI."""
    from twitch_download import TwitchDownloadError, download_assets, download_assets_multi

    out_dir = Path(args.download_dir).resolve() if getattr(args, "download_dir", None) else None
    if out_dir is not None and (
        is_dangerous_publish_path(out_dir) or is_dangerous_publish_path(out_dir.parent)
    ):
        print(f"错误: --download-dir 不能是系统目录: {out_dir}", file=sys.stderr)
        return 2

    raw_segments = list(getattr(args, "segment", None) or [])
    # Parse --cut START-END into (start_s, end_s) pairs
    cut_ranges: list[tuple[float, float]] = []
    for raw_cut in (getattr(args, "cut", None) or []):
        text = str(raw_cut or "").strip()
        if not text:
            continue
        # Accept "start-end", "start end", or "start,end"
        try:
            from twitch_download import parse_segment_line
            seg = parse_segment_line(text)
            if seg is None:
                raise TwitchDownloadError(f"无效 --cut: {raw_cut!r}")
            cut_ranges.append((seg.begin_s, seg.end_s))
        except TwitchDownloadError:
            # Try hyphen split for time-like values
            if "-" in text and " " not in text and "," not in text:
                left, _, right = text.partition("-")
                seg = parse_segment_line(f"{left} {right}")
                if seg is not None:
                    cut_ranges.append((seg.begin_s, seg.end_s))
                    continue
            raise

    emit_task_event("stage_started", stage="download", completed=0, total=1)
    try:
        multi = _parse_cli_segments(raw_segments) if raw_segments else []
        if multi and (getattr(args, "begin", None) or getattr(args, "end", None)):
            print(
                "警告: 已指定 --segment，忽略 --begin/--end",
                file=sys.stderr,
            )
        if cut_ranges and not multi:
            print(
                "警告: --cut 仅在 --segment 多段下载时生效，已忽略",
                file=sys.stderr,
            )
            cut_ranges = []
        if multi:
            result = download_assets_multi(
                str(args.download),
                multi,
                out_dir=out_dir,
                kind=str(getattr(args, "kind", "auto") or "auto"),
                quality=getattr(args, "quality", None) or None,
                oauth=getattr(args, "oauth", None),
                remove_ranges=cut_ranges or None,
                output_fps=getattr(args, "download_output_fps", None),
                encoder=str(getattr(args, "download_encoder", "auto") or "auto"),
                trim_mode=str(getattr(args, "download_trim_mode", "Safe") or "Safe"),
                media_check=str(getattr(args, "media_check", "fast") or "fast"),
                media_repair=str(getattr(args, "media_repair", "audio") or "audio"),
            )
        else:
            result = download_assets(
                str(args.download),
                out_dir=out_dir,
                kind=str(getattr(args, "kind", "auto") or "auto"),
                quality=getattr(args, "quality", None) or None,
                begin=getattr(args, "begin", None),
                end=getattr(args, "end", None),
                oauth=getattr(args, "oauth", None),
                trim_mode=str(getattr(args, "download_trim_mode", "Safe") or "Safe"),
                media_check=str(getattr(args, "media_check", "fast") or "fast"),
                media_repair=str(getattr(args, "media_repair", "audio") or "audio"),
            )
    except TwitchDownloadError as e:
        emit_task_event("stage_failed", stage="download", completed=0, total=1)
        print(f"错误: {e}", file=sys.stderr)
        return 2
    except Exception as e:
        emit_task_event("stage_failed", stage="download", completed=0, total=1)
        print(f"错误: 下载失败: {e}", file=sys.stderr)
        return 1
    global _TASK_RESULT_CONTEXT
    _TASK_RESULT_CONTEXT = {
        "mode": "download",
        "artifacts": [("video", result.video_path), ("chat_html", result.chat_html_path)],
    }
    emit_task_event("stage_completed", stage="download", completed=1, total=1)
    return _post_download_next_steps(
        result.video_path,
        result.chat_html_path,
        download_only=bool(getattr(args, "download_only", False)),
        yes=bool(getattr(args, "yes", False)),
    )


def _export_translation_json(
    *,
    burn: Path,
    video: Path,
    chat_html: Path,
    trans_json: Path,
    force: bool = False,
    offset: float | None = None,
) -> None:
    """Export via burn. Auto-reuse when JSON already has translations (unless force).

    Forward ``offset`` so export_offset metadata matches the pipeline's intended
    timeline diagnosis (identity still uses stream timestamps either way).
    """
    existing_n = _translation_nonempty_count(trans_json)
    if existing_n > 0 and not force:
        log(
            f"[1/3] 检测到已有 {existing_n} 条非空 translation，跳过导出以免覆盖: {trans_json}\n"
            f"      （继续翻译/渲染；若要强制重导加 --force-export；只渲染用 --reuse-translation）"
        )
        return
    cmd = [
        sys.executable,
        str(burn),
        "--out-dir",
        str(trans_json.parent),
        str(video),
        str(chat_html),
        "--export-translation",
        str(trans_json),
    ]
    if force:
        cmd.append("--force-export")
    if offset is not None:
        cmd.extend(["--offset", str(offset)])
    run(
        cmd,
        error_hint=(
            "导出翻译 JSON 失败。若提示已有译文被拒绝覆盖：改用 --reuse-translation，"
            "或确认后加 --force-export。并检查 HTML 是否为 TwitchDownloader 标准格式"
        ),
    )


def _fallback_manual_after_export(
    *,
    video: Path,
    chat_html: Path,
    trans_json: Path,
    review_tsv: Path,
    review_xlsx: Path,
    workdir: Path | None,
    final_output: Path,
    reason: str,
) -> None:
    """API unavailable: export review tables and stop for hand translation (same as --manual-translation tail)."""
    log(f"\n[翻译 API] {reason}")
    filled = _translation_nonempty_count(trans_json)
    total = 0
    try:
        data = json.loads(trans_json.read_text(encoding="utf-8")) if trans_json.is_file() else {}
        total = len((data.get("messages") if isinstance(data, dict) else None) or [])
    except Exception:
        total = 0
    if filled > 0:
        log(
            f"[手翻] 当前 JSON 已有 {filled}/{total or '?'} 条非空译文（可能为中途失败残留）；"
            f"导出复核表时会保留这些行，请只补空行或改错行"
        )
    else:
        log("[手翻] 导出人工复核表（不调用 LLM；translation 列为空，请自行填写）…")
    try:
        export_review_tsv(trans_json, review_tsv)
        export_review_xlsx(trans_json, review_xlsx)
    except Exception as e:
        raise PipelineError(f"错误: 导出人工复核表失败: {e}") from e
    mark_manual_translation_required()
    print("\n[OK] 已改为人工翻译流程。请编辑 XLSX 最后一列 translation：")
    print(f"     {review_xlsx}")
    print(f"     JSON: {trans_json}")
    if filled > 0:
        print(f"     提示: 已有 {filled} 条译文会写进表内，勿整列清空。")
    print("     完成后回写并渲染：")
    _hint = (
        f"{current_cli_invocation()} {quote_cli_arg(video)} {quote_cli_arg(chat_html)} "
        f"--reuse-translation --review-done --translation-json {quote_cli_arg(trans_json)} "
        f"--review-xlsx {quote_cli_arg(review_xlsx)}"
    )
    if workdir:
        _hint += f" --workdir {quote_cli_arg(workdir)}"
    _hint += f" --output {quote_cli_arg(final_output)}"
    print(f"     {_hint}")


def ensure_translate_api_or_fallback(
    *,
    video: Path,
    chat_html: Path,
    trans_json: Path,
    review_tsv: Path,
    review_xlsx: Path,
    workdir: Path | None,
    final_output: Path,
    yes: bool = False,
) -> str:
    """Before calling the translator: probe API; on failure ask continue (manual) or retry.

    Returns:
      "api" — proceed with LLM translate
      "manual" — user chose hand translation (tables already exported by caller path)
    """
    max_rounds = 8
    for _round in range(max_rounds):
        ok, msg = probe_translate_api()
        if ok:
            log(f"[翻译 API] {msg}")
            return "api"

        print(f"\n[!] 翻译 API 不可用: {msg}", flush=True)
        print("  可检查 .env 中 OPENAI_COMPAT_BASE_URL / MODEL / API_KEY，以及网络。", flush=True)

        if yes or not _stdin_is_interactive():
            # Non-interactive: fall through to manual tables so batch jobs can continue.
            print("  （非交互/--yes：改为导出人工翻译表后停止）", flush=True)
            _fallback_manual_after_export(
                video=video,
                chat_html=chat_html,
                trans_json=trans_json,
                review_tsv=review_tsv,
                review_xlsx=review_xlsx,
                workdir=workdir,
                final_output=final_output,
                reason=msg,
            )
            return "manual"

        print("  [C] 继续 → 导出未翻译表格，自行填写后再 --review-done 渲染", flush=True)
        print("  [R] 重试 → 再探测一次 API", flush=True)
        print("  [Q] 退出", flush=True)
        try:
            raw = input("请选择 [C 继续 / R 重试 / Q 退出] (默认 C): ").strip().lower()
        except EOFError:
            raw = "c"
        if not raw:
            raw = "c"
        if raw in ("r", "retry", "重试"):
            load_dotenv_if_present()
            print("  重新探测…", flush=True)
            continue
        if raw in ("q", "quit", "exit", "n", "no"):
            raise PipelineError("已取消：翻译 API 不可用。")
        # continue / c / enter / anything else → manual
        _fallback_manual_after_export(
            video=video,
            chat_html=chat_html,
            trans_json=trans_json,
            review_tsv=review_tsv,
            review_xlsx=review_xlsx,
            workdir=workdir,
            final_output=final_output,
            reason=msg,
        )
        return "manual"

    raise PipelineError("错误: 翻译 API 多次重试仍不可用。")


def handle_translate_run_failure(
    err: BaseException,
    *,
    video: Path,
    chat_html: Path,
    trans_json: Path,
    review_tsv: Path,
    review_xlsx: Path,
    workdir: Path | None,
    final_output: Path,
    translation_context: str,
    target_language: str,
    batch_size: int,
    workers: int,
    translator: Path,
    yes: bool = False,
) -> str:
    """After a mid-run translator failure: C=manual tables, R=retry once, Q=re-raise.

    Returns:
      "manual" — stopped for hand translation
      "api" — retry succeeded (caller continues pipeline)
    Raises:
      PipelineError / original err on quit or retry failure.
    """
    print(f"\n[!] 翻译调用失败: {err}", flush=True)
    if yes or not _stdin_is_interactive():
        _fallback_manual_after_export(
            video=video,
            chat_html=chat_html,
            trans_json=trans_json,
            review_tsv=review_tsv,
            review_xlsx=review_xlsx,
            workdir=workdir,
            final_output=final_output,
            reason=str(err),
        )
        return "manual"
    print("  [C] 继续 → 用当前 JSON 导出人工表（自行翻译）", flush=True)
    print("  [R] 重试 → 再调用一次翻译 API", flush=True)
    print("  [Q] 退出", flush=True)
    try:
        choice = input("请选择 [C/R/Q] (默认 C): ").strip().lower() or "c"
    except EOFError:
        choice = "c"
    if choice in ("r", "retry", "重试"):
        run(
            [
                sys.executable,
                str(translator),
                str(trans_json),
                "--context",
                translation_context,
                "--target-language",
                target_language,
                "--batch-size",
                str(batch_size),
                "--workers",
                str(workers),
            ],
            error_hint="翻译重试仍失败。可改用人工表：--manual-translation 或修好 API 后再跑",
        )
        return "api"
    if choice in ("q", "quit", "exit"):
        raise err
    _fallback_manual_after_export(
        video=video,
        chat_html=chat_html,
        trans_json=trans_json,
        review_tsv=review_tsv,
        review_xlsx=review_xlsx,
        workdir=workdir,
        final_output=final_output,
        reason=str(err),
    )
    return "manual"


def load_yaml_file(yaml_path: Path, label: str):
    try:
        import yaml
    except ImportError:
        raise SystemExit(f"错误: 使用 {label} 需要安装 PyYAML，请运行 pip install PyYAML")
    if not yaml_path.is_file():
        raise SystemExit(f"错误: {label} 文件不存在: {yaml_path}")
    try:
        data = yaml.safe_load(yaml_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, yaml.YAMLError) as e:
        raise PipelineError(f"Invalid {label} YAML {yaml_path}: {e}") from e
    if data is None:
        return {}
    if not isinstance(data, dict):
        raise PipelineError(f"Invalid {label} YAML {yaml_path}: root must be a mapping")
    return data


def load_yaml_rules(rules_path: Path):
    """Load normalizations + optional preserve_patterns from a rules YAML.

    Returns a dict: {"normalizations": [...], "preserve_patterns": [compiled re, ...]}.
    preserve_patterns skip rule application when original matches (translate path
    still runs separately; this only protects the rules-normalize pass).
    """
    data = load_yaml_file(rules_path, "规则")
    rules = []
    if "normalizations" not in data:
        raw_rules = []
    else:
        raw_rules = data["normalizations"]
    if not isinstance(raw_rules, list):
        raise PipelineError(
            f"Invalid rules YAML {rules_path}: normalizations must be a list"
        )
    for rule_index, item in enumerate(raw_rules):
        if not isinstance(item, dict):
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: normalizations[{rule_index}] must be a mapping"
            )
        targets = item.get("match", [])
        if isinstance(targets, str):
            targets = [targets]
        elif not isinstance(targets, list):
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: normalizations[{rule_index}].match must be a string or list"
            )
        if not all(isinstance(target, (str, int, float)) for target in targets):
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: normalizations[{rule_index}].match contains a non-scalar value"
            )
        translation = item.get("translation")
        if translation is None:
            continue
        rules.append({
            "name": item.get("name", "unnamed"),
            "match": {str(x) for x in targets},
            "translation": str(translation),
        })
    preserve = []
    preserve_raw = data.get("preserve_patterns")
    if preserve_raw is None:
        preserve_raw = []
    if not isinstance(preserve_raw, list):
        raise PipelineError(
            f"Invalid rules YAML {rules_path}: preserve_patterns must be a list"
        )
    for pattern_index, pat in enumerate(preserve_raw):
        try:
            preserve.append(re.compile(str(pat)))
        except re.error as e:
            raise PipelineError(
                f"Invalid rules YAML {rules_path}: preserve_patterns[{pattern_index}] is not a valid regex: {e}"
            ) from e
    return {"normalizations": rules, "preserve_patterns": preserve}


def publish_output(src_path: Path, dst_path: Path, *, backup_prev: bool = True):
    """Copy rendered output to the final path using a temp file + atomic replace.

    When backup_prev is True (default), rename an existing dst to dst.bak first and
    restore it if the replace fails — matching burn's default backup contract.
    """
    src_path = Path(src_path)
    dst_path = Path(dst_path)
    if not src_path.is_file():
        raise PipelineError(f"错误: 渲染输出不存在: {src_path}")
    if src_path.resolve() == dst_path.resolve():
        return dst_path
    dst_path.parent.mkdir(parents=True, exist_ok=True)
    backup = None
    backup_created = False
    if backup_prev and dst_path.is_file():
        backup = Path(str(dst_path) + ".bak")
        try:
            if backup.is_file():
                backup.unlink()
            dst_path.rename(backup)
            backup_created = True
            print(f"  [backup] {backup}")
        except OSError as e:
            print(f"  warning: cannot backup {dst_path}: {e}")
            backup = None
            backup_created = False
    fd, tmp_name = tempfile.mkstemp(prefix=dst_path.stem + ".", suffix=".partial.mp4", dir=str(dst_path.parent))
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        try:
            shutil.copy2(src_path, tmp_path)
            os.replace(tmp_path, dst_path)
        except OSError:
            if backup_created and backup is not None and backup.is_file() and not dst_path.is_file():
                try:
                    backup.rename(dst_path)
                    print(f"  已从备份恢复: {dst_path}")
                except OSError as restore_err:
                    print(f"  警告: 无法从备份恢复 {backup}: {restore_err}")
            raise
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except OSError:
            pass
    try:
        src_path.unlink()
    except OSError:
        pass
    return dst_path


def normalize_translation(json_path: Path, rules_path: Path | None = None):
    if not rules_path:
        print("\n[规则清洗] 未指定 --rules，跳过规则清洗。")
        return
    if DRY_RUN:
        print(f"\n[dry-run] 跳过规则清洗写入: {json_path}")
        return
    loaded = load_yaml_rules(rules_path)
    rules = loaded.get("normalizations") or []
    preserve_patterns = loaded.get("preserve_patterns") or []
    if not rules:
        print(f"\n[规则清洗] 规则文件无 normalizations: {rules_path}")
        return
    data = json.loads(json_path.read_text(encoding="utf-8"))
    changed = []
    for msg in data.get("messages", []):
        original = str(msg.get("original", ""))
        if any(p.search(original) for p in preserve_patterns):
            continue
        for rule in rules:
            if original in rule["match"] and msg.get("translation") != rule["translation"]:
                changed.append((msg.get("index"), rule["name"], original, msg.get("translation"), rule["translation"]))
                msg["translation"] = rule["translation"]
                break
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    if changed:
        print(f"\n[规则清洗] 已应用 {len(changed)} 条修改，规则文件: {rules_path}")
        for idx, rule_name, original, old, new in changed:
            print(f"  [{idx}] {rule_name}: {original!r}: {old!r} -> {new!r}")
    else:
        print(f"\n[规则清洗] 无需修改，规则文件: {rules_path}")


def load_profile(profile_path: Path):
    data = load_yaml_file(profile_path, "Profile")
    glossary_value = data.get("glossary")
    if glossary_value is not None and not isinstance(glossary_value, dict):
        raise PipelineError(
            f"Invalid Profile YAML {profile_path}: glossary must be a mapping"
        )
    preserve_value = data.get("preserve")
    if preserve_value is not None and not isinstance(preserve_value, list):
        raise PipelineError(
            f"Invalid Profile YAML {profile_path}: preserve must be a list"
        )
    style_value = data.get("translation_style")
    if style_value is not None and not isinstance(style_value, dict):
        raise PipelineError(
            f"Invalid Profile YAML {profile_path}: translation_style must be a mapping"
        )
    context_parts = []
    if data.get("context"):
        context_parts.append(str(data["context"]))

    glossary = data.get("glossary") or {}
    if glossary:
        terms = []
        for src, dst in glossary.items():
            terms.append(f"  {src} -> {dst}")
        context_parts.append(
            "**术语词典 / Glossary (必须严格遵守 / MUST follow strictly)**\n"
            + "\n".join(terms)
        )

    preserve = data.get("preserve") or []
    if preserve:
        context_parts.append("需要保留 / Preserve: " + ", ".join(map(str, preserve)))

    style = data.get("translation_style") or {}
    if style:
        style_lines = [f"{k}: {v}" for k, v in style.items()]
        context_parts.append("翻译风格 / Translation style:\n" + "\n".join(style_lines))

    return "\n\n".join(context_parts), data


def _review_issue_map(json_path: Path, max_chars: int = 90):
    """Map message index -> (severity, codes, notes) from lint, without printing a full report."""
    try:
        raw = json_path.read_text(encoding="utf-8")
        json.loads(raw)
    except FileNotFoundError:
        print(f"[WARN] 复核表 lint 跳过：找不到 {json_path}", flush=True)
        return {}
    except (OSError, UnicodeError, json.JSONDecodeError) as e:
        # Do not fail export on bad JSON here; surface why lint columns are empty.
        print(f"[WARN] 复核表 lint 跳过：无法解析 {json_path}: {e}", flush=True)
        return {}
    # Reuse lint rules quietly; suppress console noise via stdout redirect.
    import contextlib
    import io
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        try:
            issues = lint_translation(json_path, report_path=None, max_chars=max_chars)
        except SystemExit:
            issues = []
        except Exception as e:
            print(f"[WARN] 复核表 lint 失败: {e}", flush=True)
            issues = []
    by_index: dict = {}
    for issue in issues:
        idx = issue.get("index")
        bucket = by_index.setdefault(idx, {"severity": "OK", "codes": [], "notes": []})
        sev = issue.get("severity", "WARN")
        if sev == "FAIL" or bucket["severity"] != "FAIL":
            if sev == "FAIL":
                bucket["severity"] = "FAIL"
            elif bucket["severity"] != "FAIL":
                bucket["severity"] = sev
        code = str(issue.get("code", ""))
        if code:
            bucket["codes"].append(code)
        note = str(issue.get("message", ""))
        if note:
            bucket["notes"].append(note)
    return by_index


def _review_rows(json_path: Path, include_lint: bool = True, max_chars: int = 90):
    data = json.loads(json_path.read_text(encoding="utf-8"))
    issue_map = _review_issue_map(json_path, max_chars=max_chars) if include_lint else {}
    rows = []
    for msg in data.get("messages", []):
        idx = msg.get("index", "")
        info = issue_map.get(idx) or issue_map.get(str(idx)) or {}
        severity = info.get("severity", "OK") if include_lint else ""
        codes = ",".join(info.get("codes") or []) if include_lint else ""
        notes = " | ".join(info.get("notes") or []) if include_lint else ""
        rows.append([
            idx,
            msg.get("timestamp", ""),
            msg.get("author", ""),
            str(msg.get("original", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
            str(msg.get("translation", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
            severity,
            codes,
            notes,
        ])
    return rows


def export_review_tsv(json_path: Path, review_path: Path):
    """导出人工复核 TSV。translation 列可直接编辑后再导入。"""
    lines = ["index\ttimestamp\tauthor\toriginal\ttranslation\tlint_severity\tlint_codes\tlint_notes"]
    for row in _review_rows(json_path, include_lint=True):
        lines.append("\t".join(map(str, row)))
    review_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    print(f"\n[人工复核] 已导出中英对照 TSV: {review_path}")


def export_review_xlsx(json_path: Path, review_path: Path):
    """导出带列宽、换行和冻结表头的人工复核 XLSX。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise SystemExit("错误: 导出 XLSX 需要 openpyxl，请先运行 python -m pip install openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    header = ["index", "timestamp", "author", "original", "translation", "lint_severity", "lint_codes", "lint_notes"]
    ws.append(header)
    for row in _review_rows(json_path, include_lint=True):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fail_fill = PatternFill("solid", fgColor="F8CBAD")
    warn_fill = PatternFill("solid", fgColor="FFE699")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {"A": 8, "B": 10, "C": 20, "D": 50, "E": 50, "F": 12, "G": 24, "H": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[3].alignment = Alignment(vertical="top", wrap_text=True)
        row[4].alignment = Alignment(vertical="top", wrap_text=True)
        sev = str(row[5].value or "").upper()
        if sev == "FAIL":
            row[5].fill = fail_fill
        elif sev == "WARN":
            row[5].fill = warn_fill

    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 36

    review_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(review_path)
    print(f"[人工复核] 已导出更适合 Excel/WPS 的 XLSX: {review_path}")


def import_review_xlsx(json_path: Path, review_path: Path):
    """把人工复核 XLSX 的 translation 列回写到 JSON。"""
    if DRY_RUN:
        print(f"[dry-run] 跳过 XLSX 回写: {review_path} -> {json_path}")
        return
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("错误: 读取 XLSX 需要 openpyxl，请先运行 python -m pip install openpyxl") from e
    if not review_path.is_file():
        raise SystemExit(f"错误: 找不到人工复核文件: {review_path}")
    data = json.loads(json_path.read_text(encoding="utf-8"))
    by_index = {int(m.get("index")): m for m in data.get("messages", []) if str(m.get("index", "")).isdigit()}
    wb = load_workbook(review_path)
    ws = wb.active
    header = [ws.cell(row=1, column=i).value for i in range(1, 9)]
    required = ["index", "timestamp", "author", "original", "translation"]
    if header[:5] != required:
        # Backward compatible with old 5-column review sheets.
        header5 = [ws.cell(row=1, column=i).value for i in range(1, 6)]
        if header5 != required:
            raise SystemExit(
                "错误: XLSX 表头不匹配，请保持 index/timestamp/author/original/translation 五列"
                "（可选附加 lint_severity/lint_codes/lint_notes）"
            )
    changed = 0
    for row_no in range(2, ws.max_row + 1):
        idx_value = ws.cell(row=row_no, column=1).value
        if idx_value is None:
            continue
        try:
            idx = int(idx_value)
        except ValueError:
            print(f"警告: 第 {row_no} 行 index 非数字，已跳过")
            continue
        if idx not in by_index:
            print(f"警告: 第 {row_no} 行 index={idx} 不存在，已跳过")
            continue
        raw_cell = ws.cell(row=row_no, column=5).value
        translation = str(raw_cell or "").strip()
        # Empty cells must not wipe existing non-empty translations on writeback.
        existing = str(by_index[idx].get("translation", "") or "").strip()
        if not translation and existing:
            continue
        if by_index[idx].get("translation") != translation:
            by_index[idx]["translation"] = translation
            changed += 1
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[人工复核] 已从 XLSX 回写 {changed} 条修改到: {json_path}")


def import_review_tsv(json_path: Path, review_path: Path):
    """把人工复核 TSV 的 translation 列回写到 JSON。"""
    if DRY_RUN:
        print(f"[dry-run] 跳过 TSV 回写: {review_path} -> {json_path}")
        return
    if not review_path.is_file():
        raise SystemExit(f"错误: 找不到人工复核文件: {review_path}")
    data = json.loads(json_path.read_text(encoding="utf-8"))
    by_index = {int(m.get("index")): m for m in data.get("messages", []) if str(m.get("index", "")).isdigit()}
    lines = review_path.read_text(encoding="utf-8-sig").splitlines()
    if not lines:
        raise SystemExit(f"错误: 人工复核文件为空: {review_path}")
    header = lines[0].split("\t")
    if len(header) < 5 or header[:5] != ["index", "timestamp", "author", "original", "translation"]:
        raise SystemExit(
            "错误: TSV 表头不匹配，请保持 index/timestamp/author/original/translation 五列"
            "（可选附加 lint_severity/lint_codes/lint_notes）"
        )
    changed = 0
    for line_no, line in enumerate(lines[1:], start=2):
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            print(f"警告: 第 {line_no} 行列数不足，已跳过")
            continue
        try:
            idx = int(parts[0])
        except ValueError:
            print(f"警告: 第 {line_no} 行 index 非数字，已跳过")
            continue
        if idx not in by_index:
            print(f"警告: 第 {line_no} 行 index={idx} 不存在，已跳过")
            continue
        translation = parts[4].strip()
        # Empty cells must not wipe existing non-empty translations on writeback.
        existing = str(by_index[idx].get("translation", "") or "").strip()
        if not translation and existing:
            continue
        if by_index[idx].get("translation") != translation:
            by_index[idx]["translation"] = translation
            changed += 1
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[人工复核] 已从 TSV 回写 {changed} 条修改到: {json_path}")


LINT_URL_RE = re.compile(r"https?://\S+")
LINT_MENTION_RE = re.compile(r"@[A-Za-z0-9_]+")
LINT_BRACKET_TOKEN_RE = re.compile(r"\[[^\]]+\]")
LINT_PURE_EMOTE_RE = re.compile(r"^(?:\s*\[[^\]]+\]\s*)+$")


def _lint_issue(issues, idx, code, message, severity="WARN", original="", translation=""):
    issues.append({
        "index": idx,
        "severity": severity,
        "code": code,
        "message": message,
        "original": original,
        "translation": translation,
    })


def lint_translation(json_path: Path, report_path: Path | None = None, max_chars: int = 90, max_ratio: float = 2.8):
    """检查翻译 JSON 中的常见可疑问题，返回 issue 列表。"""
    if not json_path.is_file():
        raise SystemExit(f"错误: 翻译 JSON 不存在: {json_path}")
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        raise SystemExit(f"错误: JSON 解析失败: {json_path}: {e}")

    issues = []
    messages = data.get("messages")
    if not isinstance(messages, list):
        _lint_issue(issues, "", "schema_missing_messages", "顶层字段 messages 缺失或不是数组", "FAIL")
        messages = []

    seen_indexes = set()
    for pos, msg in enumerate(messages):
        if not isinstance(msg, dict):
            _lint_issue(issues, pos, "schema_message_not_object", "消息条目不是对象", "FAIL")
            continue

        idx = msg.get("index", pos)
        original = str(msg.get("original", ""))
        translation = str(msg.get("translation", "")) if msg.get("translation") is not None else ""
        original_s = original.strip()
        translation_s = translation.strip()

        for required in ["index", "original", "translation"]:
            if required not in msg:
                _lint_issue(issues, idx, "schema_missing_field", f"缺少字段: {required}", "FAIL", original, translation)

        if idx in seen_indexes:
            _lint_issue(issues, idx, "duplicate_index", "index 重复", "FAIL", original, translation)
        seen_indexes.add(idx)

        if not translation_s:
            _lint_issue(issues, idx, "empty_translation", "translation 为空", "FAIL", original, translation)
            continue

        is_pure_emote = bool(LINT_PURE_EMOTE_RE.fullmatch(original_s))
        if is_pure_emote and translation_s != original_s:
            _lint_issue(issues, idx, "pure_emote_changed", "纯 emote 消息应保持原样", "WARN", original, translation)

        original_mentions = set(LINT_MENTION_RE.findall(original))
        missing_mentions = sorted(m for m in original_mentions if m not in translation)
        if missing_mentions:
            _lint_issue(issues, idx, "mention_lost", "翻译丢失 @用户名: " + ", ".join(missing_mentions), "WARN", original, translation)

        original_urls = set(LINT_URL_RE.findall(original))
        missing_urls = sorted(u for u in original_urls if u not in translation)
        if missing_urls:
            _lint_issue(issues, idx, "url_lost", "翻译丢失 URL: " + ", ".join(missing_urls), "WARN", original, translation)

        original_brackets = set(LINT_BRACKET_TOKEN_RE.findall(original))
        missing_brackets = sorted(b for b in original_brackets if b not in translation)
        if missing_brackets:
            _lint_issue(issues, idx, "bracket_token_lost", "翻译丢失方括号 token/emote: " + ", ".join(missing_brackets), "WARN", original, translation)

        if not is_pure_emote and len(translation_s) > max_chars:
            _lint_issue(issues, idx, "too_long", f"翻译超过 {max_chars} 字，可能不适合弹幕显示", "WARN", original, translation)
        elif not is_pure_emote and original_s and len(translation_s) / max(1, len(original_s)) > max_ratio and len(translation_s) > 24:
            _lint_issue(issues, idx, "expansion_ratio_high", f"翻译长度超过原文 {max_ratio:.1f} 倍", "WARN", original, translation)

    fail_count = sum(1 for issue in issues if issue["severity"] == "FAIL")
    warn_count = sum(1 for issue in issues if issue["severity"] == "WARN")
    print(f"\n[翻译质检] 文件: {json_path}")
    print(f"  消息数: {len(messages)}")
    print(f"  FAIL: {fail_count}, WARN: {warn_count}")

    if issues:
        for issue in issues[:80]:
            print(f"  [{issue['severity']}] #{issue['index']} {issue['code']}: {issue['message']}")
        if len(issues) > 80:
            print(f"  ... 还有 {len(issues) - 80} 条未显示")
    else:
        print("  未发现确定性规则问题。")

    if report_path:
        lines = ["index\tseverity\tcode\tmessage\toriginal\ttranslation"]
        for issue in issues:
            row = [
                str(issue["index"]),
                issue["severity"],
                issue["code"],
                issue["message"],
                str(issue.get("original", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
                str(issue.get("translation", "")).replace("\t", " ").replace("\r", " ").replace("\n", " "),
            ]
            lines.append("\t".join(row))
        report_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
        print(f"  质检报告: {report_path}")

    return issues


def _doctor_script_invocation() -> str:
    """Best-effort script or console-entry path for copy-paste hints."""
    return current_cli_script()


def doctor(args):
    """检查本机运行环境。"""
    print("# 环境诊断 / Doctor")
    # Prefer the trusted portable FFmpeg directory before PATH probes.
    tools_bin = prepend_tools_ffmpeg_to_path()
    if tools_bin:
        print(f"[info] 使用可信目录中的便携 FFmpeg: {tools_bin}")

    if getattr(args, "offer_fix", False):
        offer_fixes(assume_yes=bool(getattr(args, "yes", False) or getattr(args, "fix_yes", False)))

    ok = True
    fails: list[str] = []
    warns: list[str] = []
    api_ok = False
    offset_diag = None
    video = None
    html = None
    v_dur = 0.0

    def check(name, passed, detail="", fix="", required=True):
        nonlocal ok
        status = "OK" if passed else ("FAIL" if required else "WARN")
        print(f"[{status}] {name}{(': ' + detail) if detail else ''}")
        if not passed:
            if required:
                ok = False
                fails.append(name)
            else:
                warns.append(name)
            if fix:
                print(f"      修复建议: {fix}")

    check("Python", sys.version_info >= (3, 10), sys.version.split()[0], "安装 Python 3.10 或更高版本: https://www.python.org/downloads/")
    for exe in ["ffmpeg", "ffprobe"]:
        path = safe_which(exe)
        fix = "安装 FFmpeg: https://ffmpeg.org/download.html"
        if platform.system() == "Windows":
            fix += (
                "\n      Windows: winget install --id Gyan.FFmpeg -e"
                "\n      或: choco install ffmpeg -y"
                f"\n      或: {current_cli_invocation()} --doctor --offer-fix"
                "\n      便携: 运行 --doctor --offer-fix 可安装到可信工具目录"
            )
        elif platform.system() == "Darwin":
            fix += "\n      macOS: brew install ffmpeg"
        else:
            fix += (
                "\n      Linux: sudo apt install ffmpeg fonts-noto-cjk"
                "\n      或: sudo dnf install ffmpeg"
            )
        check(exe, bool(path), path or "未找到", fix)

    packages = {
        "Pillow": "PIL",
        "beautifulsoup4": "bs4",
        "openai": "openai",
        "PyYAML": "yaml",
    }
    missing_required_pkgs: list[str] = []
    for display, module in packages.items():
        try:
            present = importlib.util.find_spec(module) is not None
        except (ValueError, ModuleNotFoundError):
            # Stub modules in tests may set __spec__ = None.
            present = module in sys.modules
        if not present and module in ("PIL", "bs4", "yaml"):
            missing_required_pkgs.append(display)
        check(
            display,
            present,
            fix=f"pip install {display}\n      或: pip install -r requirements.txt",
        )

    if getattr(args, "font_path", None) and args.font_path != "auto":
        check("常规字体", Path(args.font_path).is_file(), args.font_path, "用 --font-path 指定一个可用字体")
    elif getattr(args, "font_path", "auto") == "auto":
        reg, _ = detect_cjk_font()
        check("常规字体 (auto)", bool(reg), reg or "未检测到 CJK 字体", "用 --font-path 手动指定字体路径")
    if getattr(args, "font_bold_path", None) and args.font_bold_path != "auto":
        check("粗体字体", Path(args.font_bold_path).is_file(), args.font_bold_path, "用 --font-bold-path 指定一个可用字体")
    elif getattr(args, "font_bold_path", "auto") == "auto":
        _, bold = detect_cjk_font()
        check("粗体字体 (auto)", bool(bold), bold or "未检测到 CJK 字体", "用 --font-bold-path 手动指定字体路径")

    base_url = os.getenv("OPENAI_COMPAT_BASE_URL")
    api_key = os.getenv("OPENAI_COMPAT_API_KEY")
    model = os.getenv("OPENAI_COMPAT_MODEL")
    check(
        "翻译 Base URL",
        bool(base_url),
        base_url or "未设置",
        f"设置 OPENAI_COMPAT_BASE_URL；仅复用翻译可忽略\n      可 {current_cli_invocation()} --init 生成 .env",
        required=False,
    )
    check("翻译 Model", bool(model), model or "未设置", "设置 OPENAI_COMPAT_MODEL；仅复用翻译可忽略", required=False)
    check("翻译 API Key", bool(api_key), "已设置" if api_key else "未设置", "设置 OPENAI_COMPAT_API_KEY；仅复用翻译可忽略", required=False)
    api_ok = bool(base_url and model and api_key)

    if getattr(args, "video", None):
        video = Path(args.video).resolve()
        check("输入视频", video.is_file(), str(video), "检查视频路径")
        if video.is_file() and safe_which("ffprobe"):
            try:
                probe = subprocess.run(
                    [
                        require_executable("ffprobe"),
                        "-v",
                        "error",
                        "-show_entries",
                        "format=duration",
                        "-of",
                        "csv=p=0",
                        str(video),
                    ],
                    capture_output=True,
                    text=True,
                    timeout=45,
                )
            except (OSError, subprocess.TimeoutExpired) as exc:
                check("视频可读取", False, str(exc)[:120], "确认 ffprobe 可用且视频文件未损坏")
            else:
                check(
                    "视频可读取",
                    probe.returncode == 0,
                    ((probe.stdout or "").strip() + (probe.stderr or "").strip())[:120],
                    "确认视频文件未损坏",
                )
                try:
                    v_dur = (
                        float((probe.stdout or "").strip().splitlines()[0])
                        if probe.returncode == 0
                        else 0.0
                    )
                except (ValueError, IndexError):
                    v_dur = 0.0
    if getattr(args, "chat_html", None):
        html = Path(args.chat_html).resolve()
        check("聊天 HTML", html.is_file(), str(html), "检查 HTML 路径")
        # 时间轴对齐诊断：用与主路径相同的 parse_chat_html，而不是只认 Web data-timestamp。
        if html.is_file() and args.video and video is not None and video.is_file() and safe_which("ffprobe"):
            try:
                import tempfile as _tf

                from chat_parser import parse_chat_html as _parse_chat_html
                from chat_window import compute_time_offset, format_offset_diagnosis

                with _tf.TemporaryDirectory(prefix="doctor_chat_") as tmp:
                    chat = _parse_chat_html(str(html), tmp)
                    msgs = chat.get("messages") or []
                    if msgs and v_dur > 0:
                        first_ts = float(msgs[0].get("timestamp", 0) or 0)
                        diag = compute_time_offset(msgs, video_duration=v_dur, manual_offset=getattr(args, "offset", None))
                        offset_diag = diag
                        if diag.get("mode") == "auto":
                            check(
                                "时间轴对齐",
                                True,
                                f"首条 {first_ts:.0f}s / 视频 {v_dur:.0f}s；将自动 offset={diag['offset']:.0f}s",
                                required=False,
                            )
                        elif first_ts > v_dur:
                            check(
                                "时间轴对齐",
                                False,
                                f"首条消息 {first_ts:.0f}s > 视频时长 {v_dur:.0f}s；自动检测未触发",
                                "用 --offset <秒> 手动指定并用 --preview-frame 验证",
                                required=False,
                            )
                        elif diag.get("warnings"):
                            check(
                                "时间轴对齐",
                                False,
                                diag["warnings"][0][:160],
                                "用 --preview-clip / --offset 确认",
                                required=False,
                            )
                        else:
                            check(
                                "时间轴对齐",
                                True,
                                f"首条消息 {first_ts:.0f}s，视频时长 {v_dur:.0f}s，共 {len(msgs)} 条",
                                required=False,
                            )
                        print()
                        print(format_offset_diagnosis(diag))
                    elif not msgs:
                        check("时间轴对齐", False, "解析到 0 条消息，无法诊断偏移", "确认 HTML 为 TwitchDownloader 导出", required=False)
            except Exception as e:
                # doctor 不应因诊断失败而整体失败
                check("时间轴对齐", True, f"跳过详细诊断 ({type(e).__name__})", required=False)

    print("\n诊断结果:", "通过" if ok else "存在问题")

    # ---- 就绪清单（P1 分级）----
    min_ok, full_ok = print_readiness_report()

    # Default UX: if not ready for render, ask to help install FFmpeg (TTY only).
    # install.bat / doctor.bat / run.bat doctor all hit this path.
    offered = bool(getattr(args, "offer_fix", False))
    if not min_ok:
        ran = maybe_prompt_offer_fixes(
            already_offered=offered,
            assume_yes=bool(getattr(args, "yes", False) or getattr(args, "fix_yes", False)),
        )
        if ran or offered:
            print("\n--- 修复后复检 ---")
            if safe_which("ffmpeg"):
                fails = [f for f in fails if f != "ffmpeg"]
            if safe_which("ffprobe"):
                fails = [f for f in fails if f != "ffprobe"]
            ok = len(fails) == 0
            min_ok, full_ok = print_readiness_report()

    # ---- 推荐下一步（可复制命令）----
    script = _doctor_script_invocation()
    if fails:
        print("\n先处理 FAIL 项（上方「修复建议」可复制）：")
        for name in fails:
            print(f"  - {name}")
    if missing_required_pkgs:
        print("  pip install -r requirements.txt")
    if offset_diag and offset_diag.get("mode") == "auto":
        print(f"\n# doctor 检测到自动 offset≈{float(offset_diag.get('offset') or 0):.0f}s，请用预览核对")
    elif offset_diag and offset_diag.get("warnings"):
        print("\n# 时间轴有警告，请用 --preview-clip / --offset 确认")
    print_setup_next_steps(
        has_api=api_ok,
        has_ffmpeg=bool(safe_which("ffmpeg") and safe_which("ffprobe")),
        video=video if video and video.is_file() else None,
        chat=html if html and html.is_file() else None,
        script=script,
    )
    # Exit non-zero if classic doctor fails OR minimum render readiness fails.
    return 0 if (ok and min_ok) else 1


# argparse defaults used by job/layout/render “CLI wins” application.
PIPELINE_CLI_DEFAULTS = {
    "video": None,
    "chat_html": None,
    "context": "livestream chat",
    "target_language": "zh",
    "profile": None,
    "layout_preset": None,
    "render_preset": None,
    "lazy_message_images": False,
    "message_image_cache_size": 256,
    "max_visible": 0,
    "msg_lifetime": 14.0,
    "max_message_lines": 0,
    "min_visible_seconds": 0.0,
    "arrival_interval": 0.0,
    "stack_mode": "lanes",
    "x_ratio": 0.0,
    "y_ratio": 0.0,
    "width_ratio": 0.0,
    "height_ratio": 0.0,
    "font_size_ratio": 0.0,
    "emote_height": 22,
    "translation_json": None,
    "reuse_translation": False,
    "force_export": False,
    "strict_import": False,
    "skip_translate": False,
    "manual_translation": False,
    "render_original": False,
    "review": False,
    "review_done": False,
    "review_tsv": None,
    "review_xlsx": None,
    "lint_translation": None,
    "lint_report": None,
    "lint_max_chars": 90,
    "rules": None,
    "output": None,
    "doctor": False,
    "init": False,
    "init_job": False,
    "list_jobs": False,
    "job": None,
    "mode": "auto",
    "workdir": None,
    "dry_run": False,
    "quiet": False,
    "verbose": False,
    "x": 15,
    "y": 327,
    "width": 497,
    "height": 363,
    "font_size": 15,
    "font_path": "auto",
    "font_bold_path": "auto",
    "fps": 15,
    "output_fps": None,
    "bg_alpha": 255,
    "keep_temp": False,
    "no_backup_prev": False,
    "offset": None,
    "clean": False,
    "clean_progress": False,
    "preview_frame": None,
    "preview_image": None,
    "preview_clip": None,
    "preview_dense": False,
    "yes": False,
    "batch_size": 10,
    "workers": 4,
    "encoder": "x264",
    "video_preset": None,
    "crf": 18,
    "video_bitrate": None,
    "maxrate": None,
    "bufsize": None,
    "audio_codec": "aac",
    "audio_bitrate": "192k",
    "overlay_codec": "vp9",
    "webm_crf": 30,
    "webm_cpu_used": 4,
    "no_reuse_static_frames": False,
    "no_skip_blank_frames": False,
    "blank_hold_seconds": 0.5,
}


def _cli_flag_present(*flags: str) -> bool:
    """True if any of the given CLI flags appear in sys.argv (explicit user intent)."""
    argv = sys.argv[1:]
    for flag in flags:
        if flag in argv:
            return True
        # also match --flag=value form
        prefix = flag + "="
        if any(a.startswith(prefix) for a in argv):
            return True
    return False


def apply_mode_defaults(args) -> list[str]:
    """Apply --mode scenario defaults without overriding explicit CLI values.

    - preview: via apply_preview_first_defaults (preview_clip=10, overlay png when safe)
    - translate: stop after translate/rules/lint/review export (no burn)
    - render: require reuse-translation / render-original / skip / manual / review-only
    - full/auto: no-op defaults
    Returns list of applied field names for logging.
    """
    mode = str(getattr(args, "mode", "auto") or "auto").strip().lower()
    applied: list[str] = []
    if mode in ("auto", "full"):
        return applied

    if mode == "preview":
        preview_applied = apply_preview_first_defaults(
            args,
            PIPELINE_CLI_DEFAULTS,
            explicit_overlay_codec=_cli_flag_present("--overlay-codec"),
        )
        for name in preview_applied:
            if name == "preview_clip":
                applied.append("preview_clip=10")
            elif name == "overlay_codec":
                applied.append("overlay_codec=png")
            else:
                applied.append(name)
        return applied

    if mode == "translate":
        # Translate path: after API translate, stop before burn (like --review without review table).
        args._mode_stop_after_translate = True  # type: ignore[attr-defined]
        applied.append("stop_after_translate")
        return applied

    if mode == "render":
        # Allow paths that do not call the live translation API.
        needs_live_api = not (
            bool(getattr(args, "render_original", False))
            or bool(getattr(args, "reuse_translation", False))
            or bool(getattr(args, "skip_translate", False))
            or bool(getattr(args, "manual_translation", False))
            or bool(getattr(args, "review", False))
            or bool(getattr(args, "review_done", False))
            or (
                getattr(args, "lint_translation", None)
                and getattr(args, "lint_translation", None) != "__PIPELINE__"
            )
        )
        if needs_live_api:
            raise PipelineError(
                "错误: --mode render 不会调用翻译 API。"
                "请使用 --reuse-translation（已有翻译 JSON）或 --render-original，"
                "或改用 --mode full / --mode auto 做完整翻译出片。"
            )
        applied.append("render_only_guard")
        return applied

    raise PipelineError(f"错误: 未知 --mode {mode!r}，可选 auto|preview|translate|render|full")


def _main():
    # Activate only the trusted source/user-data portable FFmpeg directory.
    prepend_tools_ffmpeg_to_path()
    parser = argparse.ArgumentParser(description="Generate translated chat overlay video from Twitch HTML")
    parser.add_argument("video", nargs="?", help="Source video path, e.g. video.mp4")
    parser.add_argument("chat_html", nargs="?", help="Twitch chat HTML path, e.g. chat.html")
    parser.add_argument("--context", default="livestream chat", help="Background context passed to the translator")
    parser.add_argument("--target-language", default="zh", help="Target language for translation (e.g. zh, ja, ko, en). Default: zh")
    parser.add_argument("--profile", default=None, help="翻译 profile YAML，例如 profiles/default.yaml；会合并 context、glossary、preserve 和 style")
    parser.add_argument(
        "--layout-preset",
        default=None,
        help="渲染布局 YAML 或短名，例如 profiles/layout_default.yaml 或 compact；命令行布局参数优先覆盖",
    )
    parser.add_argument(
        "--render-preset",
        default=None,
        help="编码/性能 YAML 或短名，例如 profiles/render_default.yaml 或 fast；命令行参数优先覆盖",
    )
    parser.add_argument(
        "--init",
        action="store_true",
        help="首次脚手架：创建 .env（从 .env.example）与 jobs/example_job.yaml，打印推荐命令",
    )
    parser.add_argument(
        "--offer-fix",
        action="store_true",
        help="doctor 时直接进入修复流程（不先问总开关）；默认在缺 FFmpeg 时也会询问是否帮忙安装",
    )
    parser.add_argument(
        "--fix-yes",
        action="store_true",
        help="与 --offer-fix 联用：非交互默认同意修复步骤（CI/脚本用）",
    )
    parser.add_argument(
        "--offer-td-cli",
        action="store_true",
        help="可选增强: 自动下载/引导安装 TwitchDownloaderCLI 到 tools/（需确认；--yes 直接下载）",
    )
    parser.add_argument(
        "--download",
        default=None,
        metavar="URL_OR_ID",
        help="用 TwitchDownloaderCLI 下载 VOD/Clip 视频 + 嵌入表情的聊天 HTML（可选增强）",
    )
    parser.add_argument(
        "--download-dir",
        default=None,
        help="--download 输出目录（默认 downloads/<id>_<时间>/）",
    )
    parser.add_argument(
        "--download-only",
        action="store_true",
        help="与 --download 联用：只下载并打印路径后退出（不进入下一步菜单）",
    )
    parser.add_argument(
        "--quality",
        default="1080p60",
        help="--download 视频画质（默认 1080p60；不可用时 CLI 会回退）",
    )
    parser.add_argument(
        "--begin",
        default=None,
        help="--download 裁切起点（仅 VOD；如 0:01:40 或 100s）",
    )
    parser.add_argument(
        "--end",
        default=None,
        help="--download 裁切终点（仅 VOD）",
    )
    parser.add_argument(
        "--segment",
        action="append",
        default=None,
        metavar="BEGIN-END",
        help=(
            "--download 多段裁切（可重复；同一 VOD）。"
            "例: --segment 0:10:00-0:12:30 --segment 0:40:00-0:43:00；"
            "与 --begin/--end 同时出现时以 --segment 为准"
        ),
    )
    parser.add_argument(
        "--cut",
        action="append",
        default=None,
        metavar="START-END",
        help=(
            "合并后切除时间段（可重复）。"
            "例: --cut 21:01-22:59 删除合并视频的第 21 分 01 秒到 22 分 59 秒。"
            "时间轴自动前移，聊天同步裁剪。仅与 --segment 多段下载联用。"
        ),
    )
    parser.add_argument(
        "--download-output-fps",
        type=float,
        default=None,
        help="合并视频强制 CFR 帧率（如 60）。不指定则保持源帧率。",
    )
    parser.add_argument(
        "--download-encoder",
        default="auto",
        choices=["auto", "x264", "nvenc", "qsv", "amf"],
        help="合并视频编码器（默认 auto 自动探测硬件编码器）",
    )
    parser.add_argument(
        "--download-trim-mode",
        default="Safe",
        choices=["Safe", "Exact"],
        help="VOD 裁切模式；Safe（默认、推荐）避免 Exact 的时间戳偏移，Exact 仅用于明确需要精确裁切时。",
    )
    parser.add_argument(
        "--media-check",
        default="fast",
        choices=["off", "fast", "decode"],
        help="媒体健康门禁：fast=流/时长/AAC包检查（默认）；decode=额外完整解码；off=不建议，仅跳过检查。",
    )
    parser.add_argument(
        "--source-media-check",
        default="fast",
        choices=["off", "fast", "decode"],
        help="本地输入视频门禁：fast=快速检查（默认）；decode=翻译/渲染前完整解码；off=仅用于排障。",
    )
    parser.add_argument(
        "--media-repair",
        default="audio",
        choices=["off", "audio"],
        help="健康失败时自动尝试非破坏性音频时间轴修复（默认 audio）；输出 *.repaired.mp4，原下载不覆盖；off 可禁用。",
    )
    parser.add_argument(
        "--kind",
        default="auto",
        choices=["auto", "vod", "clip"],
        help="--download 源类型（默认 auto 识别）",
    )
    parser.add_argument(
        "--oauth",
        default=None,
        help="TwitchDownloaderCLI --oauth（订阅限定 VOD；勿提交到 git）",
    )
    parser.add_argument(
        "--install-td-prompt",
        action="store_true",
        help="安装脚本用：交互询问是否配置可选 TwitchDownloaderCLI（默认 No）",
    )
    parser.add_argument(
        "--init-job",
        action="store_true",
        help="引导式创建带注释的 jobs/<name>.yaml（交互问答）",
    )
    parser.add_argument(
        "--list-jobs",
        action="store_true",
        help="列出 jobs/ 下的任务配置",
    )
    parser.add_argument(
        "--job",
        default=None,
        help="从 job.yaml 加载 video/chat/output/presets 等；显式 CLI 仍优先；也可用短名",
    )
    parser.add_argument(
        "--mode",
        default="auto",
        choices=["auto", "preview", "translate", "render", "full"],
        help="场景模式: auto/full=完整流程; preview=默认10s预览; translate=只翻译; render=只渲染(需 reuse/original)",
    )
    parser.add_argument("--lazy-message-images", action="store_true", help="长片省内存：转发给 burn 的消息图 LRU 缓存模式")
    parser.add_argument("--message-image-cache-size", type=int, default=256, help="lazy 消息图缓存上限，默认 256")
    parser.add_argument(
        "--max-visible",
        type=int,
        default=0,
        help=(
            "最大同时可见消息数；默认 0=按框高/字号自动填满；"
            "显式 N 固定条数；若 N 大于框高可容纳行数会自动钳制并告警，避免弹幕叠在顶部"
        ),
    )
    parser.add_argument(
        "--msg-lifetime",
        type=positive_float_arg,
        default=14.0,
        help="消息停留秒数（仅 stack_mode=lanes；float 上浮模式忽略；必须 > 0）",
    )
    parser.add_argument("--max-message-lines", type=int, default=0, help="单条消息最多显示行数；0 表示不额外限制")
    parser.add_argument(
        "--min-visible-seconds",
        type=float,
        default=0.0,
        help="已上屏消息最短可见秒数（仅 stack_mode=lanes；float 忽略）；0 表示允许立即被顶替",
    )
    parser.add_argument("--arrival-interval", type=float, default=0.0, help="新消息最小入场间隔秒数；0 表示不限流")
    parser.add_argument(
        "--stack-mode",
        choices=("float", "lanes"),
        default="lanes",
        help="聊天堆叠: lanes=lifetime lane沉积(默认), float=Twitch上浮(仅容量顶出)",
    )
    parser.add_argument("--x-ratio", type=float, default=0.0, help="相对源视频宽度的 X 坐标；0 使用 --x")
    parser.add_argument("--y-ratio", type=float, default=0.0, help="相对源视频高度的 Y 坐标；0 使用 --y")
    parser.add_argument("--width-ratio", type=float, default=0.0, help="相对源视频宽度的 overlay 宽度；0 使用 --width")
    parser.add_argument("--height-ratio", type=float, default=0.0, help="相对源视频高度的 overlay 高度；0 使用 --height")
    parser.add_argument("--font-size-ratio", type=float, default=0.0, help="相对源视频高度的字号；0 使用 --font-size")
    parser.add_argument("--emote-height", type=int, default=22, help="emote 高度像素")
    parser.add_argument("--translation-json", default=None, help="翻译 JSON 路径，默认 <视频名>_translation.json")
    parser.add_argument("--reuse-translation", action="store_true", help="如果翻译 JSON 已存在，跳过导出和翻译，直接渲染")
    parser.add_argument(
        "--force-export",
        action="store_true",
        help="允许覆盖已有非空 translation 的 JSON（默认拒绝；防丢译）。仅影响导出步骤",
    )
    parser.add_argument(
        "--strict-import",
        action="store_true",
        help="导入翻译渲染时：author/timestamp/original 不一致则硬失败（转发给 burn；默认跳过错配）",
    )
    parser.add_argument("--skip-translate", action="store_true", help="只导出翻译 JSON，不调用翻译和渲染")
    parser.add_argument("--manual-translation", action="store_true", help="不调用 LLM；导出待翻译 JSON 和人工复核 XLSX/TSV 后停止")
    parser.add_argument("--render-original", action="store_true", help="不导出、不翻译，直接将原始聊天文本和已有 emote 渲染到视频")
    parser.add_argument("--review", action="store_true", help="LLM 翻译后导出中英对照 TSV 并停止，等待人工复核")
    parser.add_argument("--review-done", action="store_true", help="从人工复核 TSV 回写翻译后再渲染")
    parser.add_argument("--review-tsv", default=None, help="人工复核 TSV 路径，默认 <视频名>_translation_review.tsv")
    parser.add_argument("--review-xlsx", default=None, help="人工复核 XLSX 路径，默认 <视频名>_translation_review.xlsx")
    parser.add_argument(
        "--yes", "-y",
        action="store_true",
        help="非交互：翻译完成后不暂停等回车，直接渲染（默认交互会导出 Excel 并等待确认）",
    )
    parser.add_argument("--lint-translation", nargs="?", const="__PIPELINE__", default=None, help="检查翻译 JSON；可单独传 JSON 路径，或在 pipeline 中不带值使用")
    parser.add_argument("--lint-report", default=None, help="导出翻译质检 TSV 报告路径")
    parser.add_argument("--lint-max-chars", type=int, default=90, help="翻译长度告警阈值，默认 90 字")
    parser.add_argument("--rules", default=None, help="YAML 规则文件路径，例如 configs/rules.example.yaml")
    parser.add_argument("--output", default=None, help="最终输出路径；默认使用 twitch_chat_burn.py 的 <视频名>_chat.mp4")
    parser.add_argument("--doctor", action="store_true", help="检查 Python、依赖、FFmpeg、字体和翻译环境变量")
    parser.add_argument("--workdir", default=None, help="独立工作目录，所有中间文件和输出将归档到此目录")
    parser.add_argument("--dry-run", action="store_true", help="只打印计划执行步骤，不实际运行")
    parser.add_argument("--quiet", action="store_true", help="减少输出")
    parser.add_argument("--verbose", action="store_true", help="显示详细输出")

    parser.add_argument("--x", type=int, default=15)
    parser.add_argument("--y", type=int, default=327)
    parser.add_argument("--w", "--width", dest="width", type=int, default=497)
    parser.add_argument("--h", "--height", dest="height", type=int, default=363)
    parser.add_argument("--font-size", type=int, default=15)
    parser.add_argument("--font-path", default="auto", help="字体文件路径；auto 为自动检测 CJK 字体")
    parser.add_argument("--font-bold-path", default="auto", help="粗体字体路径；auto 为自动检测")
    parser.add_argument("--fps", type=int, default=15, help="弹幕 overlay 渲染帧率（默认 15；不强制成片帧率）")
    parser.add_argument(
        "--output-fps", type=float, default=None,
        help="最终成片视频帧率（可用 29.97 等分数帧率）；默认跟随源视频",
    )
    parser.add_argument("--bg-alpha", type=int, default=255, help="聊天背景透明度 0-255；255 为不透明黑底（默认），170 为半透明")
    parser.add_argument("--keep-temp", action="store_true", help="保留底层渲染中间文件，方便失败后排查/续跑")

    parser.add_argument("--no-backup-prev", action="store_true", help="不备份旧输出文件（默认自动备份为 .bak）")
    parser.add_argument("--offset", type=float, default=None, help="时间偏移修正秒数；默认交给 twitch_chat_burn.py 自动判断")
    parser.add_argument(
        "--clean",
        action="store_true",
        help="清理 --workdir 下临时文件后退出（无 workdir 时用视频目录/当前目录）：默认只删 *.partial.mp4；加 --clean-all 才删全部已结束 job_/batch_；默认不删 *.progress.json",
    )
    parser.add_argument(
        "--clean-all",
        action="store_true",
        help="与 --clean 联用：删除 workdir/out 下全部已结束的工具 job_/batch_ 目录（仍跳过 running）",
    )
    parser.add_argument(
        "--clean-progress",
        action="store_true",
        help="与 --clean 联用：同时删除 *.progress.json 进度文件",
    )
    parser.add_argument("--preview-frame", type=float, default=None, help="只导出指定秒数的一张预览图，不渲染整片")
    parser.add_argument("--preview-image", default=None, help="预览图输出路径；默认 <视频名>_preview_<秒数>s.png")
    parser.add_argument(
        "--preview-clip",
        type=float,
        default=None,
        help="只渲染 N 秒短片；默认真从 0 秒开始，可用 --preview-dense 选弹幕最密段",
    )
    parser.add_argument(
        "--preview-dense",
        action="store_true",
        help="与 --preview-clip 联用：自动选弹幕最密时间窗",
    )
    parser.add_argument("--batch-size", type=int, default=10)
    parser.add_argument("--workers", type=int, default=4)
    # Performance / encode (forwarded to twitch_chat_burn.py)
    parser.add_argument("--encoder", default="x264", choices=["auto", "x264", "nvenc", "qsv", "amf"],
                        help="最终视频编码器；auto 优先硬件，默认 x264 最稳")
    parser.add_argument("--video-preset", default=None, help="编码预设（x264/nvenc/qsv/amf 各自体系）")
    parser.add_argument("--crf", type=int, default=18, help="质量 CRF/CQ，默认 18")
    parser.add_argument("--video-bitrate", default=None, help="视频码率，如 8M")
    parser.add_argument("--maxrate", default=None, help="最大码率")
    parser.add_argument("--bufsize", default=None, help="码率缓冲")
    parser.add_argument("--audio-codec", default="aac", choices=["aac", "copy"])
    parser.add_argument("--audio-bitrate", default="192k")
    parser.add_argument("--overlay-codec", default="vp9", choices=["vp9", "png"],
                        help="聊天层：vp9 中间 WebM 或 png 直接叠加")
    parser.add_argument("--webm-crf", type=int, default=30)
    parser.add_argument("--webm-cpu-used", type=int, default=4, help="VP9 速度 0-8，默认 4")
    parser.add_argument("--no-reuse-static-frames", action="store_true")
    parser.add_argument("--no-skip-blank-frames", action="store_true")
    parser.add_argument("--blank-hold-seconds", type=float, default=0.5)
    args = parser.parse_args()
    install_process_cleanup_handlers()

    # --init / --list-jobs / --init-job / download / td guide early (no video/html required).
    if getattr(args, "init", False):
        raise SystemExit(run_init(create_job=True, run_doctor_fn=doctor, doctor_args=args))
    if getattr(args, "list_jobs", False):
        raise SystemExit(run_list_jobs())
    if getattr(args, "init_job", False):
        created = run_job_wizard()
        if created is None:
            raise SystemExit(1)
        # If wizard saved a job, offer to continue by loading it as --job when
        # no other action was requested: re-enter via env for shell launchers.
        print(f"\n提示: 运行该配置 → {current_cli_invocation()} --job \"{created}\"")
        raise SystemExit(0)
    if getattr(args, "offer_td_cli", False):
        offer_td_cli_guide(
            assume_yes=bool(getattr(args, "yes", False) or getattr(args, "fix_yes", False))
        )
        try:
            from twitch_download import find_twitchdownloader_cli

            installed = find_twitchdownloader_cli() is not None
        except ImportError:
            installed = False
        if not installed:
            print("  [ERROR] TwitchDownloaderCLI 仍不可用；安装或引导未完成。")
        raise SystemExit(0 if installed else 1)
    if getattr(args, "install_td_prompt", False):
        maybe_prompt_offer_td_cli(
            assume_yes=bool(getattr(args, "yes", False) or getattr(args, "fix_yes", False))
        )
        raise SystemExit(0)
    if getattr(args, "download", None):
        raise SystemExit(_run_download_flow(args))

    # --job fills only fields still at CLI default (explicit CLI wins).
    job_applied: list[str] = []
    if getattr(args, "job", None):
        try:
            from job_config import resolve_job_arg

            job_path = resolve_job_arg(args.job)
            args.job = str(job_path)
            job = load_job_file(job_path)
            # Apply style fields first. Media paths may be omitted (commented in YAML)
            # for reusable jobs — then CLI args or interactive ask must supply them.
            job_applied = apply_job_to_namespace(args, job, cli_defaults=PIPELINE_CLI_DEFAULTS)
            if job_applied:
                print(f"[job] 已加载: {job_path} -> {', '.join(job_applied)}")
            else:
                print(f"[job] 已加载: {job_path}（无字段应用，可能均被 CLI 覆盖）")

            # Interactive fill for missing video/chat when stdin is a real TTY and not dry-run.
            need_video = not getattr(args, "video", None)
            need_chat = not getattr(args, "chat_html", None)
            if (need_video or need_chat) and not getattr(args, "dry_run", False):
                from pathlib import Path as _P

                from job_wizard import _guess_chat_html, _prompt_path

                interactive = _stdin_is_interactive()
                if interactive:
                    print("[job] 配置未固定视频/HTML（可复用样式）。请指定本次文件（不会写回配置）：")
                    try:
                        if need_video:
                            args.video = _prompt_path("  源视频", must_exist=True)
                        if need_chat:
                            guess = _guess_chat_html(_P(args.video)) if args.video else None
                            args.chat_html = _prompt_path("  聊天 HTML", guess, must_exist=True)
                    except (EOFError, FileNotFoundError) as e:
                        raise SystemExit(
                            f"错误: 无法取得本次视频/HTML（{e}）。\n"
                            "  请传入: --job style.yaml video.mp4 chat.html\n"
                            "  或在 YAML 取消注释 video/chat_html 后重新运行。\n"
                            f"  重试: {current_cli_invocation()} --job {quote_cli_arg(job_path)}"
                        ) from e
                else:
                    raise SystemExit(
                        "错误: job 未包含 video/chat_html，且当前非交互终端。\n"
                        "  请在命令行传入: --job style.yaml video.mp4 chat.html\n"
                        "  或在 YAML 中取消注释 video/chat_html 以固定路径。\n"
                        f"  或在交互终端重新运行: {current_cli_invocation()} "
                        f"--job {quote_cli_arg(job_path)}"
                    )

            merged = {
                "video": getattr(args, "video", None),
                "chat_html": getattr(args, "chat_html", None),
            }
            media_problems = validate_job_media_paths(merged, require_existing=True)
            # Missing keys after interactive attempt
            if not merged.get("video"):
                media_problems = list(media_problems) + ["缺少 video（请传参或取消注释配置）"]
            if not merged.get("chat_html"):
                media_problems = list(media_problems) + ["缺少 chat_html（请传参或取消注释配置）"]
            if media_problems and not getattr(args, "dry_run", False):
                msg = "错误: job 输入路径不可用\n" + "\n".join(f"  - {p}" for p in media_problems)
                raise SystemExit(msg)
            if media_problems and getattr(args, "dry_run", False):
                print("[job] 警告: " + " | ".join(str(p).splitlines()[0] for p in media_problems[:2]))
            save_last_job(job_path)
        except SystemExit:
            # Preserve intentional exits (missing media / non-interactive).
            raise
        except (OSError, ValueError) as e:
            raise SystemExit(f"错误: {e}")
    args._job_applied = set(job_applied)  # type: ignore[attr-defined]

    if args.layout_preset:
        try:
            preset = load_layout_preset(args.layout_preset)
            applied = apply_layout_preset_to_namespace(
                args, preset, cli_defaults=PIPELINE_CLI_DEFAULTS
            )
            if applied:
                print(f"[layout-preset] 已加载: {args.layout_preset} -> {', '.join(applied)}")
        except (OSError, ValueError) as e:
            raise SystemExit(f"错误: {e}")

    if getattr(args, "render_preset", None):
        try:
            rpreset = load_render_preset(args.render_preset)
            rapplied = apply_render_preset_to_namespace(
                args, rpreset, cli_defaults=PIPELINE_CLI_DEFAULTS
            )
            if rapplied:
                print(f"[render-preset] 已加载: {args.render_preset} -> {', '.join(rapplied)}")
        except Exception as e:
            print(f"[render-preset] 加载失败: {e}")
            raise SystemExit(2)

    # Doctor / clean early exits before mode guards (mode=render should not block doctor).
    if args.doctor:
        raise SystemExit(doctor(args))

    companion_err = clean_companion_flags_error(args)
    if companion_err:
        print(companion_err)
        raise SystemExit(2)

    # --clean early exit: resolve out dir from --workdir (or default), no export/translate/burn.
    if getattr(args, "clean", False):
        if args.workdir:
            out_base = Path(args.workdir).resolve()
            clean_root = out_base / "temp"
            if not clean_root.is_dir():
                clean_root = out_base
        elif args.video:
            video_path = Path(args.video).expanduser()
            if not video_path.is_file():
                print(f"--clean: 视频不存在，拒绝回退到当前目录: {video_path}")
                raise SystemExit(1)
            clean_root = video_path.resolve().parent
        else:
            print("--clean: 请指定 --workdir，或提供存在的 video 路径（避免误清 cwd）")
            raise SystemExit(1)
        clean_root = Path(os.path.abspath(str(clean_root)))
        if not clean_root.is_dir():
            print(f"--clean: 目录不存在: {clean_root}")
            raise SystemExit(1)
        if is_dangerous_publish_path(clean_root):
            print(f"--clean: 拒绝在系统目录下清理: {clean_root}")
            raise SystemExit(2)
        count, freed = clean_temp_artifacts(
            clean_root,
            clean_progress=bool(getattr(args, "clean_progress", False)),
            clean_all=bool(getattr(args, "clean_all", False)),
        )
        print(f"\n清理完成: {count} 项, 释放 {freed / (1024 * 1024):.1f} MB")
        raise SystemExit(0)

    # Mode defaults after job/presets so "still at default" checks remain valid for preview overlay.
    try:
        mode_applied = apply_mode_defaults(args)
        if mode_applied and not getattr(args, "quiet", False):
            print(f"[mode={getattr(args, 'mode', 'auto')}] {', '.join(mode_applied)}")
    except PipelineError:
        raise

    try:
        args.font_path, args.font_bold_path = resolve_font_paths(args.font_path, args.font_bold_path)
    except FileNotFoundError as e:
        raise PipelineError(f"错误: {e}")

    global DRY_RUN, VERBOSE, QUIET
    DRY_RUN = args.dry_run
    VERBOSE = args.verbose
    QUIET = args.quiet
    if args.lint_translation and args.lint_translation != "__PIPELINE__" and not args.video and not args.chat_html:
        lint_path = Path(args.lint_translation).resolve()
        report_path = Path(args.lint_report).resolve() if args.lint_report else None
        issues = lint_translation(lint_path, report_path=report_path, max_chars=args.lint_max_chars)
        raise SystemExit(1 if any(i["severity"] == "FAIL" for i in issues) else 0)
    if args.lint_translation == "__PIPELINE__" and args.video and not args.chat_html:
        lint_path = Path(args.video).resolve()
        report_path = Path(args.lint_report).resolve() if args.lint_report else None
        issues = lint_translation(lint_path, report_path=report_path, max_chars=args.lint_max_chars)
        raise SystemExit(1 if any(i["severity"] == "FAIL" for i in issues) else 0)
    if not args.video or not args.chat_html:
        parser.error(
            "需要提供 video 和 chat_html；"
            "仅 --init / --doctor / --job / 单独 --lint-translation 可省略输入文件"
        )
    if args.render_original and (args.reuse_translation or args.skip_translate or args.manual_translation or args.review or args.review_done or args.lint_translation or args.rules or args.profile):
        raise PipelineError("错误: --render-original 不能与翻译、复核、规则或 profile 参数同时使用。请只保留渲染布局参数和 --output。")
    if args.manual_translation and (args.reuse_translation or args.skip_translate or args.render_original or args.review_done):
        raise PipelineError("错误: --manual-translation 只负责导出人工翻译文件，不能与复用翻译、仅导出、原文渲染或回写复核同时使用。")
    if args.review_done and not args.reuse_translation:
        raise PipelineError(
            "错误: --review-done 必须配合 --reuse-translation 使用，"
            "避免重新导出/翻译冲掉已有 JSON。请先保留翻译文件再回写复核表。"
        )
    base_dir = Path(__file__).resolve().parent
    burn = base_dir / "twitch_chat_burn.py"
    translator = base_dir / "translate_chat_openai.py"

    video = Path(args.video).resolve()
    chat_html = Path(args.chat_html).resolve()
    if not video.is_file():
        raise PipelineError(f"错误: 视频文件不存在: {video}\n  请确认路径正确，或用 TwitchDownloader 下载视频后重试。")
    if not chat_html.is_file():
        raise PipelineError(f"错误: 聊天 HTML 文件不存在: {chat_html}\n  请用 TwitchDownloader 导出聊天 HTML，确保选择 HTML 格式。")
    explicit_output = Path(args.output).resolve() if args.output else None
    if explicit_output == video:
        raise PipelineError(
            "错误: --output 不能与源视频指向同一文件；请选择新的输出文件名，避免覆盖原片。"
        )
    if not burn.is_file():
        raise PipelineError(f"错误: 找不到核心脚本: {burn}\n  请确认从项目根目录运行，或检查 scripts/ 目录完整性。")
    if not translator.is_file() and not args.skip_translate and not args.reuse_translation and not args.manual_translation and not args.render_original:
        raise PipelineError(f"错误: 找不到翻译脚本: {translator}\n  请确认 scripts/ 目录完整性。")

    validate_source_media(video, mode=args.source_media_check, dry_run=args.dry_run)

    workdir = None
    if args.workdir:
        workdir = Path(args.workdir).resolve()
        if is_dangerous_publish_path(workdir):
            raise PipelineError(f"错误: --workdir 不能是系统目录: {workdir}")
        workdir.mkdir(parents=True, exist_ok=True)
        (workdir / "temp").mkdir(exist_ok=True)
        log(f"[workdir] 使用工作目录: {workdir}")

    def wd(default_path, filename=None):
        if workdir:
            return workdir / (filename or default_path.name)
        return default_path

    # Explicit paths always win. --workdir only relocates implicit defaults.
    trans_json = Path(args.translation_json).resolve() if args.translation_json else wd(video.with_name(video.stem + "_translation.json"))
    review_tsv = Path(args.review_tsv).resolve() if args.review_tsv else wd(video.with_name(video.stem + "_translation_review.tsv"))
    review_xlsx = Path(args.review_xlsx).resolve() if args.review_xlsx else review_tsv.with_suffix(".xlsx")
    output_default = video.with_name(video.stem + "_chat.mp4")
    if explicit_output is not None:
        final_output = explicit_output
    elif workdir:
        final_output = workdir / (video.stem + "_chat.mp4")
    else:
        final_output = output_default
    if is_dangerous_publish_path(final_output) or is_dangerous_publish_path(final_output.parent):
        raise PipelineError(f"错误: --output 不能写到系统目录: {final_output}")
    if final_output == video:
        raise PipelineError(
            "错误: --output 不能与源视频指向同一文件；请选择新的输出文件名，避免覆盖原片。"
        )

    global _TASK_RESULT_CONTEXT
    _TASK_RESULT_CONTEXT = {
        "mode": str(getattr(args, "mode", "auto") or "auto"),
        "artifacts": [
            ("video", final_output),
            ("translation_json", trans_json),
            ("review_xlsx", review_xlsx),
            ("review_tsv", review_tsv),
            ("preview_image", getattr(args, "preview_image", None)),
        ],
    }

    if args.render_original:
        log("[1/1] 不使用 LLM，直接渲染原始聊天文本和 HTML 中已有 emote")
        cmd = [
            sys.executable, str(burn), str(video), str(chat_html),
            "--x", str(args.x), "--y", str(args.y), "--w", str(args.width), "--h", str(args.height),
            "--font-size", str(args.font_size),
            "--font-path", args.font_path,
            "--font-bold-path", args.font_bold_path,
            "--bg-alpha", str(args.bg_alpha),
        ]
        append_shared_burn_args(cmd, args)
        if args.keep_temp:
            cmd.append("--keep-temp")
        if args.no_backup_prev:
            cmd.append("--no-backup-prev")
        if workdir:
            cmd.extend(["--out-dir", str(workdir / "temp")])
        if args.offset is not None:
            cmd.extend(["--offset", str(args.offset)])
        if args.preview_frame is not None:
            cmd.extend(["--preview-frame", str(args.preview_frame)])
            if args.preview_image:
                cmd.extend(["--preview-image", str(Path(args.preview_image).resolve())])
        if args.preview_clip is not None:
            cmd.extend(["--preview-clip", str(args.preview_clip)])
        if getattr(args, "preview_dense", False):
            cmd.append("--preview-dense")
        # no --import-translation on original path → no --strict-import
        run(cmd, error_hint="渲染失败，请检查视频文件、FFmpeg 和字体路径是否正确")
        if DRY_RUN:
            return
        if args.preview_frame is not None:
            log("\n[OK] 原始聊天预览图已生成。")
            return
        rendered_output = (workdir / "temp" / (video.stem + "_chat.mp4")) if workdir else output_default
        if final_output != rendered_output:
            publish_output(
                rendered_output,
                final_output,
                backup_prev=not bool(getattr(args, "no_backup_prev", False)),
            )
        log(f"\n[OK] 原始聊天 overlay 已输出到: {final_output}")
        if (
            args.preview_frame is None
            and args.preview_clip is None
            and str(getattr(args, "mode", "auto") or "auto") not in ("preview",)
        ):
            log("提示: 下次可先 --preview-clip 10 或 --mode preview 确认 offset/布局，再出长片")
        return

    if args.manual_translation:
        log(f"[1/2] 导出待人工翻译 JSON: {trans_json}")
        _export_translation_json(
            burn=burn,
            video=video,
            chat_html=chat_html,
            trans_json=trans_json,
            force=bool(getattr(args, "force_export", False)),
            offset=getattr(args, "offset", None),
        )
        if DRY_RUN:
            log("\n[dry-run] 跳过复核表导出和后续步骤。")
            return
        log("\n[2/2] 导出人工复核表（无需 LLM）")
        export_review_tsv(trans_json, review_tsv)
        export_review_xlsx(trans_json, review_xlsx)
        mark_manual_translation_required()
        print("\n[OK] 请优先编辑 XLSX 最后一列 translation：")
        print(f"     {review_xlsx}")
        print("     完成后使用以下命令回写并渲染：")
        _hint = (
            f"{current_cli_invocation()} {quote_cli_arg(video)} {quote_cli_arg(chat_html)} "
            f"--reuse-translation --review-done --translation-json {quote_cli_arg(trans_json)} "
            f"--review-xlsx {quote_cli_arg(review_xlsx)}"
        )
        if workdir:
            _hint += f" --workdir {quote_cli_arg(workdir)}"
        _hint += f" --output {quote_cli_arg(final_output)}"
        print(f"     {_hint}")
        return

    profile_context = ""
    if args.profile:
        profile_path = resolve_public_resource(args.profile, subdir="profiles")
        profile_context, profile_data = load_profile(profile_path)
        print(f"[profile] 已加载: {profile_path} ({profile_data.get('label') or profile_data.get('name') or 'unnamed'})")
    translation_context = "\n\n".join(part for part in [args.context, profile_context] if part)

    if args.reuse_translation:
        if not trans_json.is_file():
            raise PipelineError(f"错误: --reuse-translation 指定但翻译 JSON 不存在: {trans_json}\n  请先运行不带 --reuse-translation 的命令生成翻译，或用 --manual-translation 导出后人工填写。")
        log(f"[1/3] 复用翻译 JSON: {trans_json}")
    else:
        log(f"[1/3] 导出待翻译 JSON: {trans_json}")
        _export_translation_json(
            burn=burn,
            video=video,
            chat_html=chat_html,
            trans_json=trans_json,
            force=bool(getattr(args, "force_export", False)),
            offset=getattr(args, "offset", None),
        )
        if DRY_RUN:
            log("\n[dry-run] 跳过翻译和渲染步骤。")
            return

        if args.skip_translate:
            print(f"\n[OK] 已导出待翻译 JSON，未继续翻译/渲染: {trans_json}")
            return

        # Allow choosing translate mode even with bad/missing API: probe first,
        # then continue with hand-translation tables or retry.
        api_mode = ensure_translate_api_or_fallback(
            video=video,
            chat_html=chat_html,
            trans_json=trans_json,
            review_tsv=review_tsv,
            review_xlsx=review_xlsx,
            workdir=workdir,
            final_output=final_output,
            yes=bool(getattr(args, "yes", False)),
        )
        if api_mode == "manual":
            return

        log(f"\n[2/3] 调用 OpenAI-compatible 翻译器: {trans_json}")
        try:
            run([
                sys.executable, str(translator), str(trans_json),
                "--context", translation_context,
                "--target-language", args.target_language,
                "--batch-size", str(args.batch_size),
                "--workers", str(args.workers),
            ], error_hint="翻译失败，请检查 OPENAI_COMPAT_* 环境变量是否正确设置，网络是否可达")
        except PipelineError as e:
            # Mid-run API failure: same C/R/Q contract as pre-flight probe.
            mid = handle_translate_run_failure(
                e,
                video=video,
                chat_html=chat_html,
                trans_json=trans_json,
                review_tsv=review_tsv,
                review_xlsx=review_xlsx,
                workdir=workdir,
                final_output=final_output,
                translation_context=translation_context,
                target_language=args.target_language,
                batch_size=args.batch_size,
                workers=args.workers,
                translator=translator,
                yes=bool(getattr(args, "yes", False)),
            )
            if mid == "manual":
                return
        if DRY_RUN:
            log("\n[dry-run] 跳过渲染步骤。")
            return

    rules_path = resolve_public_resource(args.rules, subdir="configs") if args.rules else None
    normalize_translation(trans_json, rules_path=rules_path)

    if args.review_done:
        if review_xlsx.is_file():
            import_review_xlsx(trans_json, review_xlsx)
        else:
            import_review_tsv(trans_json, review_tsv)

    if args.lint_translation:
        report_path = Path(args.lint_report).resolve() if args.lint_report else None
        issues = lint_translation(trans_json, report_path=report_path, max_chars=args.lint_max_chars)
        if any(issue["severity"] == "FAIL" for issue in issues):
            raise PipelineError("错误: 翻译质检存在 FAIL，请修复后再渲染；如需只查看报告，可单独运行 --lint-translation。")

    if args.review:
        export_review_tsv(trans_json, review_tsv)
        export_review_xlsx(trans_json, review_xlsx)
        mark_manual_translation_required()
        print("  请优先编辑 XLSX 的最后一列 translation；TSV 仅作为兼容备份。")
        print("\n[OK] 已停在人工复核环节，尚未渲染视频。")
        print("     修改 XLSX 后运行同一命令并把 --review 换成 --review-done。")
        return

    if getattr(args, "_mode_stop_after_translate", False):
        if args.reuse_translation:
            log(f"\n[OK] --mode translate + --reuse-translation：已完成规则/质检，未渲染。JSON: {trans_json}")
        else:
            log(f"\n[OK] --mode translate：翻译已完成，未渲染。JSON: {trans_json}")
        log(
            f"     下一步渲染: {current_cli_invocation()} {quote_cli_arg(video)} {quote_cli_arg(chat_html)} "
            f"--mode render --reuse-translation --translation-json {quote_cli_arg(trans_json)} "
            f"--output {quote_cli_arg(final_output)}"
        )
        return

    # Default UX: after a *fresh* API translation, export Excel and wait for Enter
    # so the user can skim before a long render. Skip pause when:
    #   --yes / -y, dry-run, non-TTY, --reuse-translation (already translated),
    #   or --review-done (user already came back from Excel).
    did_fresh_translate = (
        not args.render_original
        and not args.reuse_translation
        and not args.review_done
        and not args.skip_translate
    )
    if did_fresh_translate:
        action = pause_after_translation_for_review(
            trans_json=trans_json,
            review_xlsx=review_xlsx,
            review_tsv=review_tsv,
            auto_continue=bool(getattr(args, "yes", False)),
            video=video,
            chat_html=chat_html,
            args=args,
            workdir=workdir,
            burn=burn,
        )
        if action == "stop":
            return
        # If user edited XLSX during the pause, pull changes back into JSON.
        if review_xlsx.is_file() and not getattr(args, "yes", False) and _stdin_is_interactive():
            try:
                import_review_xlsx(trans_json, review_xlsx)
                log(f"[复核] 已从 Excel 回写翻译: {review_xlsx}")
            except Exception as e:
                log(f"[WARN] 回写 Excel 失败（使用当前 JSON 继续）: {e}")

    log("\n[3/3] 渲染并合成翻译后的 chat overlay 视频")
    cmd = [
        sys.executable, str(burn), str(video), str(chat_html),
        "--x", str(args.x), "--y", str(args.y), "--w", str(args.width), "--h", str(args.height),
        "--font-size", str(args.font_size),
        "--font-path", args.font_path,
        "--font-bold-path", args.font_bold_path,
        "--bg-alpha", str(args.bg_alpha),
        "--import-translation", str(trans_json),
    ]
    append_strict_import_arg(cmd, args)
    append_shared_burn_args(cmd, args)
    if args.keep_temp:
        cmd.append("--keep-temp")
    if args.no_backup_prev:
        cmd.append("--no-backup-prev")
    if workdir:
        cmd.extend(["--out-dir", str(workdir / "temp")])
    if args.offset is not None:
        cmd.extend(["--offset", str(args.offset)])
    if args.preview_frame is not None:
        cmd.extend(["--preview-frame", str(args.preview_frame)])
        if args.preview_image:
            cmd.extend(["--preview-image", str(Path(args.preview_image).resolve())])
    if args.preview_clip is not None:
        cmd.extend(["--preview-clip", str(args.preview_clip)])
    if getattr(args, "preview_dense", False):
        cmd.append("--preview-dense")
    run(cmd, error_hint="渲染失败，请检查视频文件、FFmpeg 和字体路径是否正确")

    if args.preview_frame is not None:
        log("\n[OK] 预览图已生成。")
        return

    rendered_output = (workdir / "temp" / (video.stem + "_chat.mp4")) if workdir else output_default
    if final_output != rendered_output:
        publish_output(
            rendered_output,
            final_output,
            backup_prev=not bool(getattr(args, "no_backup_prev", False)),
        )
        log(f"\n[OK] 已输出到: {final_output}")
    else:
        log(f"\n[OK] 输出: {final_output}")
    if (
        args.preview_frame is None
        and args.preview_clip is None
        and str(getattr(args, "mode", "auto") or "auto") not in ("preview",)
    ):
        log("提示: 下次可先 --preview-clip 10 或 --mode preview 确认 offset/布局，再出长片")


def _terminal_exit_code(value: object) -> int:
    try:
        return int(value) if value is not None else 0
    except (TypeError, ValueError):
        return 1


def _write_terminal_result(state: str, returncode: int) -> None:
    terminal_state = str(_TASK_RESULT_CONTEXT.get("terminal_state") or state)
    write_task_result(
        state=terminal_state,
        mode=str(_TASK_RESULT_CONTEXT.get("mode", "unknown")),
        returncode=returncode,
        artifacts=list(_TASK_RESULT_CONTEXT.get("artifacts", [])),
    )


def main():
    """Run the pipeline and publish an optional narrow terminal result manifest."""
    global _TASK_RESULT_CONTEXT
    _TASK_RESULT_CONTEXT = {"mode": "unknown", "artifacts": []}
    try:
        result = _main()
    except SystemExit as exc:
        code = _terminal_exit_code(exc.code)
        _write_terminal_result("succeeded" if code == 0 else "failed", code)
        raise
    except BaseException:
        _write_terminal_result("failed", 1)
        raise
    code = _terminal_exit_code(result)
    _write_terminal_result("succeeded" if code == 0 else "failed", code)
    return result


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)
        sys.stderr.reconfigure(encoding="utf-8", line_buffering=True)
    except Exception:
        pass
    try:
        main()
    except PipelineError as e:
        print(f"\n{e}", file=sys.stderr)
        sys.exit(1)
